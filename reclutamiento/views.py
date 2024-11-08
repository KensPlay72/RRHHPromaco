from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db import models
from django.contrib.auth.hashers import make_password
from django.contrib.auth.hashers import check_password
from django.utils import timezone
from django.shortcuts import redirect
from django.contrib.auth import logout
import json
import os
import random
import string
from django.conf import settings
from django.http import HttpResponse
from pathlib import Path
import mimetypes
from django.db.models import Q
from django.contrib.sessions.models import Session
import pandas as pd
from dateutil import parser
import uuid
from datetime import datetime, timedelta
from openpyxl import load_workbook
from django.views import View
from django.core.files.storage import FileSystemStorage
from num2words import num2words
from django.utils.formats import number_format
from django.db.models.functions import TruncDate
from django.db.models import Avg
from django.db.models import Count
from datetime import date
import calendar
from calendar import month_name
import locale

from .models import (DepartamentoHonduras,MunicipioHonduras, ContratacionEmpleados,Inventariotelefonos, Telefonia,
                     ContratacionEmpleadosmulti,Cesantias,Spicosmart,PerfilesPuestos,Requisa,Jefes, Colaboradores, HorarioJefes,
                     RegistroAsistencia, Asistencia)

from .models import (Sucursal, Empresas, Unidad_Negocio, Departamento, Ciudades, Puestos, Prioridad, 
                     Modo, Motivo, MedioReclutamiento, TipoContrato, Users, BolsaEmpleos, ControlDeTiempo, ControlDePlazas)

# def contar_sesiones_activas(request):
#     count = Session.objects.count()
#     return JsonResponse({'sesiones_activas': count})

#--------LOGIN VIEW--------#
def login_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            dni = data.get('dni')
            password = data.get('password')

            if not dni or not password:
                return JsonResponse({'error': 'Debe ingresar DNI y contraseña.'}, status=400)

            # Buscar al usuario por su DNI manualmente
            user = Users.objects.filter(dni=dni).first()

            # Verificar si el usuario existe
            if user is None:
                return JsonResponse({'error': 'DNI No encontrado'}, status=404)

            # Verificar el estado del usuario antes de la autenticación
            if user.estado.lower() != 'activo':
                return JsonResponse({'error': 'Usuario inactivo.'}, status=403)

            # Verificar la contraseña usando check_password
            if not check_password(password, user.password):
                return JsonResponse({'error': 'Contraseña incorrecta.'}, status=401)

            # Comprobar si la contraseña es la predeterminada
            if password == '12345678':
                return JsonResponse({'showModal': True}, status=200)

            # Guardar la información del usuario en la sesión manualmente
            request.session['user_id'] = user.id
            request.session['dni'] = user.dni
            request.session['username'] = user.username

            # Guardar los permisos principales
            request.session['plazas'] = user.plazas
            request.session['users'] = user.users
            request.session['bolsaempleo'] = user.bolsaempleo
            request.session['contrataciones'] = user.contrataciones
            request.session['cesantias'] = user.cesantias
            request.session['perfilpuesto'] = user.perfilpuesto
            request.session['listas'] = user.listas
            request.session['requisa'] = user.requisa
            request.session['inventario'] = user.inventario

            # Guardar los permisos de "ver" y "escribir" para cada sección
            request.session['plazas_ver'] = user.plazas_ver
            request.session['plazas_escribir'] = user.plazas_escribir
            request.session['bolsaempleo_ver'] = user.bolsaempleo_ver
            request.session['bolsaempleo_escribir'] = user.bolsaempleo_escribir
            request.session['contrataciones_ver'] = user.contrataciones_ver
            request.session['contrataciones_escribir'] = user.contrataciones_escribir
            request.session['listas_ver'] = user.listas_ver
            request.session['listas_escribir'] = user.listas_escribir
            request.session['contrataciones_multi_ver'] = user.contrataciones_multi_ver
            request.session['contrataciones_multi_escribir'] = user.contrataciones_multi_escribir
            request.session['cesantias_ver'] = user.cesantias_ver
            request.session['cesantias_escribir'] = user.cesantias_escribir
            request.session['perfilpuesto_ver'] = user.perfilpuesto_ver
            request.session['perfilpuesto_escribir'] = user.perfilpuesto_escribir
            request.session['requisa_ver'] = user.requisa_ver
            request.session['requisa_escribir'] = user.requisa_escribir
            request.session['inventario_ver'] = user.inventario_ver
            request.session['inventario_escribir'] = user.inventario_escribir

            # Establecer permisos adicionales (con nombres más descriptivos)
            request.session['has_plazas_permission'] = user.plazas
            request.session['has_users_permission'] = user.users
            request.session['has_roles_permission'] = user.roles
            request.session['has_bolsa_empleo_permission'] = user.bolsaempleo
            request.session['has_contrataciones_permission'] = user.contrataciones
            request.session['has_listas_permission'] = user.listas
            request.session['has_contrataciones_multi_permission'] = user.contrataciones_multi
            request.session['has_cesantias_permission'] = user.cesantias
            request.session['has_perfilpuesto_permission'] = user.perfilpuesto
            request.session['has_requisa_permission'] = user.requisa
            request.session['has_inventario_permission'] = user.inventario

            # Establecer la sesión como iniciada manualmente
            request.session['user_authenticated'] = True

            # Actualizar el último inicio de sesión
            user.last_login = timezone.now()
            user.fechaactualizacion = timezone.now()
            user.save()

            return JsonResponse({'redirect': '/Dashboard', 'IdUsuario': user.id}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            # Imprimir el error para depurar
            print(f'Error en login: {e}')
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'login.html')

@csrf_exempt
def change_password_view(request):
    if request.method == 'POST':
        try:
            print(f'Request body: {request.body}')
            data = json.loads(request.body)
            dni = data.get('dni')
            new_password = data.get('contrasenaeditar')
            confirm_password = data.get('new_password_confirmation')

            # Imprimir los valores recibidos para depuración
            print(f'Datos recibidos - DNI: {dni}, Nueva Contraseña: {new_password}, Confirmación: {confirm_password}')

            # Validaciones...
            if not dni or not new_password or not confirm_password:
                return JsonResponse({'error': 'Todos los campos son obligatorios.'}, status=400)

            if new_password != confirm_password:
                return JsonResponse({'error': 'Las contraseñas no coinciden.'}, status=400)

            if len(new_password) < 8:
                return JsonResponse({'error': 'La nueva contraseña debe tener al menos 8 caracteres.'}, status=400)

            if new_password == '12345678':
                return JsonResponse({'error': 'La nueva contraseña no puede ser 12345678.'}, status=400)

            # Buscar al usuario por DNI
            user = get_object_or_404(Users, dni=dni)
            print(f'Usuario encontrado: {user}')

            # Actualizar la contraseña
            user.password = make_password(new_password)
            user.save()
            print(f'Contraseña actualizada para el usuario {user.username}')

            return JsonResponse({'message': 'Contraseña actualizada correctamente.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Formato de JSON no válido.'}, status=400)
        except Users.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado.'}, status=404)
        except Exception as e:
            print(f'Error en la vista: {e}')  # Depuración
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

def logout_view(request):
    logout(request)
    return redirect('login')

#--------DASHBOARD VIEW--------#
def dashboard_view(request):
    # Filtros
    empresa = request.GET.get('empresa')
    año = request.GET.get('año')
    mes_corte = request.GET.get('mes_corte')
    sucursal = request.GET.get('sucursal')

    # Base QuerySet
    plazas = ControlDePlazas.objects.all()

    # Aplicar filtros si existen
    if empresa:
        plazas = plazas.filter(empresa__nombre_empresa=empresa)
    if año:
        plazas = plazas.filter(año=año)
    if mes_corte:
        plazas = plazas.filter(mes_corte=mes_corte)
    if sucursal:
        plazas = plazas.filter(sucursal__nombre_sucursal=sucursal)

    # Conteos después de aplicar los filtros
    canceladas_count = plazas.filter(estatus='CANCELADA').count()
    en_proceso_count = plazas.filter(estatus='EN PROCESO').count()
    cerradas_count = plazas.filter(estatus='CERRADA').count()

    # Promedio de tiempo_efectivo_cobertura
    promedio_tiempo_efectivo_cobertura = plazas.filter(estatus='CERRADA').aggregate(Avg('tiempo_efectivo_cobertura'))['tiempo_efectivo_cobertura__avg']

    # Conteo de plazas cerradas por género
    genero_count = plazas.filter(estatus='CERRADA').aggregate(
        total_masculino=Count('id', filter=Q(genero='M')),
        total_femenino=Count('id', filter=Q(genero='F'))
    )

    # Promedio de edades
    promedio_edad = plazas.aggregate(Avg('edad'))['edad__avg']

    # Obtener empresas, años, meses de corte, y sucursales únicos para los selectores
    empresas = ControlDePlazas.objects.values('empresa__nombre_empresa').distinct()
    años = ControlDePlazas.objects.values('año').distinct()
    meses_corte = ControlDePlazas.objects.values('mes_corte').distinct()
    sucursales = ControlDePlazas.objects.values('sucursal__nombre_sucursal').distinct()

    # Consultar el estado de pago en cesantías
    cesantias_count = Cesantias.objects.values('estado_pago').annotate(total=Count('id')).filter(estado_pago__in=['AUTORIZADO', 'CALCULADO', 'PAGADO'])

    # Crear un diccionario para los resultados de cesantías
    estado_pago_counts = {estado['estado_pago']: estado['total'] for estado in cesantias_count}

    # Nuevo conteo para BolsaEmpleos basado en el estado
    bolsa_estado_count = BolsaEmpleos.objects.values('estado').annotate(total=Count('id')).filter(estado__in=['CONTACTADO', 'DESCARTADO', 'ENTREVISTADO', 'REGISTRADO'])
    
    # Crear un diccionario para los resultados de BolsaEmpleos
    bolsa_estado_counts = {estado['estado']: estado['total'] for estado in bolsa_estado_count}

    # Obtener la fecha actual
    fecha_actual = date.today()

    # Pasar los datos al contexto
    context = {
        'canceladas_count': canceladas_count,
        'en_proceso_count': en_proceso_count,
        'cerradas_count': cerradas_count,
        'empresas': empresas,
        'años': años,
        'meses_corte': meses_corte,
        'sucursales': sucursales,
        'promedio_tiempo_efectivo_cobertura': promedio_tiempo_efectivo_cobertura,
        'total_masculino': genero_count['total_masculino'],
        'total_femenino': genero_count['total_femenino'],
        'promedio_edad': promedio_edad,
        'fecha_actual': fecha_actual,
        'estado_pago_labels': list(estado_pago_counts.keys()),  # Etiquetas para el gráfico de cesantías
        'estado_pago_values': list(estado_pago_counts.values()),  # Valores para el gráfico de cesantías
        'bolsa_estado_labels': list(bolsa_estado_counts.keys()),  # Etiquetas para el nuevo gráfico de BolsaEmpleos
        'bolsa_estado_values': list(bolsa_estado_counts.values()),  # Valores para el nuevo gráfico de BolsaEmpleos
    }
    
    return render(request, 'dashboard.html', context)


#--------SIN ACCESO VIEW--------#
def sin_acceso_view(request):
    return render(request, 'errors/sinacceso.html')
    
#--------SUCURSALES VIEW--------#
@csrf_exempt
def sucursales_view(request, id=None):
    if request.method == 'GET':
        # Manejando la vista GET para mostrar y buscar sucursales
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  # Obtener el estado del query param
        sucursales = Sucursal.objects.all()

        if search:
            sucursales = sucursales.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_sucursal__icontains=search) 
            )

        if estado:
            sucursales = sucursales.filter(estado__icontains=estado)  # Filtrar por estado

        paginator = Paginator(sucursales, 10)  # Paginación de 10 elementos por página
        page_number = request.GET.get('page')  # Obtener el número de página del query param
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        total_pages = paginator.num_pages
        current_page = page_obj.number

        # Calcular el rango de páginas a mostrar
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste para mostrar siempre 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        context = {
            'sucursales': page_obj,  
            'search': search,  # Pasar el valor de búsqueda actual al contexto
            'estado': estado,  # Pasar el valor del estado actual al contexto
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }
        return render(request, 'modulos/sucursales.html', context)

    elif request.method == 'POST':
        # Manejando la creación de una nueva sucursal
        try:
            data = json.loads(request.body)
            nombre_sucursal = data.get('nombre_sucursal')
            estado = data.get('estado')  # Obtener el estado del cuerpo de la solicitud

            if not nombre_sucursal:
                return JsonResponse({'success': False, 'message': 'El campo nombre_sucursal es obligatorio.'}, status=400)

            if Sucursal.objects.filter(nombre_sucursal=nombre_sucursal).exists():
                return JsonResponse({'success': False, 'message': 'La sucursal ya está registrada.'}, status=409)

            # Crear la nueva sucursal con el estado especificado
            sucursal = Sucursal.objects.create(
                nombre_sucursal=nombre_sucursal,
                estado=estado  # Asignar el valor del campo estado
            )
            
            return JsonResponse({'success': True, 'message': 'Sucursal creada exitosamente', 'sucursal': {
                'id': sucursal.id,
                'nombre_sucursal': sucursal.nombre_sucursal,
                'estado': sucursal.estado  # Incluir el estado en la respuesta
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        # Manejando la actualización de una sucursal existente
        try:
            sucursal = get_object_or_404(Sucursal, id=id)
            data = json.loads(request.body)
            nombre_sucursal = data.get('nombre_sucursal')
            estado = data.get('estado')  # Obtener el estado del cuerpo de la solicitud
            
            if not nombre_sucursal:
                return JsonResponse({'success': False, 'message': 'El campo nombre_sucursal es obligatorio.'}, status=400)

            if Sucursal.objects.filter(nombre_sucursal=nombre_sucursal).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'La sucursal ya está registrada.'}, status=409)

            sucursal.nombre_sucursal = nombre_sucursal
            sucursal.estado = estado  # Actualizar el estado
            sucursal.save()

            return JsonResponse({'success': True, 'message': 'Sucursal actualizada exitosamente', 'sucursal': {
                'id': sucursal.id,
                'nombre_sucursal': sucursal.nombre_sucursal,
                'estado': sucursal.estado  # Incluir el estado en la respuesta
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        # Método no permitido
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
#--------EMPRESAS VIEW--------#
@csrf_exempt
def empresas_view(request, id=None):
    if request.method == 'GET':
        # Manejando la vista GET para mostrar y buscar empresas
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  # Obtener el estado del query param
        empresas = Empresas.objects.all()

        if search:
            empresas = empresas.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_empresa__icontains=search)
            )

        if estado:
            empresas = empresas.filter(estado__icontains=estado)  # Filtrar por estado

        paginator = Paginator(empresas, 10)  # Paginación de 10 elementos por página
        page_number = request.GET.get('page')  # Obtener el número de página del query param
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste para mostrar siempre 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        context = {
            'empresas': page_obj,  
            'search': search, 
            'estado': estado,  
            'page_range': page_range, 
        }
        return render(request, 'modulos/empresas.html', context)

    elif request.method == 'POST':
        # Manejando la creación de una nueva empresa
        try:
            data = json.loads(request.body)
            nombre_empresa = data.get('nombre_empresa')
            estado = data.get('estado') 

            if not nombre_empresa:
                return JsonResponse({'success': False, 'message': 'El campo nombre_empresa es obligatorio.'}, status=400)

            if Empresas.objects.filter(nombre_empresa=nombre_empresa).exists():
                return JsonResponse({'success': False, 'message': 'La empresa ya está registrada.'}, status=409)

            # Crear la nueva empresa con el estado especificado
            empresa = Empresas.objects.create(
                nombre_empresa=nombre_empresa,
                estado=estado  # Asignar el valor del campo estado
            )
            
            return JsonResponse({'success': True, 'message': 'Empresa creada exitosamente', 'empresa': {
                'id': empresa.id,
                'nombre_empresa': empresa.nombre_empresa,
                'estado': empresa.estado  # Incluir el estado en la respuesta
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        # Manejando la actualización de una empresa existente
        try:
            empresa = get_object_or_404(Empresas, id=id)
            data = json.loads(request.body)
            nombre_empresa = data.get('nombre_empresa')
            estado = data.get('estado')  
            
            if not nombre_empresa:
                return JsonResponse({'success': False, 'message': 'El campo nombre_empresa es obligatorio.'}, status=400)

            if Empresas.objects.filter(nombre_empresa=nombre_empresa).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'La empresa ya está registrada.'}, status=409)

            empresa.nombre_empresa = nombre_empresa
            empresa.estado = estado  # Actualizar el estado
            empresa.save()

            return JsonResponse({'success': True, 'message': 'Empresa actualizada exitosamente', 'empresa': {
                'id': empresa.id,
                'nombre_empresa': empresa.nombre_empresa,
                'estado': empresa.estado  # Incluir el estado en la respuesta
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    else:
        # Método no permitido
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405) 
#--------UNIDAD DE NEGOCIO VIEW--------#
@csrf_exempt
def unidad_negocio_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  # Obtener el estado del query param
        unidades_negocio = Unidad_Negocio.objects.all()

        if search:
            unidades_negocio = unidades_negocio.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_unidad_de_negocio__icontains=search)
            )

        if estado:
            unidades_negocio = unidades_negocio.filter(estado__icontains=estado)  # Filtrar por estado

        paginator = Paginator(unidades_negocio, 10)  # Paginación de 10 elementos por página
        page_number = request.GET.get('page')  # Obtener el número de página del query param
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste para mostrar siempre 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        context = {
            'unidades_negocio': page_obj,  
            'search': search,  
            'estado': estado,
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }
        return render(request, 'modulos/unidad_negocio.html', context)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            nombre_unidad_de_negocio = data.get('nombre_unidad_negocio')
            estado = data.get('estado')  

            if not nombre_unidad_de_negocio:
                return JsonResponse({'success': False, 'message': 'El campo nombre_unidad_negocio es obligatorio.'}, status=400)

            if Unidad_Negocio.objects.filter(nombre_unidad_de_negocio=nombre_unidad_de_negocio).exists():
                return JsonResponse({'success': False, 'message': 'La unidad de negocio ya está registrada.'}, status=409)

            unidad_negocio = Unidad_Negocio.objects.create(
                nombre_unidad_de_negocio=nombre_unidad_de_negocio,
                estado=estado  
            )
            
            return JsonResponse({'success': True, 'message': 'Unidad de negocio creada exitosamente', 'unidad_negocio': {
                'id': unidad_negocio.id,
                'nombre_unidad_negocio': unidad_negocio.nombre_unidad_de_negocio,
                'estado': unidad_negocio.estado  
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        try:
            unidad_negocio = get_object_or_404(Unidad_Negocio, id=id)
            data = json.loads(request.body)
            nombre_unidad_de_negocio = data.get('nombre_unidad_negocio')
            estado = data.get('estado')  

            if not nombre_unidad_de_negocio:
                return JsonResponse({'success': False, 'message': 'El campo nombre_unidad_negocio es obligatorio.'}, status=400)

            if Unidad_Negocio.objects.filter(nombre_unidad_de_negocio=nombre_unidad_de_negocio).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'La unidad de negocio ya está registrada.'}, status=409)

            unidad_negocio.nombre_unidad_de_negocio = nombre_unidad_de_negocio
            unidad_negocio.estado = estado  
            unidad_negocio.save()

            return JsonResponse({'success': True, 'message': 'Unidad de negocio actualizada exitosamente', 'unidad_negocio': {
                'id': unidad_negocio.id,
                'nombre_unidad_negocio': unidad_negocio.nombre_unidad_de_negocio,
                'estado': unidad_negocio.estado  
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
#--------UNIDAD DE NEGOCIO VIEW--------#
@csrf_exempt
def departamentos_view(request, id=None):
    if request.method == 'GET':
        # Manejando la vista GET para mostrar y buscar departamentos
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  
        departamentos = Departamento.objects.all()

        if search:
            departamentos = departamentos.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_departamento__icontains=search)
            )

        if estado:
            departamentos = departamentos.filter(estado__icontains=estado)  

        paginator = Paginator(departamentos, 10)  # Paginación de 10 elementos por página
        page_number = request.GET.get('page')  
        page_obj = paginator.get_page(page_number)  

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste para mostrar siempre 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        context = {
            'departamentos': page_obj,  
            'search': search,  
            'estado': estado,
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }
        return render(request, 'modulos/departamentos.html', context)

    elif request.method == 'POST':
        # Manejando la creación de un nuevo departamento
        try:
            data = json.loads(request.body)
            nombre_departamento = data.get('nombre_departamento')
            estado = data.get('estado')  

            if not nombre_departamento:
                return JsonResponse({'success': False, 'message': 'El campo nombre_departamento es obligatorio.'}, status=400)

            if Departamento.objects.filter(nombre_departamento=nombre_departamento).exists():
                return JsonResponse({'success': False, 'message': 'El departamento ya está registrado.'}, status=409)

            # Crear el nuevo departamento con el estado especificado
            departamento = Departamento.objects.create(
                nombre_departamento=nombre_departamento,
                estado=estado  
            )
            
            return JsonResponse({'success': True, 'message': 'Departamento creado exitosamente', 'departamento': {
                'id': departamento.id,
                'nombre_departamento': departamento.nombre_departamento,
                'estado': departamento.estado  
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        try:
            # Manejando la actualización de un departamento existente
            departamento = get_object_or_404(Departamento, id=id)
            data = json.loads(request.body)
            nombre_departamento = data.get('nombre_departamento')
            estado = data.get('estado')

            if not nombre_departamento:
                return JsonResponse({'success': False, 'message': 'El campo nombre_departamento es obligatorio.'}, status=400)

            if Departamento.objects.filter(nombre_departamento=nombre_departamento).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'El departamento ya está registrado.'}, status=409)

            departamento.nombre_departamento = nombre_departamento
            departamento.estado = estado  # Actualizar el estado
            departamento.save()

            return JsonResponse({'success': True, 'message': 'Departamento actualizado exitosamente', 'departamento': {
                'id': departamento.id,
                'nombre_departamento': departamento.nombre_departamento,
                'estado': departamento.estado  # Incluir el estado en la respuesta
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
#--------CIUDADES VIEW--------#
@csrf_exempt
def ciudades_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')  # Capturar el valor del parámetro de búsqueda
        estado = request.GET.get('estado', '')  # Capturar el valor del parámetro de estado
        ciudades = Ciudades.objects.all()

        # Filtrar según búsqueda y estado
        if search:
            ciudades = ciudades.filter(
                models.Q(id__icontains=search) |
                models.Q(nombre_ciudades__icontains=search)
            )

        if estado:
            ciudades = ciudades.filter(estado__icontains=estado)

        # Paginación
        paginator = Paginator(ciudades, 10)  # Paginación de 10 elementos por página
        page_number = request.GET.get('page')  # Capturar el número de página
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajustar el rango para mostrar siempre un máximo de 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        context = {
            'ciudades': page_obj,
            'search': search,
            'estado': estado,
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }
        return render(request, 'modulos/ciudades.html', context)

    elif request.method == 'POST':
        # Manejando la creación de una nueva ciudad
        try:
            data = json.loads(request.body)
            nombre_ciudades = data.get('nombre_ciudades')
            estado = data.get('estado')

            if not nombre_ciudades:
                return JsonResponse({'success': False, 'message': 'El campo nombre_ciudades es obligatorio.'}, status=400)

            if Ciudades.objects.filter(nombre_ciudades=nombre_ciudades).exists():
                return JsonResponse({'success': False, 'message': 'La ciudad ya está registrada.'}, status=409)

            # Crear la nueva ciudad con el estado especificado
            ciudad = Ciudades.objects.create(
                nombre_ciudades=nombre_ciudades,
                estado=estado
            )

            return JsonResponse({'success': True, 'message': 'Ciudad creada exitosamente', 'ciudad': {
                'id': ciudad.id,
                'nombre_ciudades': ciudad.nombre_ciudades,
                'estado': ciudad.estado
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        try:
            # Manejando la actualización de una ciudad existente
            ciudad = get_object_or_404(Ciudades, id=id)
            data = json.loads(request.body)
            nombre_ciudades = data.get('nombre_ciudades')
            estado = data.get('estado')

            if not nombre_ciudades:
                return JsonResponse({'success': False, 'message': 'El campo nombre_ciudades es obligatorio.'}, status=400)

            if Ciudades.objects.filter(nombre_ciudades=nombre_ciudades).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'La ciudad ya está registrada.'}, status=409)

            ciudad.nombre_ciudades = nombre_ciudades
            ciudad.estado = estado  # Actualizar el estado
            ciudad.save()

            return JsonResponse({'success': True, 'message': 'Ciudad actualizada exitosamente', 'ciudad': {
                'id': ciudad.id,
                'nombre_ciudades': ciudad.nombre_ciudades,
                'estado': ciudad.estado  # Incluir el estado en la respuesta
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
#--------PUESTOS VIEW--------#
@csrf_exempt
def puestos_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  
        puestos = Puestos.objects.all()

        if search:
            puestos = puestos.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_puestos__icontains=search) 
            )

        if estado:
            puestos = puestos.filter(estado__icontains=estado)  

        paginator = Paginator(puestos, 10)  # Número de elementos por página
        page_number = request.GET.get('page')  
        page_obj = paginator.get_page(page_number)  

        total_pages = paginator.num_pages
        current_page = page_obj.number

        # Calcular el rango de páginas a mostrar
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste para mostrar siempre 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = list(range(start_page, end_page + 1))

        context = {
            'puestos': page_obj,  
            'search': search,  
            'estado': estado,
            'page_range': page_range,  # Pasar el rango de páginas como lista
        }
        return render(request, 'modulos/puestos.html', context)

    elif request.method == 'POST':
        # Manejando la creación de un nuevo puesto
        try:
            data = json.loads(request.body)
            nombre_puestos = data.get('nombre_puestos')
            estado = data.get('estado')  

            if not nombre_puestos:
                return JsonResponse({'success': False, 'message': 'El campo nombre_puestos es obligatorio.'}, status=400)

            if Puestos.objects.filter(nombre_puestos=nombre_puestos).exists():
                return JsonResponse({'success': False, 'message': 'El puesto ya está registrado.'}, status=409)

            # Crear el nuevo puesto con el estado especificado
            puesto = Puestos.objects.create(
                nombre_puestos=nombre_puestos,
                estado=estado  
            )
            
            return JsonResponse({'success': True, 'message': 'Puesto creado exitosamente', 'puesto': {
                'id': puesto.id,
                'nombre_puestos': puesto.nombre_puestos,
                'estado': puesto.estado  
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        try:
            puesto = get_object_or_404(Puestos, id=id)
            data = json.loads(request.body)
            nombre_puestos = data.get('nombre_puestos')
            estado = data.get('estado')

            if not nombre_puestos:
                return JsonResponse({'success': False, 'message': 'El campo nombre_puestos es obligatorio.'}, status=400)

            if Puestos.objects.filter(nombre_puestos=nombre_puestos).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'El puesto ya está registrado.'}, status=409)

            puesto.nombre_puestos = nombre_puestos
            puesto.estado = estado  # Actualizar el estado
            puesto.save()

            return JsonResponse({'success': True, 'message': 'Puesto actualizado exitosamente', 'puesto': {
                'id': puesto.id,
                'nombre_puestos': puesto.nombre_puestos,
                'estado': puesto.estado  
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
#--------PRIORIDADES VIEW--------#   
@csrf_exempt
def prioridades_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')
        estado = request.GET.get('estado', '')
        prioridades = Prioridad.objects.all()

        if search:
            prioridades = prioridades.filter(
                models.Q(id__icontains=search) |
                models.Q(nombre_prioridad__icontains=search)
            )

        if estado:
            prioridades = prioridades.filter(estado__icontains=estado)

        paginator = Paginator(prioridades, 10)  # Paginación de 10 elementos por página
        page_number = request.GET.get('page')  # Obtener el número de página del query param
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste para mostrar siempre un máximo de 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear el rango de páginas a mostrar

        context = {
            'prioridades': page_obj,
            'search': search,
            'estado': estado,
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }
        return render(request, 'modulos/prioridades.html', context)

    elif request.method == 'POST':
        # Manejando la creación de una nueva prioridad
        try:
            data = json.loads(request.body)
            nombre_prioridad = data.get('nombre_prioridad')
            estado = data.get('estado')  # Obtener el estado del cuerpo de la solicitud

            if not nombre_prioridad:
                return JsonResponse({'success': False, 'message': 'El campo nombre_prioridad es obligatorio.'}, status=400)

            if Prioridad.objects.filter(nombre_prioridad=nombre_prioridad).exists():
                return JsonResponse({'success': False, 'message': 'La prioridad ya está registrada.'}, status=409)

            # Crear la nueva prioridad con el estado especificado
            prioridad = Prioridad.objects.create(
                nombre_prioridad=nombre_prioridad,
                estado=estado  # Asignar el valor del campo estado
            )

            return JsonResponse({'success': True, 'message': 'Prioridad creada exitosamente', 'prioridad': {
                'id': prioridad.id,
                'nombre_prioridad': prioridad.nombre_prioridad,
                'estado': prioridad.estado  # Incluir el estado en la respuesta
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        # Manejando la actualización de una prioridad existente
        try:
            prioridad = get_object_or_404(Prioridad, id=id)
            data = json.loads(request.body)
            nombre_prioridad = data.get('nombre_prioridad')
            estado = data.get('estado')  # Obtener el estado del cuerpo de la solicitud

            if not nombre_prioridad:
                return JsonResponse({'success': False, 'message': 'El campo nombre_prioridad es obligatorio.'}, status=400)

            if Prioridad.objects.filter(nombre_prioridad=nombre_prioridad).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'La prioridad ya está registrada.'}, status=409)

            prioridad.nombre_prioridad = nombre_prioridad
            prioridad.estado = estado  # Actualizar el estado
            prioridad.save()

            return JsonResponse({'success': True, 'message': 'Prioridad actualizada exitosamente', 'prioridad': {
                'id': prioridad.id,
                'nombre_prioridad': prioridad.nombre_prioridad,
                'estado': prioridad.estado  # Incluir el estado en la respuesta
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    else:
        # Método no permitido
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)  
#--------MODO VIEW--------#   
@csrf_exempt
def modos_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')
        estado = request.GET.get('estado', '')
        modos = Modo.objects.all()

        if search:
            modos = modos.filter(
                models.Q(id__icontains=search) |
                models.Q(nombre_modo__icontains=search)
            )

        if estado:
            modos = modos.filter(estado__icontains=estado)

        paginator = Paginator(modos, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'modos': page_obj,
            'search': search,
            'estado': estado,
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }
        return render(request, 'modulos/modos.html', context)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            nombre_modo = data.get('nombre_modo')
            estado = data.get('estado')

            if not nombre_modo:
                return JsonResponse({'success': False, 'message': 'El campo nombre_modo es obligatorio.'}, status=400)

            if Modo.objects.filter(nombre_modo=nombre_modo).exists():
                return JsonResponse({'success': False, 'message': 'El modo ya está registrado.'}, status=409)

            modo = Modo.objects.create(
                nombre_modo=nombre_modo,
                estado=estado
            )

            return JsonResponse({'success': True, 'message': 'Modo creado exitosamente', 'modo': {
                'id': modo.id,
                'nombre_modo': modo.nombre_modo,
                'estado': modo.estado
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        try:
            modo = get_object_or_404(Modo, id=id)
            data = json.loads(request.body)
            nombre_modo = data.get('nombre_modo')
            estado = data.get('estado')

            if not nombre_modo:
                return JsonResponse({'success': False, 'message': 'El campo nombre_modo es obligatorio.'}, status=400)

            if Modo.objects.filter(nombre_modo=nombre_modo).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'El modo ya está registrado.'}, status=409)

            modo.nombre_modo = nombre_modo
            modo.estado = estado
            modo.save()

            return JsonResponse({'success': True, 'message': 'Modo actualizado exitosamente', 'modo': {
                'id': modo.id,
                'nombre_modo': modo.nombre_modo,
                'estado': modo.estado
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
#--------MOTIVO VIEW--------#       
@csrf_exempt
def motivos_view(request, id=None):
    if request.method == 'GET':
    # Obtener parámetros de búsqueda y estado desde la URL
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  
    
    # Obtener todos los registros de motivos
        motivos = Motivo.objects.all()

    # Aplicar filtro de búsqueda si se proporcionó
        if search:
            motivos = motivos.filter(
                models.Q(id__icontains=search) |
                models.Q(nombre_motivo__icontains=search)
            )

    # Aplicar filtro de estado si se proporcionó
        if estado:
            motivos = motivos.filter(estado__icontains=estado)

    # Paginación: 10 registros por página
        paginator = Paginator(motivos, 10)  
        page_number = request.GET.get('page')  # Obtener el número de página desde la URL
        page_obj = paginator.get_page(page_number)  # Obtener la página actual

    # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

    # Ajustar los límites del rango de páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Rango de páginas a mostrar

    # Preparar contexto para la plantilla
        context = {
            'motivos': page_obj,  # Pasar los registros paginados al contexto
            'search': search,  # Pasar el término de búsqueda al contexto
            'estado': estado,  # Pasar el estado al contexto
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }

    # Renderizar la plantilla con el contexto
        return render(request, 'modulos/motivos.html', context)
    elif request.method == 'POST':
        # Manejando la creación de un nuevo motivo
        try:
            data = json.loads(request.body)
            nombre_motivo = data.get('nombre_motivo')
            estado = data.get('estado')

            if not nombre_motivo:
                return JsonResponse({'success': False, 'message': 'El campo nombre_motivo es obligatorio.'}, status=400)

            if Motivo.objects.filter(nombre_motivo=nombre_motivo).exists():
                return JsonResponse({'success': False, 'message': 'El motivo ya está registrado.'}, status=409)

            motivo = Motivo.objects.create(
                nombre_motivo=nombre_motivo,
                estado=estado
            )
            
            return JsonResponse({'success': True, 'message': 'Motivo creado exitosamente', 'motivo': {
                'id': motivo.id,
                'nombre_motivo': motivo.nombre_motivo,
                'estado': motivo.estado
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        # Manejando la actualización de un motivo existente
        try:
            motivo = get_object_or_404(Motivo, id=id)
            data = json.loads(request.body)
            nombre_motivo = data.get('nombre_motivo')
            estado = data.get('estado')

            if not nombre_motivo:
                return JsonResponse({'success': False, 'message': 'El campo nombre_motivo es obligatorio.'}, status=400)

            if Motivo.objects.filter(nombre_motivo=nombre_motivo).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'El motivo ya está registrado.'}, status=409)

            motivo.nombre_motivo = nombre_motivo
            motivo.estado = estado
            motivo.save()

            return JsonResponse({'success': True, 'message': 'Motivo actualizado exitosamente', 'motivo': {
                'id': motivo.id,
                'nombre_motivo': motivo.nombre_motivo,
                'estado': motivo.estado
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
#--------MEDIO RECLUTAMIENTO VIEW--------#
@csrf_exempt
def medios_reclutamiento_view(request, id=None):
    if request.method == 'GET':
    # Obtener los parámetros de búsqueda y estado desde la URL
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  
    
    # Obtener todos los registros de medios de reclutamiento
        medios_reclutamiento = MedioReclutamiento.objects.all()

    # Filtrar por búsqueda si se proporciona
        if search:
            medios_reclutamiento = medios_reclutamiento.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_medio_de_reclutamiento__icontains=search)
            )

    # Filtrar por estado si se proporciona
        if estado:
            medios_reclutamiento = medios_reclutamiento.filter(estado__icontains=estado)

    # Paginación de 10 registros por página
        paginator = Paginator(medios_reclutamiento, 10)  
        page_number = request.GET.get('page')  # Obtener el número de página desde la URL
        page_obj = paginator.get_page(page_number)  # Obtener la página actual

    # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

    # Ajustar los límites del rango de páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Rango de páginas a mostrar

    # Preparar el contexto para la plantilla
        context = {
            'medios_reclutamiento': page_obj,
            'search': search,
            'estado': estado,
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }

    # Renderizar la plantilla con el contexto
        return render(request, 'modulos/medios_reclutamiento.html', context)
    elif request.method == 'POST':
        # Manejando la creación de un nuevo medio de reclutamiento
        try:
            data = json.loads(request.body)
            nombre_medio_de_reclutamiento = data.get('nombre_medio_de_reclutamiento')
            estado = data.get('estado')

            if not nombre_medio_de_reclutamiento:
                return JsonResponse({'success': False, 'message': 'El campo nombre_medio_de_reclutamiento es obligatorio.'}, status=400)

            if MedioReclutamiento.objects.filter(nombre_medio_de_reclutamiento=nombre_medio_de_reclutamiento).exists():
                return JsonResponse({'success': False, 'message': 'El medio de reclutamiento ya está registrado.'}, status=409)

            medio = MedioReclutamiento.objects.create(
                nombre_medio_de_reclutamiento=nombre_medio_de_reclutamiento,
                estado=estado
            )
            
            return JsonResponse({'success': True, 'message': 'Medio de Reclutamiento creado exitosamente', 'medio': {
                'id': medio.id,
                'nombre_medio_de_reclutamiento': medio.nombre_medio_de_reclutamiento,
                'estado': medio.estado
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        # Manejando la actualización de un medio de reclutamiento existente
        try:
            medio = get_object_or_404(MedioReclutamiento, id=id)
            data = json.loads(request.body)
            nombre_medio_de_reclutamiento = data.get('nombre_medio_de_reclutamiento')
            estado = data.get('estado')

            if not nombre_medio_de_reclutamiento:
                return JsonResponse({'success': False, 'message': 'El campo nombre_medio_de_reclutamiento es obligatorio.'}, status=400)

            if MedioReclutamiento.objects.filter(nombre_medio_de_reclutamiento=nombre_medio_de_reclutamiento).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'El medio de reclutamiento ya está registrado.'}, status=409)

            medio.nombre_medio_de_reclutamiento = nombre_medio_de_reclutamiento
            medio.estado = estado
            medio.save()

            return JsonResponse({'success': True, 'message': 'Medio de Reclutamiento actualizado exitosamente', 'medio': {
                'id': medio.id,
                'nombre_medio_de_reclutamiento': medio.nombre_medio_de_reclutamiento,
                'estado': medio.estado
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)       
#--------TIPO CONTRATO VIEW--------#
@csrf_exempt
def tipo_contrato_view(request, id=None):
    if request.method == 'GET':
    # Obtener los parámetros de búsqueda y estado desde la URL
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  
    
    # Obtener todos los registros de tipos de contrato
        tipos_contrato = TipoContrato.objects.all()

    # Filtrar por búsqueda si se proporciona
        if search:
            tipos_contrato = tipos_contrato.filter(
                models.Q(id__icontains=search) |  
                models.Q(nombre_tipo_de_contrato__icontains=search)
            )

    # Filtrar por estado si se proporciona
        if estado:
            tipos_contrato = tipos_contrato.filter(estado__icontains=estado)

    # Paginación de 10 registros por página
        paginator = Paginator(tipos_contrato, 10)  
        page_number = request.GET.get('page')  # Obtener el número de página desde la URL
        page_obj = paginator.get_page(page_number)  # Obtener la página actual

    # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

    # Ajustar los límites del rango de páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Rango de páginas a mostrar

    # Preparar el contexto para la plantilla
        context = {
            'tipos_contrato': page_obj,  # Pasar los registros paginados al contexto
            'search': search,  # Pasar el término de búsqueda al contexto
            'estado': estado,  # Pasar el estado al contexto
            'page_range': page_range,  # Pasar el rango de páginas al contexto
        }

    # Renderizar la plantilla con el contexto
        return render(request, 'modulos/tipo_contrato.html', context)
    elif request.method == 'POST':
        # Manejando la creación de un nuevo tipo de contrato
        try:
            data = json.loads(request.body)
            nombre_tipo_de_contrato = data.get('nombre_tipo_de_contrato')
            estado = data.get('estado')

            if not nombre_tipo_de_contrato:
                return JsonResponse({'success': False, 'message': 'El campo nombre_tipo_de_contrato es obligatorio.'}, status=400)

            if TipoContrato.objects.filter(nombre_tipo_de_contrato=nombre_tipo_de_contrato).exists():
                return JsonResponse({'success': False, 'message': 'El tipo de contrato ya está registrado.'}, status=409)

            tipo_contrato = TipoContrato.objects.create(
                nombre_tipo_de_contrato=nombre_tipo_de_contrato,
                estado=estado
            )
            
            return JsonResponse({'success': True, 'message': 'Tipo de Contrato creado exitosamente', 'tipo_contrato': {
                'id': tipo_contrato.id,
                'nombre_tipo_de_contrato': tipo_contrato.nombre_tipo_de_contrato,
                'estado': tipo_contrato.estado
            }}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    elif request.method == 'PUT' and id is not None:
        # Manejando la actualización de un tipo de contrato existente
        try:
            tipo_contrato = get_object_or_404(TipoContrato, id=id)
            data = json.loads(request.body)
            nombre_tipo_de_contrato = data.get('nombre_tipo_de_contrato')
            estado = data.get('estado')

            if not nombre_tipo_de_contrato:
                return JsonResponse({'success': False, 'message': 'El campo nombre_tipo_de_contrato es obligatorio.'}, status=400)

            if TipoContrato.objects.filter(nombre_tipo_de_contrato=nombre_tipo_de_contrato).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'El tipo de contrato ya está registrado.'}, status=409)

            tipo_contrato.nombre_tipo_de_contrato = nombre_tipo_de_contrato
            tipo_contrato.estado = estado
            tipo_contrato.save()

            return JsonResponse({'success': True, 'message': 'Tipo de Contrato actualizado exitosamente', 'tipo_contrato': {
                'id': tipo_contrato.id,
                'nombre_tipo_de_contrato': tipo_contrato.nombre_tipo_de_contrato,
                'estado': tipo_contrato.estado
            }}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
    
#--------PISCOSMART VIEW--------#
def pruebapsicosmart_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')
        estado = request.GET.get('estado', '')

        pruebas = Spicosmart.objects.all()

        if search:
            pruebas = pruebas.filter(
                models.Q(nombre_prueba__icontains=search) |
                models.Q(nivel__icontains=search) |
                models.Q(prueba_mide__icontains=search) |
                models.Q(estado__icontains=search) 
            )

        if estado:
            pruebas = pruebas.filter(estado__icontains=estado)

        paginator = Paginator(pruebas, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'pruebas': page_obj,
            'search': search,
            'estado': estado,
            'page_range': page_range,
        }

        return render(request, 'modulos/psicosmart.html', context)
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            nivel = data.get('nivel')
            nombre_prueba = data.get('nombre_prueba')
            prueba_mide = data.get('prueba_mide')
            estado = data.get('estado')

            if not nombre_prueba:
                return JsonResponse({'success': False, 'message': 'El campo nombre_prueba es obligatorio.'}, status=400)

            nueva_prueba = Spicosmart.objects.create(
                nivel=nivel,
                nombre_prueba=nombre_prueba,
                prueba_mide=prueba_mide,
                estado=estado
            )

            return JsonResponse({
                'success': True, 
                'message': 'Prueba creada exitosamente', 
                'prueba': {
                    'id': nueva_prueba.id,
                    'nivel': nueva_prueba.nivel,
                    'nombre_prueba': nueva_prueba.nombre_prueba,
                    'prueba_mide': nueva_prueba.prueba_mide,
                    'estado': nueva_prueba.estado
                }
            }, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            prueba = Spicosmart.objects.get(id=id)

            prueba.nivel = data.get('nivel', prueba.nivel)
            prueba.nombre_prueba = data.get('nombre_prueba', prueba.nombre_prueba)
            prueba.prueba_mide = data.get('prueba_mide', prueba.prueba_mide)
            prueba.estado = data.get('estado', prueba.estado)

            prueba.save()

            return JsonResponse({'success': True, 'message': 'Prueba actualizada exitosamente'})
        except Spicosmart.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Prueba no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
#--------USERS VIEW--------#
@csrf_exempt
def users_view(request, id=None):
    if request.method == 'GET':
        # Obtener los parámetros de búsqueda y estado desde la URL
        search = request.GET.get('search', '')  
        estado = request.GET.get('estado', '')  
        users = Users.objects.all()

        # Aplicar filtro de búsqueda si se proporciona
        if search:
            users = users.filter(
                models.Q(username__icontains=search) |  
                models.Q(dni__icontains=search)
            )

        # Aplicar filtro por estado si se proporciona
        if estado:
            users = users.filter(estado__icontains=estado) 

        # Paginación de 10 registros por página
        paginator = Paginator(users, 10)  
        page_number = request.GET.get('page')  # Obtener el número de página desde la URL
        page_obj = paginator.get_page(page_number)  # Obtener la página actual

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste para mostrar siempre 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        context = {
            'users': page_obj,  
            'search': search,  
            'estado': estado,
            'start_page': start_page,  
            'end_page': end_page,      
            'total_pages': total_pages,
        }
        return render(request, 'users.html', context)

    elif request.method == 'POST':
        # Crear un nuevo usuario
        try:
            data = json.loads(request.body)
            username = data.get('username')
            dni = data.get('dni')
            password = data.get('password')
            estado = data.get('estado')
            
            # Permisos
            plazas = data.get('plazas', False)
            plazas_ver = data.get('plazas_ver', False)
            plazas_escribir = data.get('plazas_escribir', False)

            users_perm = data.get('users', False)
            roles_perm = data.get('roles', False)

            bolsaempleo = data.get('bolsaempleo', False)
            bolsaempleo_ver = data.get('bolsaempleo_ver', False)
            bolsaempleo_escribir = data.get('bolsaempleo_escribir', False)

            contrataciones = data.get('contrataciones', False)
            contrataciones_ver = data.get('contrataciones_ver', False)
            contrataciones_escribir = data.get('contrataciones_escribir', False)

            listas = data.get('listas', False)
            listas_ver = data.get('listas_ver', False)
            listas_escribir = data.get('listas_escribir', False)

            contrataciones_multi = data.get('contrataciones_multi', False)
            contrataciones_multi_ver = data.get('contrataciones_multi_ver', False)
            contrataciones_multi_escribir = data.get('contrataciones_multi_escribir', False)

            cesantias = data.get('cesantias', False)
            cesantias_ver = data.get('cesantias_ver', False)
            cesantias_escribir = data.get('cesantias_escribir', False)

            # Nuevos permisos para "Perfil de Puesto"
            perfilpuesto = data.get('perfilpuesto', False)
            perfilpuesto_ver = data.get('perfilpuesto_ver', False)
            perfilpuesto_escribir = data.get('perfilpuesto_escribir', False)

            # Nuevos permisos para "Requisas"
            requisas = data.get('requisas', False)
            requisas_ver = data.get('requisas_ver', False)
            requisas_escribir = data.get('requisas_escribir', False)

            # Nuevos permisos para "Inventario"
            inventario = data.get('inventario', False)
            inventario_ver = data.get('inventario_ver', False)
            inventario_escribir = data.get('inventario_escribir', False)

            # Validación de datos
            if not username or not dni or not password:
                return JsonResponse({'success': False, 'message': 'Todos los campos son obligatorios.'}, status=400)

            if Users.objects.filter(dni=dni).exists():
                return JsonResponse({'success': False, 'message': 'El usuario ya está registrado.'}, status=409)

            # Crear el nuevo usuario
            user = Users.objects.create(
                username=username,
                dni=dni,
                password=make_password(password),
                estado=estado,
                plazas=plazas,
                plazas_ver=plazas_ver,
                plazas_escribir=plazas_escribir,
                users=users_perm,
                roles=roles_perm,
                bolsaempleo=bolsaempleo,
                bolsaempleo_ver=bolsaempleo_ver,
                bolsaempleo_escribir=bolsaempleo_escribir,
                contrataciones=contrataciones,
                contrataciones_ver=contrataciones_ver,
                contrataciones_escribir=contrataciones_escribir,
                listas=listas,
                listas_ver=listas_ver,
                listas_escribir=listas_escribir,
                contrataciones_multi=contrataciones_multi,
                contrataciones_multi_ver=contrataciones_multi_ver,
                contrataciones_multi_escribir=contrataciones_multi_escribir,
                cesantias=cesantias,
                cesantias_ver=cesantias_ver,
                cesantias_escribir=cesantias_escribir,
                # Guardar permisos de "Perfil de Puesto"
                perfilpuesto=perfilpuesto,
                perfilpuesto_ver=perfilpuesto_ver,
                perfilpuesto_escribir=perfilpuesto_escribir,
                # Guardar permisos de "Requisas"
                requisa=requisas,
                requisa_ver=requisas_ver,
                requisa_escribir=requisas_escribir,
                # Guardar permisos de "Inventario"
                inventario=inventario,
                inventario_ver=inventario_ver,
                inventario_escribir=inventario_escribir
            )
            
            return JsonResponse({'success': True, 'message': 'Usuario creado exitosamente'}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)


    elif request.method == 'PUT' and id is not None:
        try:
            user = get_object_or_404(Users, id=id)
            data = json.loads(request.body)
            username = data.get('username')
            dni = data.get('dni')
            estado = data.get('estado')

            # Permisos existentes
            plazas = data.get('plazas', False)
            plazas_ver = data.get('plazas_ver', False)
            plazas_escribir = data.get('plazas_escribir', False)
            users_perm = data.get('users', False)
            roles_perm = data.get('roles', False)
            bolsaempleo = data.get('bolsaempleo', False)
            bolsaempleo_ver = data.get('bolsaempleo_ver', False)
            bolsaempleo_escribir = data.get('bolsaempleo_escribir', False)
            contrataciones = data.get('contrataciones', False)
            contrataciones_ver = data.get('contrataciones_ver', False)
            contrataciones_escribir = data.get('contrataciones_escribir', False)
            listas = data.get('listas', False)
            listas_ver = data.get('listas_ver', False)
            listas_escribir = data.get('listas_escribir', False)

            # Nuevos permisos para "Contrataciones Multiservicios"
            contrataciones_multi = data.get('contrataciones_multi', False)
            contrataciones_multi_ver = data.get('contrataciones_multi_ver', False)
            contrataciones_multi_escribir = data.get('contrataciones_multi_escribir', False)

            # Nuevos permisos para "Cesantías"
            cesantias = data.get('cesantias', False)
            cesantias_ver = data.get('cesantias_ver', False)
            cesantias_escribir = data.get('cesantias_escribir', False)

            # Nuevos permisos para "Perfil de Puesto"
            perfilpuesto = data.get('perfilpuesto', False)
            perfilpuesto_ver = data.get('perfilpuesto_ver', False)
            perfilpuesto_escribir = data.get('perfilpuesto_escribir', False)

            # Nuevos permisos para "Requisas"
            requisas = data.get('requisas', False)
            requisas_ver = data.get('requisas_ver', False)
            requisas_escribir = data.get('requisas_escribir', False)

            # Nuevos permisos para "Inventario"
            inventario = data.get('inventario', False)
            inventario_ver = data.get('inventario_ver', False)
            inventario_escribir = data.get('inventario_escribir', False)

            # Validación de datos
            if not username or not dni:
                return JsonResponse({'success': False, 'message': 'Todos los campos son obligatorios.'}, status=400)

            if Users.objects.filter(dni=dni).exclude(id=id).exists():
                return JsonResponse({'success': False, 'message': 'El usuario ya está registrado.'}, status=409)

            # Actualizar solo los datos permitidos
            user.username = username
            user.dni = dni
            user.estado = estado
            user.plazas = plazas
            user.plazas_ver = plazas_ver
            user.plazas_escribir = plazas_escribir
            user.users = users_perm
            user.roles = roles_perm
            user.bolsaempleo = bolsaempleo
            user.bolsaempleo_ver = bolsaempleo_ver
            user.bolsaempleo_escribir = bolsaempleo_escribir
            user.contrataciones = contrataciones
            user.contrataciones_ver = contrataciones_ver
            user.contrataciones_escribir = contrataciones_escribir
            user.listas = listas
            user.listas_ver = listas_ver
            user.listas_escribir = listas_escribir

            # Actualizar los nuevos permisos de "Contrataciones Multiservicios"
            user.contrataciones_multi = contrataciones_multi
            user.contrataciones_multi_ver = contrataciones_multi_ver
            user.contrataciones_multi_escribir = contrataciones_multi_escribir

            # Actualizar los permisos de "Cesantías"
            user.cesantias = cesantias
            user.cesantias_ver = cesantias_ver
            user.cesantias_escribir = cesantias_escribir

            # Actualizar los permisos de "Perfil de Puesto"
            user.perfilpuesto = perfilpuesto
            user.perfilpuesto_ver = perfilpuesto_ver
            user.perfilpuesto_escribir = perfilpuesto_escribir

            # Actualizar los permisos de "Requisas"
            user.requisa = requisas
            user.requisa_ver = requisas_ver
            user.requisa_escribir = requisas_escribir

            # Actualizar los permisos de "Inventario"
            user.inventario = inventario
            user.inventario_ver = inventario_ver
            user.inventario_escribir = inventario_escribir

            # Guardar los cambios
            user.save()

            return JsonResponse({'success': True, 'message': 'Usuario actualizado exitosamente'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)


#--------RESET PASSWORD VIEW--------#    
@csrf_exempt
def reset_password_view(request, id):
    if request.method == 'PUT':
        try:
            # Obtener el usuario por ID
            user = get_object_or_404(Users, id=id)
            
            # Actualizar la contraseña a '12345678' y encriptarla
            user.password = make_password('12345678')
            user.save()

            return JsonResponse({
                'success': True,
                'message': 'Contraseña reestablecida con éxito',
            }, status=200)

        except Exception as e:
            import traceback
            error_message = traceback.format_exc()  # Obtener detalles completos del error
            return JsonResponse({
                'success': False,
                'message': str(e),
                'error': error_message  # Devolver los detalles del error en la respuesta
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': 'Método no permitido',
    }, status=405)

#--------BOLSA DE EMPLEO VIEW--------#    
@csrf_exempt
def bolsaempleo_view(request, id=None):
    if request.method == 'GET':
        # Obtener ciudades y medios de reclutamiento activos
        ciudades = Ciudades.objects.filter(estado='ACTIVO').order_by('nombre_ciudades')
        medios_reclutamiento = MedioReclutamiento.objects.filter(estado='ACTIVO').order_by('nombre_medio_de_reclutamiento')

        # Obtener todos los puestos
        puestos = Puestos.objects.all()

        # Obtener solo los puestos que están en BolsaEmpleos
        puestos_ids = set()
        for empleo in BolsaEmpleos.objects.all():
            puestos_ids.update(empleo.puestosaspira or [])
            puestos_ids.update(empleo.puestosaplica or [])
        
        puestos_en_bolsa = Puestos.objects.filter(id__in=puestos_ids)

        # Obtener los distintos estados únicos presentes en la base de datos
        estados = BolsaEmpleos.objects.values_list('estado', flat=True).distinct()

        # Parámetros de búsqueda
        search = request.GET.get('search', '').upper()  # Convertir a mayúsculas para evitar problemas de case-sensitive
        estado = request.GET.getlist('estado', [])  # Obtener la lista de estados seleccionados
        puestos_aspira = request.GET.getlist('puestosaspira', [])
        puestos_aplica = request.GET.getlist('puestosaplica', [])
        ciudad_id = request.GET.getlist('ciudad', [])  # Filtro de ciudad

        # Obtener la queryset inicial
        bolsa_empleos = BolsaEmpleos.objects.all().prefetch_related('ciudad', 'medio_reclutamiento').order_by('-id')

        # Filtrar por ciudad si se proporciona y no está vacío
        if ciudad_id and all(ciudad_id):  # Asegúrate de que ciudad_id no tenga valores vacíos
            bolsa_empleos = bolsa_empleos.filter(ciudad__id__in=ciudad_id)

        # Filtrar por búsqueda
        if search:
            bolsa_empleos = bolsa_empleos.filter(
                Q(id__icontains=search) |
                Q(dni__icontains=search) |
                Q(nombre_candidato__icontains=search) |
                Q(telefono__icontains=search) |
                Q(telefono2__icontains=search) |
                Q(estado__icontains=search) |
                Q(ciudad__nombre_ciudades__icontains=search) |
                Q(medio_reclutamiento__nombre_medio_de_reclutamiento__icontains=search) |
                Q(edad__icontains=search) |
                Q(fecha_nacimiento__icontains=search) |
                Q(estadocivil__icontains=search) |
                Q(nhijos__icontains=search) |
                Q(direccion__icontains=search) |
                Q(mediomovilizacion__icontains=search) |
                Q(experiencia__icontains=search) |
                Q(observacion__icontains=search) |
                Q(ruta__icontains=search) |
                Q(nombredocumento__icontains=search) |
                Q(puestosaspira__icontains=search) |  # Busca en puestosaspira (parcial)
                Q(puestosaplica__icontains=search)    # Busca en puestosaplica (parcial)
            )

        # Filtrar por estado (múltiples estados seleccionados)
        if estado:
            bolsa_empleos = bolsa_empleos.filter(estado__in=estado)

        # Filtrado manual por puestos que aspira
        if puestos_aspira:
            bolsa_empleos = [
                empleo for empleo in bolsa_empleos
                if any(int(puesto) in (int(p) for p in empleo.puestosaspira or []) for puesto in puestos_aspira)
            ]

        # Filtrado manual por puestos que aplica
        if puestos_aplica:
            bolsa_empleos = [
                empleo for empleo in bolsa_empleos
                if any(int(puesto) in (int(p) for p in (empleo.puestosaplica or [])) for puesto in puestos_aplica)
            ]

        # Paginación: mostrar 10 registros por página
        paginator = Paginator(bolsa_empleos, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar en la paginación
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Convertir IDs de puestosaspira y puestosaplica a nombres de puestos
        for empleo in page_obj:
            empleo.puestosaspira_nombres = [
                Puestos.objects.get(id=puesto_id).nombre_puestos 
                for puesto_id in (empleo.puestosaspira or [])
            ]
            
            empleo.puestosaplica_nombres = [
                Puestos.objects.get(id=puesto_id).nombre_puestos 
                for puesto_id in (empleo.puestosaplica or [])
            ]

        # Contexto para renderizar la plantilla
        context = {
            'puestos': puestos,
            'puestos_en_bolsa': puestos_en_bolsa,
            'ciudades': ciudades,
            'medios_reclutamiento': medios_reclutamiento,
            'bolsa_empleos': page_obj,
            'page_range': page_range,
            'search': search,
            'estado': estado,  # Incluir el filtro de estado en el contexto
            'puestos_aspira': puestos_aspira,
            'puestos_aplica': puestos_aplica,
            'ciudad_id': ciudad_id,
            'estados': estados,  # Incluir los estados disponibles para el filtro
        }
        return render(request, 'bolsaempleo.html', context)
    elif request.method == 'POST':
        try:
            data = json.loads(request.POST.get('data'))  # Datos JSON del formulario
            dni = data.get('dni')
            nombre_candidato = data.get('nombre_candidato')
            puestosaspira = data.get('puestoaspira', [])
            telefono = data.get('telefono')
            telefono2 = data.get('telefono2')
            estado = data.get('estado')
            ciudad_id = data.get('ciudad')
            medio_reclutamiento_id = data.get('mediosReclutamiento')
            edad = data.get('edad') or None  # Asignar None si está vacío
            fecha_nacimiento = data.get('fechanacimiento') or None  # Asignar None si está vacío
            estadocivil = data.get('estadocivil')
            nhijos = data.get('nhijos') or None  # Asignar None si está vacío
            direccion = data.get('direccion')
            experiencia = data.get('experiencia')
            observacion = data.get('observacion')
            mediomovilizacion = data.get('mediomovilizacion')
            archivo = request.FILES.get('cv')  # Archivo PDF subido

            # Validación de campos obligatorios
            if not dni or not nombre_candidato or not telefono or not puestosaspira or not estado:
                return JsonResponse({'success': False, 'message': 'Los campos DNI, Nombre del Candidato, Teléfono, Puesto al que Aspira y Estado son obligatorios.'}, status=400)
            
            # Validar que telefono y telefono2 no sean iguales entre sí
            if telefono == telefono2:
                return JsonResponse({'success': False, 'message': 'El teléfono principal y el secundario no pueden ser iguales.'}, status=400)

            # Validar que telefono y telefono2 no existan ya en la base de datos
            if BolsaEmpleos.objects.filter(telefono=telefono).exists() or BolsaEmpleos.objects.filter(telefono2=telefono).exists():
                return JsonResponse({'success': False, 'message': 'El número de teléfono principal ya está registrado.'}, status=400)

            if telefono2 and (BolsaEmpleos.objects.filter(telefono=telefono2).exists() or BolsaEmpleos.objects.filter(telefono2=telefono2).exists()):
                return JsonResponse({'success': False, 'message': 'El número de teléfono secundario ya está registrado.'}, status=400)
            
            # Validar formato de fecha
            if fecha_nacimiento:
                try:
                    # Intentar convertir a objeto de fecha
                    fecha_nacimiento = timezone.datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'El formato de la fecha de nacimiento es inválido. Debe ser YYYY-MM-DD.'}, status=400)

            # Guardar archivo en ruta especificada con nombre único
            ruta = None
            nombredocumento = None
            if archivo:
                if archivo.content_type != 'application/pdf':
                    return JsonResponse({'success': False, 'message': 'Solo se permiten documentos en formato PDF.'}, status=400)
                
                # Generar nombre único para el archivo con números aleatorios
                random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                nuevo_nombre_archivo = f"{random_str}_{archivo.name}"
                
                # Guardar archivo en reclutamiento/static/document
                ruta_relativa = os.path.join('reclutamiento', 'static', 'document', nuevo_nombre_archivo)
                ruta_archivo = os.path.join(settings.BASE_DIR, ruta_relativa)
                
                # Crear directorio si no existe
                os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)
                
                # Guardar el archivo en la ruta especificada
                with open(ruta_archivo, 'wb+') as destination:
                    for chunk in archivo.chunks():
                        destination.write(chunk)
                
                ruta = ruta_relativa
                nombredocumento = nuevo_nombre_archivo

            # Crear nueva entrada en BolsaEmpleos
            bolsa_empleo = BolsaEmpleos.objects.create(
                dni=dni,
                nombre_candidato=nombre_candidato,
                puestosaspira=puestosaspira,
                telefono=telefono,
                telefono2=telefono2,
                estado=estado,
                ciudad_id=ciudad_id,
                medio_reclutamiento_id=medio_reclutamiento_id,
                edad=edad,
                fecha_nacimiento=fecha_nacimiento,
                estadocivil=estadocivil,
                nhijos=nhijos,
                direccion=direccion,
                experiencia=experiencia,
                observacion=observacion,
                mediomovilizacion=mediomovilizacion,
                ruta=ruta,
                nombredocumento=nombredocumento
            )
            return JsonResponse({'success': True, 'message': 'Candidato registrado exitosamente'}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
          
def bolsaempleo_update_post_view(request, id=None):
    if request.method == 'POST' and id is not None:
        try:
            # Imprimir el contenido de la solicitud para depuración
            print(f"Contenido de request.POST: {request.POST}")
            print(f"Contenido de request.FILES: {request.FILES}")

            # Acceder al campo 'data' dentro de request.POST
            data_json = request.POST.get('data')

            if not data_json:
                return JsonResponse({'success': False, 'message': 'No se recibieron datos.'}, status=400)

            data = json.loads(data_json)

            # Imprimir datos para depuración
            print(f"Datos recibidos: {data}")

            dni = data.get('dni')
            nombre_candidato = data.get('nombre_candidato')
            puestosaspira = data.get('puestoaspira', [])
            puestosaplica = data.get('puestoaplica', [])
            telefono = data.get('telefono')
            estado = data.get('estado')

            # Validación de campos obligatorios
            if not dni or not nombre_candidato or not telefono or not puestosaspira or not estado:
                return JsonResponse({'success': False, 'message': 'Los campos DNI, Nombre del Candidato, Teléfono, Puesto al que Aspira y Estado son obligatorios.'}, status=400)

            # Buscar el candidato en la base de datos
            try:
                bolsa_empleo = BolsaEmpleos.objects.get(id=id)
            except BolsaEmpleos.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'El candidato no existe.'}, status=404)

            # Validar formato de fecha de nacimiento si se proporciona
            fecha_nacimiento = data.get('fechanacimiento')
            if fecha_nacimiento:
                try:
                    fecha_nacimiento = timezone.datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'El formato de la fecha de nacimiento es inválido. Debe ser YYYY-MM-DD.'}, status=400)

            # Obtener el archivo subido
            archivo = request.FILES.get('cveditar')  # Archivo PDF subido, si existe
            print('Archivo: ', archivo)

            # Si se sube un nuevo archivo, eliminar el anterior y guardar el nuevo
            if archivo:
                if archivo.content_type != 'application/pdf':
                    return JsonResponse({'success': False, 'message': 'Solo se permiten documentos en formato PDF.'}, status=400)

                # Eliminar archivo anterior si existe
                if bolsa_empleo.ruta and os.path.exists(os.path.join(settings.BASE_DIR, bolsa_empleo.ruta)):
                    os.remove(os.path.join(settings.BASE_DIR, bolsa_empleo.ruta))

                # Generar nombre único para el nuevo archivo
                random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                nuevo_nombre_archivo = f"{random_str}_{archivo.name}"

                # Guardar nuevo archivo
                ruta_relativa = os.path.join('reclutamiento', 'static', 'document', nuevo_nombre_archivo)
                ruta_archivo = os.path.join(settings.BASE_DIR, ruta_relativa)

                os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)

                with open(ruta_archivo, 'wb+') as destination:
                    for chunk in archivo.chunks():
                        destination.write(chunk)

                bolsa_empleo.ruta = ruta_relativa
                bolsa_empleo.nombredocumento = nuevo_nombre_archivo

            # Actualizar los campos del modelo
            bolsa_empleo.dni = dni
            bolsa_empleo.nombre_candidato = nombre_candidato
            bolsa_empleo.puestosaspira = puestosaspira
            bolsa_empleo.puestosaplica = puestosaplica
            bolsa_empleo.telefono = telefono
            bolsa_empleo.estado = estado
            bolsa_empleo.telefono2 = data.get('telefono2') or bolsa_empleo.telefono2
            bolsa_empleo.ciudad_id = data.get('ciudad') or bolsa_empleo.ciudad_id
            bolsa_empleo.medio_reclutamiento_id = data.get('mediosReclutamiento') or bolsa_empleo.medio_reclutamiento_id
            bolsa_empleo.edad = data.get('edad') or bolsa_empleo.edad
            bolsa_empleo.fecha_nacimiento = fecha_nacimiento or bolsa_empleo.fecha_nacimiento
            bolsa_empleo.estadocivil = data.get('estadocivil') or bolsa_empleo.estadocivil
            nhijos_value = data.get('nhijos')
            if nhijos_value is None:
                bolsa_empleo.nhijos = None  # Asignar None si el valor es None
            else:
                bolsa_empleo.nhijos = nhijos_value            
            bolsa_empleo.direccion = data.get('direccion') or bolsa_empleo.direccion
            bolsa_empleo.experiencia = data.get('experiencia') or bolsa_empleo.experiencia
            bolsa_empleo.observacion = data.get('observacion') or bolsa_empleo.observacion
            bolsa_empleo.mediomovilizacion = data.get('mediomovilizacion') or bolsa_empleo.mediomovilizacion

            # Guardar los cambios en la base de datos
            bolsa_empleo.save()

            return JsonResponse({'success': True, 'message': 'Candidato actualizado exitosamente'}, status=200)

        except Exception as e:
            print(f"Error inesperado: {e}")  # Mensaje de error
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def descargar_pdf(request, archivo_id):
    # Obtener el registro de la bolsa de empleos correspondiente al archivo
    try:
        registro = BolsaEmpleos.objects.get(id=archivo_id)
    except BolsaEmpleos.DoesNotExist:
        return HttpResponse(status=404, content="El archivo no existe.")

    # Verificar si el registro tiene una ruta válida
    if not registro.ruta:
        return HttpResponse(status=404, content="El archivo no tiene una ruta válida.")

    # Obtener la ruta completa del archivo
    archivo_path = Path(settings.BASE_DIR) / registro.ruta

    # Verificar si el archivo existe en el sistema de archivos
    if not archivo_path.exists():
        return HttpResponse(status=404, content="El archivo no se encuentra en el sistema de archivos.")

    # Leer el archivo y prepararlo para la descarga
    with open(archivo_path, 'rb') as archivo:
        response = HttpResponse(archivo.read(), content_type=mimetypes.guess_type(str(archivo_path))[0])
        response['Content-Disposition'] = f'attachment; filename="{archivo_path.name}"'
        return response

def exportar_bolsa_empleos(request):
    # Obtener todos los registros de BolsaEmpleos
    datos = (
        BolsaEmpleos.objects
        .prefetch_related('ciudad', 'medio_reclutamiento')
        .values(
            'id', 
            'dni', 
            'nombre_candidato',
            'puestosaspira',  
            'puestosaplica',
            'telefono', 
            'telefono2', 
            'estado', 
            'ciudad__nombre_ciudades', 
            'medio_reclutamiento__nombre_medio_de_reclutamiento', 
            'edad', 
            'fecha_nacimiento', 
            'estadocivil', 
            'nhijos', 
            'direccion', 
            'mediomovilizacion', 
            'experiencia', 
            'observacion',
            'fechacreacion'  
        )
    )

    # Convertir los datos a un DataFrame de pandas
    df = pd.DataFrame(list(datos))
    
    # Convertir IDs a nombres de puestos
    if 'puestosaspira' in df.columns:
        df['puestosaspira'] = df['puestosaspira'].apply(lambda x: 
            ', '.join(Puestos.objects.get(id=puesto_id).nombre_puestos for puesto_id in x) 
            if isinstance(x, list) else ''
        )
    else:
        df['puestosaspira'] = ''

    if 'puestosaplica' in df.columns:
        df['puestosaplica'] = df['puestosaplica'].apply(lambda x: 
            ', '.join(Puestos.objects.get(id=puesto_id).nombre_puestos for puesto_id in x) 
            if isinstance(x, list) else ''
        )
    else:
        df['puestosaplica'] = ''

    # Crear un archivo Excel
    excel_file = "bolsa_empleos.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={excel_file}'

    # Escribir el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='BolsaEmpleos')

    return response

#--------CONTROL TIEMPO VIEW--------#    
def control_tiempo_view(request):
    if request.method == 'GET':
        # Consultar las relaciones, ordenadas alfabéticamente
        unidades_negocio = Unidad_Negocio.objects.filter(estado='ACTIVO').order_by('nombre_unidad_de_negocio')
        departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento')
        puestos = Puestos.objects.filter(estado='ACTIVO').order_by('nombre_puestos')

        # Obtener todos los registros de ControlDeTiempo
        controles_tiempo = ControlDeTiempo.objects.all().prefetch_related('unidad_de_negocio', 'departamento', 'puestos')

        # Parámetros de búsqueda (opcional)
        search = request.GET.get('search', '')
        if search:
            controles_tiempo = controles_tiempo.filter(
                Q(unidad_de_negocio__nombre_unidad_de_negocio__icontains=search) |
                Q(departamento__nombre_departamento__icontains=search) |
                Q(puestos__nombre_puestos__icontains=search) |
                Q(tiempo__icontains=search)
            )

        # Paginación
        paginator = Paginator(controles_tiempo, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Contexto para renderizar la plantilla
        context = {
            'unidadesNegocio': unidades_negocio,
            'departamentos': departamentos,
            'puestos': puestos,
            'controlesTiempo': page_obj,
            'page_range': page_range,
            'search': search,
        }
        return render(request, 'controltiempo.html', context)
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Obtener los IDs del puesto y unidad de negocio
            departamento_id = data['departamento']
            puestos_id = data['puesto']
            unidad_de_negocio_id = data['unidadNegocio']
            tiempo = data['tiempo']
            estado = data['estado']  # Captura el estado del cuerpo de la solicitud

            # Obtener los nombres de la unidad de negocio y del puesto
            unidad_negocio = Unidad_Negocio.objects.get(id=unidad_de_negocio_id)
            puesto = Puestos.objects.get(id=puestos_id)

            # Crear el campo unid_puesto combinando los nombres
            unid_puesto = f"{unidad_negocio.nombre_unidad_de_negocio}-{puesto.nombre_puestos}"

            # Crear el registro
            ControlDeTiempo.objects.create(
                unidad_de_negocio_id=unidad_de_negocio_id,
                puestos_id=puestos_id,
                unid_puesto=unid_puesto,
                departamento_id=departamento_id,
                tiempo=tiempo,
                estado=estado  # Agregar el estado al registro
            )
            return JsonResponse({'success': True, 'message': 'Control de Tiempo registrado correctamente.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
        
def control_tiempo_update(request, id):
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            control = ControlDeTiempo.objects.get(id=id)

            # Obtener los nombres de la unidad de negocio y el puesto
            unidad_negocio = Unidad_Negocio.objects.get(id=data['unidadnegocio'])
            puesto = Puestos.objects.get(id=data['puestos'])

            # Actualizar los campos
            control.unid_puesto = f"{unidad_negocio.nombre_unidad_de_negocio}-{puesto.nombre_puestos}"  # Combina nombres
            control.departamento_id = data['departamento']
            control.puestos_id = data['puestos']
            control.unidad_de_negocio_id = data['unidadnegocio']
            control.tiempo = data['tiempo']
            control.estado = data['estado']
            control.save()

            return JsonResponse({'success': True, 'message': 'Control de Tiempo actualizado correctamente.'})

        except ControlDeTiempo.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Control de Tiempo no encontrado.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

#--------CONTROL PLAZAS VIEW--------#  
@csrf_exempt  
def control_plazas_view(request, id=None):
    if request.method == 'GET':
        control_de_plazas = ControlDePlazas.objects.all().order_by('-id')

        # Obtener los registros de ControlDePlazas y valores únicos para los filtros
        estatus_options = set(control_de_plazas.values_list('estatus', flat=True))
        año_options = set(control_de_plazas.values_list('año', flat=True))
        mes_corte_options = set(control_de_plazas.values_list('mes_corte', flat=True))
        mes_solicitud_options = set(control_de_plazas.values_list('mes_solicitud', flat=True))

        # Obtener todas las sucursales, empresas, etc. para el formulario
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_unidades_negocio = Unidad_Negocio.objects.filter(estado='ACTIVO').order_by('nombre_unidad_de_negocio').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_puestos = Puestos.objects.filter(estado='ACTIVO').order_by('nombre_puestos').distinct()

        # Obtener opciones para Modo, Motivo, Prioridad, Tipo de Contrato y Medio de Reclutamiento
        modos = Modo.objects.filter(estado='ACTIVO').order_by('nombre_modo').distinct()
        motivos = Motivo.objects.filter(estado='ACTIVO').order_by('nombre_motivo').distinct()
        prioridades = Prioridad.objects.filter(estado='ACTIVO').order_by('nombre_prioridad').distinct()
        tipos_contrato = TipoContrato.objects.filter(estado='ACTIVO').order_by('nombre_tipo_de_contrato').distinct()
        mediosReclutamiento = MedioReclutamiento.objects.filter(estado='ACTIVO').order_by('nombre_medio_de_reclutamiento').distinct()

        # Filtrar solo las sucursales, empresas, unidades de negocio, departamentos y puestos que estén en ControlDePlazas
        sucursales_options = Sucursal.objects.filter(
            id__in=control_de_plazas.values_list('sucursal__id', flat=True)
        ).distinct().order_by('nombre_sucursal')

        empresas_options = Empresas.objects.filter(
            id__in=control_de_plazas.values_list('empresa__id', flat=True)
        ).distinct().order_by('nombre_empresa')

        unidades_negocio_options = Unidad_Negocio.objects.filter(
            id__in=control_de_plazas.values_list('unidad_de_negocio__id', flat=True)
        ).distinct().order_by('nombre_unidad_de_negocio')

        departamentos_options = Departamento.objects.filter(
            id__in=control_de_plazas.values_list('departamento__id', flat=True)
        ).distinct().order_by('nombre_departamento')

        puestos_options = Puestos.objects.filter(
            id__in=control_de_plazas.values_list('puestos__id', flat=True)
        ).distinct().order_by('nombre_puestos')

        # Parámetros de búsqueda (opcional)
        search = request.GET.get('search', '')
        sucursal_filter = request.GET.getlist('sucursal')
        empresa_filter = request.GET.getlist('empresa')
        año_filter = request.GET.getlist('año')
        mes_corte_filter = request.GET.getlist('mes_corte')
        mes_solicitud_filter = request.GET.getlist('mes_solicitud')
        puesto_filter = request.GET.getlist('puesto')
        departamento_filter = request.GET.getlist('departamento')
        estatus_filter = request.GET.getlist('estatus')

        # Aplicar filtros en ControlDePlazas
        if search:
            control_de_plazas = control_de_plazas.filter(
                Q(sucursal__nombre_sucursal__icontains=search) |
                Q(empresa__nombre_empresa__icontains=search) |
                Q(año__icontains=search) |
                Q(mes_corte__icontains=search) |
                Q(mes_solicitud__icontains=search) |
                Q(puestos__nombre_puestos__icontains=search) |
                Q(departamento__nombre_departamento__icontains=search) |
                Q(estatus__icontains=search)
            )

        # Filtrar por Sucursal
        if sucursal_filter:
            control_de_plazas = control_de_plazas.filter(sucursal__id__in=sucursal_filter)

        # Filtrar por Empresa
        if empresa_filter:
            control_de_plazas = control_de_plazas.filter(empresa__id__in=empresa_filter)

        # Filtrar por Año
        if año_filter:
            control_de_plazas = control_de_plazas.filter(año__in=año_filter)

        # Filtrar por Mes de Corte
        if mes_corte_filter:
            control_de_plazas = control_de_plazas.filter(mes_corte__in=mes_corte_filter)

        # Filtrar por Mes de Solicitud
        if mes_solicitud_filter:
            control_de_plazas = control_de_plazas.filter(mes_solicitud__in=mes_solicitud_filter)

        # Filtrar por Puesto
        if puesto_filter:
            control_de_plazas = control_de_plazas.filter(puestos__id__in=puesto_filter)

        # Filtrar por Departamento
        if departamento_filter:
            control_de_plazas = control_de_plazas.filter(departamento__id__in=departamento_filter)

        # Filtrar por Estatus
        if estatus_filter:
            control_de_plazas = control_de_plazas.filter(estatus__in=estatus_filter)

        # Paginación
        paginator = Paginator(control_de_plazas, 10) 
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Contexto para renderizar la plantilla
        context = {
            'sucursales': sucursales_options,
            'empresas': empresas_options,
            'unidadesNegocio': unidades_negocio_options,
            'departamentos': departamentos_options,
            'puestos': puestos_options,
            'estatus_options': list(estatus_options),  # Convertir set a list para el contexto
            'año_options': list(año_options),  # Convertir set a list para el contexto
            'mes_corte_options': list(mes_corte_options),  # Convertir set a list para el contexto
            'mes_solicitud_options': list(mes_solicitud_options),  # Convertir set a list para el contexto
            'controlDePlazas': page_obj,
            'page_range': page_range,
            'search': search,
            'sucursal_filter': sucursal_filter,
            'empresa_filter': empresa_filter,
            'año_filter': año_filter,
            'mes_corte_filter': mes_corte_filter,
            'mes_solicitud_filter': mes_solicitud_filter,
            'puesto_filter': puesto_filter,
            'departamento_filter': departamento_filter,
            'estatus_filter': estatus_filter,
            'all_sucursales': all_sucursales,
            'all_empresas': all_empresas,
            'all_unidades_negocio': all_unidades_negocio,
            'all_departamentos': all_departamentos,
            'all_puestos': all_puestos,
            'modos': modos,
            'motivos': motivos,
            'prioridades': prioridades,
            'tipos_contrato': tipos_contrato,
            'mediosReclutamiento': mediosReclutamiento,
        }

        return render(request, 'controlplazas.html', context)
    if request.method == 'POST':
        try:
            # Validación de los datos de entrada
            data = json.loads(request.POST.get('data'))

            # Consultar la instancia de empresa (puede ser None)
            empresa = Empresas.objects.filter(id=data.get('empresa')).first()

            # Consultar la instancia de sucursal (puede ser None)
            sucursal = Sucursal.objects.filter(id=data.get('sucursal')).first()

            # Consultar Unidad de Negocio (puede ser None)
            unidad_negocio = Unidad_Negocio.objects.filter(id=data.get('unidad_de_negocio')).first()

            # Consultar Puesto (puede ser None)
            puesto = Puestos.objects.filter(id=data.get('puestos')).first()

            # Consultar otros modelos relacionados (pueden ser None)
            modo = Modo.objects.filter(id=data.get('modo')).first()
            motivo_ingreso = Motivo.objects.filter(id=data.get('motivo_ingreso')).first()
            departamento = Departamento.objects.filter(id=data.get('departamento')).first() if data.get('departamento') else None
            prioridad = Prioridad.objects.filter(id=data.get('prioridad')).first() if data.get('prioridad') else None
            medio_reclutamiento = MedioReclutamiento.objects.filter(id=data.get('medio_reclutamiento')).first() if data.get('medio_reclutamiento') else None
            tipo_contrato = TipoContrato.objects.filter(id=data.get('tipo_contrato')).first() if data.get('tipo_contrato') else None

            # Construir unidad_puesto (puede ser None)
            unidad_puesto = f"{unidad_negocio.nombre_unidad_de_negocio} - {puesto.nombre_puestos}" if unidad_negocio and puesto else None

            # Consultar ControlDeTiempo para obtener el tiempo de cobertura (puede ser None)
            control_de_tiempo = ControlDeTiempo.objects.filter(
                unidad_de_negocio=unidad_negocio,
                puestos=puesto,
                unid_puesto=unidad_puesto
            ).first()
            tiempo_cobertura = control_de_tiempo.tiempo if control_de_tiempo else None

            # Manejo de las fechas
            fecha_solicitud = parser.parse(data.get('fecha_solicitud')) if data.get('fecha_solicitud') else datetime.now()
            fecha_cobertura = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else datetime.now()
            fecha_ingreso = parser.parse(data.get('fecha_ingreso')) if data.get('fecha_ingreso') else datetime.now()

            # Calcular tiempos efectivos
            tiempo_efectivo_cobertura = (fecha_cobertura - fecha_solicitud).days if fecha_solicitud and fecha_cobertura else None
            tiempo_efectivo_cobertura = tiempo_efectivo_cobertura if tiempo_efectivo_cobertura > 0 else None

            tiempo_efectivo_fecha_ingreso = (fecha_ingreso - fecha_solicitud).days if fecha_solicitud and fecha_ingreso else None
            tiempo_efectivo_fecha_ingreso = tiempo_efectivo_fecha_ingreso if tiempo_efectivo_fecha_ingreso > 0 else None

            # Calcular fecha límite de cobertura
            fecha_limite_cobertura_str = (fecha_solicitud + timedelta(days=tiempo_cobertura)).date() if fecha_solicitud and tiempo_cobertura else None

            # Capturar y convertir las fechas
            fecha_limite_cobertura_cal = datetime.combine(fecha_limite_cobertura_str, datetime.min.time()) if fecha_limite_cobertura_str else None
            fecha_cobertura_cal = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else None

            # Calcular el tiempo disponible
            if fecha_limite_cobertura_cal and fecha_cobertura_cal:
                tiempo_disponible = (fecha_limite_cobertura_cal - fecha_cobertura_cal).days
                tiempo_disponible = tiempo_disponible if tiempo_disponible >= 0 else None
            else:
                tiempo_disponible = None

            # Determinar el estatus
            estatus = 'EN PROCESO' if not data.get('fecha_cobertura') else 'CERRADA' if 'fecha_ingreso' in data else 'ESPERA DE INGRESO'

            # Asignar cantidad solicitada y cantidad cubierta
            cantidad_solicitada = 1
            cantidad_cubierta = 1 if estatus == 'CERRADA' else None

            # Manejo de la imagen
            imagen_dni = request.FILES.get('imagen_dni')
            nombre_imagen = None
            ruta_imagen = None

            if imagen_dni:
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen_dni.name}"
                ruta_imagen = f"reclutamiento/static/img/controlplazas/{nombre_imagen}"

                with open(ruta_imagen, 'wb+') as destino:
                    for chunk in imagen_dni.chunks():
                        destino.write(chunk)

            # Manejo de la imagen del reverso
            imagen_dni_reverso = request.FILES.get('imagen_dni_reverso')
            nombre_imagen_reverso = None
            ruta_imagen_reverso = None

            if imagen_dni_reverso:
                nombre_imagen_reverso = f"{uuid.uuid4().hex[:4]}_{imagen_dni_reverso.name}"  # Generar un nombre único
                ruta_imagen_reverso = f"reclutamiento/static/img/controlplazas/{nombre_imagen_reverso}"

                with open(ruta_imagen_reverso, 'wb+') as destino:
                    for chunk in imagen_dni_reverso.chunks():
                        destino.write(chunk)

            # Crear el registro en la base de datos
            control_plaza = ControlDePlazas.objects.create(
                analista=data.get('analistas'),
                sucursal=sucursal,  # Puede ser None
                empresa=empresa,  # Puede ser None
                unidad_de_negocio=unidad_negocio,  # Puede ser None
                año=data.get('año'),
                mes_corte=data.get('mes_corte'),
                mes_solicitud=data.get('mes_solicitud'),
                modo=modo,
                motivo_ingreso=motivo_ingreso,
                nombre_reemplazo=data.get('nombre_reemplazo'),
                puestos=puesto,  # Puede ser None
                departamento=departamento,  # Puede ser None
                prioridad=prioridad,  # Puede ser None
                unidad_puesto=unidad_puesto,
                tiempo_cobertura=tiempo_cobertura,
                fecha_solicitud=data.get('fecha_solicitud'),
                fecha_inicio_busqueda=parser.parse(data.get('fecha_inicio_busqueda')) if data.get('fecha_inicio_busqueda') else None,
                fecha_cobertura=data.get('fecha_cobertura'),
                fecha_ingreso=data.get('fecha_ingreso'),
                tiempo_efectivo_cobertura=tiempo_efectivo_cobertura,
                tiempo_efectivo_fecha_ingreso=tiempo_efectivo_fecha_ingreso,
                fecha_limite_cobertura=fecha_limite_cobertura_str,
                tiempo_disponible=tiempo_disponible,
                estatus=estatus,
                cantidad_solicitada=cantidad_solicitada,
                cantidad_cubierta=cantidad_cubierta,
                fuente_reclutamiento=data.get('fuente_reclutamiento'),
                nombre_contratado=data.get('nombre_contratado'),
                dni=data.get('dni'),
                genero=data.get('genero'),
                edad=data.get('edad'),
                fecha_nacimiento=data.get('fecha_nacimiento'),
                medio_reclutamiento=medio_reclutamiento,  # Puede ser None
                salario=data.get('salario'),
                combustible=data.get('combustible'),
                depreciacion=data.get('depreciacion'),
                comision=data.get('comision'),
                bono=data.get('bono'),
                hospedaje=data.get('hospedaje'),
                tipo_contrato=tipo_contrato,  # Puede ser None
                ruta=ruta_imagen,
                nombreimagen=nombre_imagen,
                ruta1=ruta_imagen_reverso,
                nombreimagen1=nombre_imagen_reverso,
            )

            return JsonResponse({'success': True, 'message': 'Plaza registrada exitosamente', 'controlPlaza': control_plaza.id}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    if request.method == 'DELETE':
        try:
            id = request.path.split('/')[-2]
            control_plaza = ControlDePlazas.objects.get(id=id)
            control_plaza.delete()
            return JsonResponse({'success': True, 'message': 'Plaza eliminada exitosamente'}, status=200)

        except ControlDePlazas.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Control de plazas no encontrado'}, status=404)
        except Exception as e:
            # Asegúrate de capturar cualquier excepción inesperada
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    if request.method == 'PATCH':
        try:
            control_plaza = ControlDePlazas.objects.get(id=id)
            data = json.loads(request.body)  # Obtiene los datos del cuerpo de la solicitud
            estatus = data.get('estatus', None)

            if estatus:
                control_plaza.estatus = estatus  # Actualiza el estatus
                control_plaza.save()  # Guarda los cambios

            return JsonResponse({'success': True, 'message': 'Estatus actualizado exitosamente'}, status=200)

        except ControlDePlazas.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Control de plazas no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt  
def control_plazas_update(request, id):
    if request.method == 'POST':
        try:
            # Obtener el objeto ControlDePlazas que se va a actualizar
            control_plaza = ControlDePlazas.objects.get(id=id)

            # Validación de los datos de entrada
            data = json.loads(request.POST.get('data'))

            # Consultar la instancia de empresa (puede ser None)
            empresa = Empresas.objects.filter(id=data.get('empresa')).first()

            # Consultar la instancia de sucursal (puede ser None)
            sucursal = Sucursal.objects.filter(id=data.get('sucursal')).first()

            # Consultar Unidad de Negocio (puede ser None)
            unidad_negocio = Unidad_Negocio.objects.filter(id=data.get('unidad_de_negocio')).first()

            # Consultar Puesto (puede ser None)
            puesto = Puestos.objects.filter(id=data.get('puestos')).first()

            # Consultar otros modelos relacionados (pueden ser None)
            modo = Modo.objects.filter(id=data.get('modo')).first()
            motivo_ingreso = Motivo.objects.filter(id=data.get('motivo_ingreso')).first()
            departamento = Departamento.objects.filter(id=data.get('departamento')).first() if data.get('departamento') else None
            prioridad = Prioridad.objects.filter(id=data.get('prioridad')).first() if data.get('prioridad') else None
            medio_reclutamiento = MedioReclutamiento.objects.filter(id=data.get('medio_reclutamiento')).first() if data.get('medio_reclutamiento') else None
            tipo_contrato = TipoContrato.objects.filter(id=data.get('tipo_contrato')).first() if data.get('tipo_contrato') else None

            # Construir unidad_puesto (puede ser None)
            unidad_puesto = f"{unidad_negocio.nombre_unidad_de_negocio} - {puesto.nombre_puestos}" if unidad_negocio and puesto else None

            # Consultar ControlDeTiempo para obtener el tiempo de cobertura (puede ser None)
            control_de_tiempo = ControlDeTiempo.objects.filter(
                unidad_de_negocio=unidad_negocio,
                puestos=puesto,
                unid_puesto=unidad_puesto
            ).first()
            tiempo_cobertura = control_de_tiempo.tiempo if control_de_tiempo else None

            # Manejo de las fechas
            fecha_solicitud = parser.parse(data.get('fecha_solicitud')) if data.get('fecha_solicitud') else datetime.now()
            fecha_cobertura = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else datetime.now()
            fecha_ingreso = parser.parse(data.get('fecha_ingreso')) if data.get('fecha_ingreso') else datetime.now()

            # Calcular tiempos efectivos
            tiempo_efectivo_cobertura = (fecha_cobertura - fecha_solicitud).days if fecha_solicitud and fecha_cobertura else None
            tiempo_efectivo_cobertura = tiempo_efectivo_cobertura if tiempo_efectivo_cobertura > 0 else None

            tiempo_efectivo_fecha_ingreso = (fecha_ingreso - fecha_solicitud).days if fecha_solicitud and fecha_ingreso else None
            tiempo_efectivo_fecha_ingreso = tiempo_efectivo_fecha_ingreso if tiempo_efectivo_fecha_ingreso > 0 else None

            # Calcular fecha límite de cobertura
            fecha_limite_cobertura_str = (fecha_solicitud + timedelta(days=tiempo_cobertura)).date() if fecha_solicitud and tiempo_cobertura else None

            # Capturar y convertir las fechas
            fecha_limite_cobertura_cal = datetime.combine(fecha_limite_cobertura_str, datetime.min.time()) if fecha_limite_cobertura_str else None
            fecha_cobertura_cal = parser.parse(data.get('fecha_cobertura')) if data.get('fecha_cobertura') else None

            # Calcular el tiempo disponible
            if fecha_limite_cobertura_cal and fecha_cobertura_cal:
                tiempo_disponible = (fecha_limite_cobertura_cal - fecha_cobertura_cal).days
                tiempo_disponible = tiempo_disponible if tiempo_disponible >= 0 else None
            else:
                tiempo_disponible = None

            # Determinar el estatus
            estatus = 'EN PROCESO' if not data.get('fecha_cobertura') else 'CERRADA' if 'fecha_ingreso' in data else 'ESPERA DE INGRESO'

            # Asignar cantidad solicitada y cantidad cubierta
            cantidad_solicitada = 1
            cantidad_cubierta = 1 if estatus == 'CERRADA' else None

            # Manejo de la imagen
            imagen_dni = request.FILES.get('imagen_dni')
            if imagen_dni:
                # Eliminar la imagen anterior si existe
                if control_plaza.nombreimagen:
                    old_image_path = os.path.join('reclutamiento/static/img/controlplazas', control_plaza.nombreimagen)
                    if os.path.isfile(old_image_path):
                        os.remove(old_image_path)

                # Guardar la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen_dni.name}"  # Toma solo los primeros 4 caracteres
                ruta_imagen = f"reclutamiento/static/img/controlplazas/{nombre_imagen}"

                with open(ruta_imagen, 'wb+') as destino:
                    for chunk in imagen_dni.chunks():
                        destino.write(chunk)

                # Actualizar el nombre y la ruta de la imagen
                control_plaza.ruta = ruta_imagen
                control_plaza.nombreimagen = nombre_imagen

            imagen_dni_reverso = request.FILES.get('imagen_dni_reverso')
            if imagen_dni_reverso:
                # Eliminar la imagen anterior si existe
                if control_plaza.nombreimagen1:  # Asegúrate de usar el campo correcto
                    old_image_path_reverso = os.path.join('reclutamiento/static/img/controlplazas', control_plaza.nombreimagen1)
                    if os.path.isfile(old_image_path_reverso):
                        os.remove(old_image_path_reverso)

                # Guardar la nueva imagen del reverso
                nombre_imagen_reverso = f"{uuid.uuid4().hex[:4]}_{imagen_dni_reverso.name}"  # Toma solo los primeros 4 caracteres
                ruta_imagen_reverso = f"reclutamiento/static/img/controlplazas/{nombre_imagen_reverso}"

                with open(ruta_imagen_reverso, 'wb+') as destino:
                    for chunk in imagen_dni_reverso.chunks():
                        destino.write(chunk)

                # Actualizar el nombre y la ruta de la imagen del reverso
                control_plaza.ruta1 = ruta_imagen_reverso
                control_plaza.nombreimagen1 = nombre_imagen_reverso

            # Actualizar el registro en la base de datos
            control_plaza.empresa = empresa
            control_plaza.sucursal = sucursal
            control_plaza.unidad_de_negocio = unidad_negocio
            control_plaza.año = data.get('año')
            control_plaza.mes_corte = data.get('mes_corte')
            control_plaza.mes_solicitud = data.get('mes_solicitud')
            control_plaza.modo = modo
            control_plaza.motivo_ingreso = motivo_ingreso
            control_plaza.nombre_reemplazo = data.get('nombre_reemplazo')
            control_plaza.puestos = puesto
            control_plaza.departamento = departamento
            control_plaza.prioridad = prioridad
            control_plaza.unidad_puesto = unidad_puesto
            control_plaza.tiempo_cobertura = tiempo_cobertura
            control_plaza.fecha_solicitud = fecha_solicitud
            control_plaza.fecha_inicio_busqueda = parser.parse(data.get('fecha_inicio_busqueda')) if data.get('fecha_inicio_busqueda') else None
            control_plaza.fecha_cobertura = fecha_cobertura
            control_plaza.fecha_ingreso = fecha_ingreso
            control_plaza.tiempo_efectivo_cobertura = tiempo_efectivo_cobertura
            control_plaza.tiempo_efectivo_fecha_ingreso = tiempo_efectivo_fecha_ingreso
            control_plaza.fecha_limite_cobertura = fecha_limite_cobertura_str
            control_plaza.tiempo_disponible = tiempo_disponible
            control_plaza.estatus = estatus
            control_plaza.cantidad_solicitada = cantidad_solicitada
            control_plaza.cantidad_cubierta = cantidad_cubierta
            control_plaza.fuente_reclutamiento = data.get('fuente_reclutamiento')
            control_plaza.nombre_contratado = data.get('nombre_contratado')
            control_plaza.dni = data.get('dni')
            control_plaza.genero = data.get('genero')
            control_plaza.edad = data.get('edad')
            control_plaza.fecha_nacimiento = data.get('fecha_nacimiento')
            control_plaza.medio_reclutamiento = medio_reclutamiento
            control_plaza.salario = data.get('salario')
            control_plaza.combustible = data.get('combustible')
            control_plaza.depreciacion = data.get('depreciacion')
            control_plaza.comision = data.get('comision')
            control_plaza.bono = data.get('bono')
            control_plaza.hospedaje = data.get('hospedaje')
            control_plaza.tipo_contrato = tipo_contrato

            control_plaza.save()  # Guardar los cambios

            return JsonResponse({'success': True, 'message': 'Plaza actualizada exitosamente'}, status=200)

        except ControlDePlazas.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Control de plazas no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def descargar_imagen(request, imagen_id):
    try:
        registro = ControlDePlazas.objects.get(id=imagen_id)
    except ControlDePlazas.DoesNotExist:
        return HttpResponse(status=404, content="El registro no existe.")

    imagenes = {}

    if registro.nombreimagen:
        imagen_path = Path(settings.BASE_DIR) / registro.ruta
        if imagen_path.exists():
            # Cambiar a una URL accesible a través del servidor
            imagenes['frontal'] = f"/static/img/controlplazas/{registro.nombreimagen}"

    if registro.nombreimagen1:
        imagen_path_reverso = Path(settings.BASE_DIR) / registro.ruta1
        if imagen_path_reverso.exists():
            # Cambiar a una URL accesible a través del servidor
            imagenes['reverso'] = f"/static/img/controlplazas/{registro.nombreimagen1}"

    if not imagenes:
        return HttpResponse(status=404, content="No hay imágenes disponibles para descargar.")

    return JsonResponse({'success': True, 'imagenes': imagenes})

class ExportExcelPlazas(View):
    def get(self, request, *args, **kwargs):
        template_path = os.path.join('reclutamiento/static/templates/formatoplazaoriginal.xlsx')
        
        # Cargar la plantilla de Excel
        try:
            wb = load_workbook(template_path)
            sheet = wb['CUADRO GENERAL DE PLAZAS']  # Ajusta esto según el nombre de tu hoja
        except Exception as e:
            return HttpResponse(f"Error al cargar la plantilla de Excel: {e}", status=500)

        # Obtener los datos de ControlDePlazas
        control_de_plazas = ControlDePlazas.objects.all().order_by('-id')
        
        # Rango de la fila donde se comenzarán a escribir los datos
        start_row = 9
        
        # Escribir los datos en la hoja de Excel
        for idx, plaza in enumerate(control_de_plazas, start=start_row):
            try:
                sheet.cell(row=idx, column=2, value=plaza.analista)
                sheet.cell(row=idx, column=3, value=plaza.sucursal.nombre_sucursal if plaza.sucursal else '')
                sheet.cell(row=idx, column=4, value=plaza.empresa.nombre_empresa if plaza.empresa else '')
                sheet.cell(row=idx, column=5, value=plaza.unidad_de_negocio.nombre_unidad_de_negocio if plaza.unidad_de_negocio else '')
                sheet.cell(row=idx, column=6, value=plaza.año)
                sheet.cell(row=idx, column=7, value=plaza.mes_corte)
                sheet.cell(row=idx, column=8, value=plaza.mes_solicitud)
                sheet.cell(row=idx, column=9, value=plaza.modo.nombre_modo if plaza.modo else '')
                sheet.cell(row=idx, column=10, value=plaza.motivo_ingreso.nombre_motivo if plaza.motivo_ingreso else '')
                sheet.cell(row=idx, column=11, value=plaza.nombre_reemplazo)
                sheet.cell(row=idx, column=12, value=plaza.puestos.nombre_puestos if plaza.puestos else '')
                sheet.cell(row=idx, column=13, value=plaza.departamento.nombre_departamento if plaza.departamento else '')
                sheet.cell(row=idx, column=14, value=plaza.prioridad.nombre_prioridad if plaza.prioridad else '')
                sheet.cell(row=idx, column=15, value=plaza.unidad_puesto)
                sheet.cell(row=idx, column=16, value=plaza.tiempo_cobertura)
                sheet.cell(row=idx, column=17, value=plaza.fecha_solicitud)
                sheet.cell(row=idx, column=18, value=plaza.fecha_inicio_busqueda)
                sheet.cell(row=idx, column=19, value=plaza.fecha_cobertura)
                sheet.cell(row=idx, column=20, value=plaza.fecha_ingreso)
                sheet.cell(row=idx, column=21, value=plaza.tiempo_efectivo_cobertura)
                sheet.cell(row=idx, column=22, value=plaza.tiempo_efectivo_fecha_ingreso)
                sheet.cell(row=idx, column=23, value=plaza.fecha_limite_cobertura)
                sheet.cell(row=idx, column=24, value=plaza.tiempo_disponible)
                sheet.cell(row=idx, column=25, value=plaza.estatus)
                sheet.cell(row=idx, column=26, value=plaza.cantidad_solicitada)
                sheet.cell(row=idx, column=27, value=plaza.cantidad_cubierta)
                sheet.cell(row=idx, column=28, value=plaza.fuente_reclutamiento)
                sheet.cell(row=idx, column=29, value=plaza.nombre_contratado)
                sheet.cell(row=idx, column=30, value=plaza.dni)
                sheet.cell(row=idx, column=31, value=plaza.genero)
                sheet.cell(row=idx, column=32, value=plaza.edad)
                sheet.cell(row=idx, column=33, value=plaza.medio_reclutamiento.nombre_medio_de_reclutamiento if plaza.medio_reclutamiento else '')
                sheet.cell(row=idx, column=34, value=plaza.salario)
                sheet.cell(row=idx, column=35, value=plaza.combustible)
                sheet.cell(row=idx, column=36, value=plaza.depreciacion)
                sheet.cell(row=idx, column=37, value=plaza.comision)
                sheet.cell(row=idx, column=38, value=plaza.bono)
                sheet.cell(row=idx, column=39, value=plaza.hospedaje)
                sheet.cell(row=idx, column=40, value=plaza.tipo_contrato.nombre_tipo_de_contrato if plaza.tipo_contrato else '')
            except Exception as e:
                return HttpResponse(f"Error escribiendo datos en Excel en la fila {idx}: {e}", status=500)

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Control_de_Plazas.xlsx"'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

#--------CONTRATACIONES--------#  
def contratacionesform_view(request):
    if request.method == 'GET':
    # Obtener todas las sucursales, empresas, etc. para el formulario
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_unidades_negocio = Unidad_Negocio.objects.filter(estado='ACTIVO').order_by('nombre_unidad_de_negocio').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_puestos = Puestos.objects.filter(estado='ACTIVO').order_by('nombre_puestos').distinct()

        # Obtener opciones para Modo, Motivo, Prioridad, Tipo de Contrato y Medio de Reclutamiento
        all_modos = Modo.objects.filter(estado='ACTIVO').order_by('nombre_modo').distinct()
        all_motivos = Motivo.objects.filter(estado='ACTIVO').order_by('nombre_motivo').distinct()
        all_prioridades = Prioridad.objects.filter(estado='ACTIVO').order_by('nombre_prioridad').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(estado='ACTIVO').order_by('nombre_tipo_de_contrato').distinct()
        all_mediosReclutamiento = MedioReclutamiento.objects.filter(estado='ACTIVO').order_by('nombre_medio_de_reclutamiento').distinct()

        all_municipios = MunicipioHonduras.objects.all().order_by('nombre_municipio').distinct()
        all_departamentos_hn = DepartamentoHonduras.objects.all().order_by('nombre_departamentohonduras').distinct()

        context = {
            'all_sucursales': all_sucursales,
            'all_empresas': all_empresas,
            'all_unidades_negocio': all_unidades_negocio,
            'all_departamentos': all_departamentos,
            'all_puestos': all_puestos,
            'all_modos': all_modos,
            'all_motivos': all_motivos,
            'all_prioridades': all_prioridades,
            'all_tipos_contrato': all_tipos_contrato,
            'all_mediosReclutamiento': all_mediosReclutamiento,
            'all_municipios': all_municipios, 
            'all_departamentos_hn': all_departamentos_hn,
        }
        return render(request, 'contrataciones/registrarcontrataciones.html', context)
    
    if request.method == 'POST':
        try:
            # Obtener campos individuales y asignar None si están vacíos
            primer_nombre = request.POST.get('primerNombre') or None
            segundo_nombre = request.POST.get('segundoNombre') or None
            primer_apellido = request.POST.get('primerApellido') or None
            segundo_apellido = request.POST.get('segundoApellido') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None
            municipio = request.POST.get('municipio') or None
            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccion') or None
            dni = request.POST.get('cedula') or None
            estado_civil = request.POST.get('estadoCivil') or None
            hijos = 1 if request.POST.get('hijos') == '1' else 0
            profesion = request.POST.get('profesion') or None
            correo = request.POST.get('correo') or None
            departamento = request.POST.get('departamento') or None
            telefono = request.POST.get('telefono') or None

            # Datos de emergencia
            emergencia1 = request.POST.get('emergencia1') or None
            parentesco1 = request.POST.get('parentesco1') or None
            telefono_emergencia1 = request.POST.get('telefonoEmergencia1') or None
            emergencia2 = request.POST.get('emergencia2') or None
            parentesco2 = request.POST.get('parentesco2') or None
            telefono_emergencia2 = request.POST.get('telefonoEmergencia2') or None

            # Nivel educativo
            nivel_educativo = request.POST.get('alfabeta') or None
            ultimo_grado_estudio = request.POST.get('ultimoGrado') or None
            ultimogradodetalle = request.POST.get('ultimogradodetalle') or None
            padecimiento = 1 if request.POST.get('padecimiento') == '1' else 0
            detalle_enfermedad = request.POST.get('detalleEnfermedad') or None

            # Datos laborales
            puesto = request.POST.get('puesto') or None
            unidad_negocio = request.POST.get('unidadnegocio') or None
            salario = request.POST.get('salario') or None
            comision = request.POST.get('comision') or None
            bofa = request.POST.get('bofa') or None
            sucursal = request.POST.get('sucursal') or None
            tipo_contrato = request.POST.get('tipocontrato') or None
            fecha_ingreso = request.POST.get('fechaIngreso') or None
            fecha_vencimiento = request.POST.get('fechavencimiento') or None
            departamento_empresa = request.POST.get('departamentoEmpresa') or None
            direccion_empresa = request.POST.get('direccionempresa') or None
            nombre_empresa = request.POST.get('nombreempresa') or None
            telefono_empresa = request.POST.get('telefonoempresa') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombreBeneficiario') or None
            identidad_beneficiario = request.POST.get('identidadBeneficiario') or None
            parentesco_beneficiario = request.POST.get('parentescoBeneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentajeBeneficiario') or None

            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Si hay imagen, generar un nombre único para el archivo
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"  # Usando UUID para nombre único
                ruta_imagen = f"reclutamiento/static/img/contrataciones/{nombre_imagen}"

                # Crear directorio si no existe
                try:
                    os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                    # Guardar la imagen en la ruta especificada
                    with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                        for chunk in imagen.chunks():
                            destino.write(chunk)
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Error al guardar la imagen: {str(e)}'}, status=500)
            else:
                # Si no hay imagen, mantener nombre_imagen y ruta_imagen como None
                nombre_imagen = None
                ruta_imagen = None

            # Crear nueva entrada en ContratacionEmpleados
            contratacion = ContratacionEmpleados.objects.create(
                tipo_contratacion=request.POST.get('tipoIngreso') or None,
                nombre1=primer_nombre,
                nombre2=segundo_nombre,
                apellido1=primer_apellido,
                apellido2=segundo_apellido,
                fecha_nacimiento=fecha_nacimiento,
                municipio_id=municipio,
                genero=genero,
                direccionexacta=direccion,
                dni=dni,
                estado_civil=estado_civil,
                hijos=hijos,
                profecion_oficio=profesion,
                correo=correo,
                departamento_id=departamento,
                telefono=telefono,
                nombre1_emergencia=emergencia1,
                parentesco1=parentesco1,
                telefonoemergencia1=telefono_emergencia1,
                nombre2_emergencia=emergencia2,
                parentesco2=parentesco2,
                telefonoemergencia2=telefono_emergencia2,
                nivel_educativo=nivel_educativo,
                ultimo_grado_estudio=ultimo_grado_estudio,
                ultimogradodetalle=ultimogradodetalle,
                padecimiento=padecimiento,
                detalle_enfermedad=detalle_enfermedad,
                puestos_id=puesto,
                unidad_de_negocio_id=unidad_negocio,
                sucursal_id=sucursal,
                departamento_empresa_id=departamento_empresa,
                tipo_contrato_id=tipo_contrato,
                salario=salario,
                comision=comision,
                bofa=bofa,
                fecha_ingreso=fecha_ingreso,
                fecha_vencimiento=fecha_vencimiento,
                direccionempresa=direccion_empresa,
                nombre_empresa_id=nombre_empresa,
                telefono_empresa=telefono_empresa,
                nombre_beneficiario=nombre_beneficiario,
                dni_beneficiario=identidad_beneficiario,
                parentesco_beneficiario=parentesco_beneficiario,
                porcentaje=porcentaje_beneficiario,
                ruta=ruta_imagen,
                nombreimagen=nombre_imagen,
            )

            return JsonResponse({'success': True, 'message': 'Candidato registrado exitosamente'}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        
def contratacionestable_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')

        # Consulta los datos de ContratacionEmpleados y las relaciones de ForeignKey
        contrataciones = ContratacionEmpleados.objects.select_related(
            'puestos', 
            'unidad_de_negocio', 
            'sucursal', 
            'municipio',  # Agregar el municipio para el filtro de ciudad
        )

        # Obtener valores únicos para los filtros
        municipios_options = set(contrataciones.values_list('municipio__nombre_municipio', flat=True))
        puestos_options = set(contrataciones.values_list('puestos__nombre_puestos', flat=True))
        unidades_negocio_options = set(contrataciones.values_list('unidad_de_negocio__nombre_unidad_de_negocio', flat=True))
        sucursales_options = set(contrataciones.values_list('sucursal__nombre_sucursal', flat=True))
        fechas_ingreso_options = set(contrataciones.values_list('fecha_ingreso', flat=True))

        # Filtrar por los filtros seleccionados (si se aplican)
        municipio_filter = request.GET.getlist('municipio')
        puesto_filter = request.GET.getlist('puesto')
        unidad_negocio_filter = request.GET.getlist('unidad_negocio')
        sucursal_filter = request.GET.getlist('sucursal')
        fecha_ingreso_filter = request.GET.getlist('fecha_ingreso')

        # Aplicar filtros a la consulta
        if municipio_filter:
            contrataciones = contrataciones.filter(municipio__nombre_municipio__in=municipio_filter)
        
        if puesto_filter:
            contrataciones = contrataciones.filter(puestos__nombre_puestos__in=puesto_filter)
        
        if unidad_negocio_filter:
            contrataciones = contrataciones.filter(unidad_de_negocio__nombre_unidad_de_negocio__in=unidad_negocio_filter)
        
        if sucursal_filter:
            contrataciones = contrataciones.filter(sucursal__nombre_sucursal__in=sucursal_filter)

        if fecha_ingreso_filter:
            contrataciones = contrataciones.filter(fecha_ingreso__in=fecha_ingreso_filter)

        if search:
            contrataciones = contrataciones.filter(
                models.Q(nombre1__icontains=search) |
                models.Q(nombre2__icontains=search) |
                models.Q(apellido1__icontains=search) |
                models.Q(apellido2__icontains=search) |
                models.Q(telefono__icontains=search) |
                models.Q(dni__icontains=search) |
                models.Q(correo__icontains=search)
            )

        # Paginación
        paginator = Paginator(contrataciones, 10)  # 10 registros por página
        page_number = request.GET.get('page')  # Capturar el número de página actual
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        # Contexto para la plantilla
        context = {
            'contrataciones': page_obj,  # Pasar el objeto paginado
            'search': search,  # Pasar el valor de búsqueda si aplica
            'page_range': page_range,  # Pasar el rango de páginas
            'municipios_options': list(municipios_options),
            'puestos_options': list(puestos_options),
            'unidades_negocio_options': list(unidades_negocio_options),
            'sucursales_options': list(sucursales_options),
            'fechas_ingreso_options': list(fechas_ingreso_options),
            'municipio_filter': municipio_filter,
            'puesto_filter': puesto_filter,
            'unidad_negocio_filter': unidad_negocio_filter,
            'sucursal_filter': sucursal_filter,
            'fecha_ingreso_filter': fecha_ingreso_filter,
        }

        return render(request, 'contrataciones/contrataciones.html', context)

    if request.method == 'DELETE':
        if id:
            contratacion = get_object_or_404(ContratacionEmpleados, id=id)

            # Eliminar la imagen si existe
            if contratacion.ruta:
                ruta_imagen = os.path.join(settings.BASE_DIR, contratacion.ruta)

                if os.path.exists(ruta_imagen):
                    os.remove(ruta_imagen)

            contratacion.delete()
            return JsonResponse({'success': True, 'message': 'Contratación eliminada con éxito.'})

        return JsonResponse({'success': False, 'message': 'ID de contratación no encontrado.'})
    
def imprimir_contratacion_view(request, contratacion_id):
    contratacion = get_object_or_404(ContratacionEmpleados, id=contratacion_id)
    
    context = {
        'contratacion': contratacion,
    }
    return render(request, 'contrataciones/imprimir_contratacion.html', context)

def updatecontrataciones_view(request, contratacion_id):
    contratacion = get_object_or_404(ContratacionEmpleados, id=contratacion_id)

    # Si es una solicitud GET, se muestra el formulario con los datos actuales
    if request.method == 'GET':
        context = {
            'contratacion': contratacion,
            'all_municipios': MunicipioHonduras.objects.all(),
            'all_departamentos': DepartamentoHonduras.objects.all(),
            'all_puestos': Puestos.objects.all(),
            'all_unidades_negocio': Unidad_Negocio.objects.all(),
            'all_sucursales': Sucursal.objects.all(),
            'all_tipos_contrato': TipoContrato.objects.all(),
            'all_empresas': Empresas.objects.all(),
            'all_departamentoempresa': Departamento.objects.all(),
        }
        return render(request, 'contrataciones/updatecontrataciones.html', context)

    if request.method == 'POST':
        try:
            # Obtener campos individuales y asignar None si están vacíos
            primer_nombre = request.POST.get('primerNombre') or None
            segundo_nombre = request.POST.get('segundoNombre') or None
            primer_apellido = request.POST.get('primerApellido') or None
            segundo_apellido = request.POST.get('segundoApellido') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None

            # Claves foráneas deben ser números o None
            municipio = request.POST.get('municipio')
            municipio_id = int(municipio) if municipio and municipio.isdigit() else None

            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccion') or None
            dni = request.POST.get('cedula') or None
            estado_civil = request.POST.get('estadoCivil') or None
            hijos = 1 if request.POST.get('hijos') == '1' else 0
            profesion = request.POST.get('profesion') or None
            correo = request.POST.get('correo') or None

            departamento = request.POST.get('departamento')
            departamento_id = int(departamento) if departamento and departamento.isdigit() else None

            telefono = request.POST.get('telefono') or None

            # Datos laborales
            puesto = request.POST.get('puesto')
            puesto_id = int(puesto) if puesto and puesto.isdigit() else None

            unidad_negocio = request.POST.get('unidadnegocio')
            unidad_negocio_id = int(unidad_negocio) if unidad_negocio and unidad_negocio.isdigit() else None

            sucursal = request.POST.get('sucursal')
            sucursal_id = int(sucursal) if sucursal and sucursal.isdigit() else None

            tipo_contrato = request.POST.get('tipocontrato')
            tipo_contrato_id = int(tipo_contrato) if tipo_contrato and tipo_contrato.isdigit() else None

            fecha_ingreso = request.POST.get('fechaIngreso') or None
            fecha_vencimiento = request.POST.get('fechavencimiento') or None

            departamento_empresa = request.POST.get('departamentoEmpresa')
            departamento_empresa_id = int(departamento_empresa) if departamento_empresa and departamento_empresa.isdigit() else None

            direccion_empresa = request.POST.get('direccionempresa') or None
            nombre_empresa = request.POST.get('nombreempresa')
            nombre_empresa_id = int(nombre_empresa) if nombre_empresa and nombre_empresa.isdigit() else None

            telefono_empresa = request.POST.get('telefonoempresa') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombreBeneficiario') or None
            identidad_beneficiario = request.POST.get('identidadBeneficiario') or None
            parentesco_beneficiario = request.POST.get('parentescoBeneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentajeBeneficiario') or None

            # Variables para imagen
            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Si ya existe una imagen, eliminarla
                if contratacion.ruta:  # Verifica si hay una imagen ya existente en la ruta
                    ruta_anterior = os.path.join(settings.BASE_DIR, contratacion.ruta)
                    if os.path.exists(ruta_anterior):
                        try:
                            os.remove(ruta_anterior)  # Elimina la imagen anterior
                        except Exception as e:
                            return JsonResponse({'success': False, 'message': f'Error al eliminar la imagen anterior: {str(e)}'}, status=500)

                # Generar un nombre único para la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/contrataciones/{nombre_imagen}"

                # Crear el directorio si no existe
                try:
                    os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                    # Guardar la nueva imagen en la ruta especificada
                    with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                        for chunk in imagen.chunks():
                            destino.write(chunk)

                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Error al guardar la nueva imagen: {str(e)}'}, status=500)

                # Actualizar la ruta y nombre de imagen en el modelo de contratación
                contratacion.ruta = ruta_imagen
                contratacion.nombreimagen = nombre_imagen
            else:
                # Si no se proporciona una nueva imagen, mantener los valores anteriores
                nombre_imagen = contratacion.nombreimagen
                ruta_imagen = contratacion.ruta


            # Valida "padecimiento"
            padecimiento = 1 if request.POST.get('padecimiento') == 'si' else 0

            # Actualizar la entrada existente en ContratacionEmpleados
            contratacion.tipo_contratacion = request.POST.get('tipoIngreso') or None
            contratacion.nombre1 = primer_nombre
            contratacion.nombre2 = segundo_nombre
            contratacion.apellido1 = primer_apellido
            contratacion.apellido2 = segundo_apellido
            contratacion.fecha_nacimiento = fecha_nacimiento
            contratacion.municipio_id = municipio_id
            contratacion.genero = genero
            contratacion.direccionexacta = direccion
            contratacion.dni = dni
            contratacion.estado_civil = estado_civil
            contratacion.hijos = hijos
            contratacion.profecion_oficio = profesion
            contratacion.correo = correo
            contratacion.departamento_id = departamento_id
            contratacion.telefono = telefono
            contratacion.nombre1_emergencia = request.POST.get('emergencia1') or None
            contratacion.parentesco1 = request.POST.get('parentesco1') or None
            contratacion.telefonoemergencia1 = request.POST.get('telefonoEmergencia1') or None
            contratacion.nombre2_emergencia = request.POST.get('emergencia2') or None
            contratacion.parentesco2 = request.POST.get('parentesco2') or None
            contratacion.telefonoemergencia2 = request.POST.get('telefonoEmergencia2') or None
            contratacion.nivel_educativo = request.POST.get('alfabeta') or None
            contratacion.ultimo_grado_estudio = request.POST.get('ultimoGrado') or None
            contratacion.ultimogradodetalle = request.POST.get('ultimogradodetalle') or None
            contratacion.padecimiento = padecimiento
            contratacion.detalle_enfermedad = request.POST.get('detalleEnfermedad') or None
            contratacion.puestos_id = puesto_id
            contratacion.unidad_de_negocio_id = unidad_negocio_id
            contratacion.sucursal_id = sucursal_id
            contratacion.departamento_empresa_id = departamento_empresa_id
            contratacion.tipo_contrato_id = tipo_contrato_id
            contratacion.salario = request.POST.get('salario') or None
            contratacion.comision = request.POST.get('comision') or None
            contratacion.bofa = request.POST.get('bofa') or None
            contratacion.fecha_ingreso = fecha_ingreso
            contratacion.fecha_vencimiento = fecha_vencimiento
            contratacion.direccionempresa = direccion_empresa
            contratacion.nombre_empresa_id = nombre_empresa_id
            contratacion.telefono_empresa = telefono_empresa
            contratacion.nombre_beneficiario = nombre_beneficiario
            contratacion.dni_beneficiario = identidad_beneficiario
            contratacion.parentesco_beneficiario = parentesco_beneficiario
            contratacion.porcentaje = porcentaje_beneficiario
            contratacion.ruta = ruta_imagen
            contratacion.nombreimagen = nombre_imagen

            # Guardar cambios en la base de datos
            contratacion.save()

            return JsonResponse({'success': True, 'message': 'Candidato actualizado exitosamente'}, status=200)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

class ExportExcelContratacion(View):
    def get(self, request, *args, **kwargs):
        template_path = os.path.join('reclutamiento/static/templates/FormatoContratacion.xlsx')
        
        # Cargar la plantilla de Excel
        try:
            wb = load_workbook(template_path) 
            sheet = wb.active  # Ajusta si la hoja tiene un nombre específico
        except Exception as e:
            return HttpResponse(f"Error al cargar la plantilla de Excel: {e}", status=500)

        # Obtener los datos de ContratacionEmpleados
        contrataciones = ContratacionEmpleados.objects.all().order_by('-id')
        
        # Definir la celda de inicio (C5)
        start_row = 5
        start_col = 3  # Columna C es la columna 3

        # Escribir los datos en la hoja de Excel
        for idx, contratacion in enumerate(contrataciones, start=start_row):
            try:
                sheet.cell(row=idx, column=start_col, value=contratacion.id)
                sheet.cell(row=idx, column=start_col + 1, value=contratacion.tipo_contratacion)
                sheet.cell(row=idx, column=start_col + 2, value=contratacion.nombre1)
                sheet.cell(row=idx, column=start_col + 3, value=contratacion.nombre2)
                sheet.cell(row=idx, column=start_col + 4, value=contratacion.apellido1)
                sheet.cell(row=idx, column=start_col + 5, value=contratacion.apellido2)
                sheet.cell(row=idx, column=start_col + 6, value=contratacion.fecha_nacimiento)
                
                # Departamento Honduras (asegúrate de que el campo en DepartamentoHonduras se llame 'nombre_departamento')
                sheet.cell(row=idx, column=start_col + 7, value=contratacion.departamento.nombre_departamentohonduras if contratacion.departamento else '')
                
                # Municipio Honduras (asegúrate de que el campo en MunicipioHonduras se llame 'nombre_municipio')
                sheet.cell(row=idx, column=start_col + 8, value=contratacion.municipio.nombre_municipio if contratacion.municipio else '')
                
                sheet.cell(row=idx, column=start_col + 9, value=contratacion.genero)
                sheet.cell(row=idx, column=start_col + 10, value=contratacion.dni)
                sheet.cell(row=idx, column=start_col + 11, value=contratacion.telefono)
                sheet.cell(row=idx, column=start_col + 12, value=contratacion.estado_civil)
                sheet.cell(row=idx, column=start_col + 13, value=contratacion.profecion_oficio)
                
                # Datos de contacto de emergencia
                sheet.cell(row=idx, column=start_col + 14, value=contratacion.nombre1_emergencia)
                sheet.cell(row=idx, column=start_col + 15, value=contratacion.parentesco1)
                sheet.cell(row=idx, column=start_col + 16, value=contratacion.telefonoemergencia1)
                sheet.cell(row=idx, column=start_col + 17, value=contratacion.nombre2_emergencia)
                sheet.cell(row=idx, column=start_col + 18, value=contratacion.parentesco2)
                sheet.cell(row=idx, column=start_col + 19, value=contratacion.telefonoemergencia2)
                
                # Dirección exacta, educación y salud
                sheet.cell(row=idx, column=start_col + 20, value=contratacion.direccionexacta)
                sheet.cell(row=idx, column=start_col + 21, value=contratacion.nivel_educativo)
                sheet.cell(row=idx, column=start_col + 22, value=contratacion.ultimo_grado_estudio)
                sheet.cell(row=idx, column=start_col + 23, value=contratacion.ultimogradodetalle)
                
                # Combinar padecimiento y detalle_enfermedad
                padecimiento_detalle = f"{'Si' if contratacion.padecimiento else 'No'} - {contratacion.detalle_enfermedad}"
                sheet.cell(row=idx, column=start_col + 24, value=padecimiento_detalle)
                
                # Puesto
                sheet.cell(row=idx, column=start_col + 25, value=contratacion.puestos.nombre_puestos if contratacion.puestos else '')
                sheet.cell(row=idx, column=start_col + 26, value=contratacion.salario)
                sheet.cell(row=idx, column=start_col + 27, value=contratacion.sucursal.nombre_sucursal if contratacion.sucursal else '')
                sheet.cell(row=idx, column=start_col + 28, value=contratacion.fecha_ingreso)
                
                # Beneficiario
                sheet.cell(row=idx, column=start_col + 29, value=contratacion.nombre_beneficiario)
                sheet.cell(row=idx, column=start_col + 30, value=contratacion.dni_beneficiario)
                sheet.cell(row=idx, column=start_col + 31, value=contratacion.parentesco_beneficiario)
                sheet.cell(row=idx, column=start_col + 32, value=contratacion.porcentaje)
                
                # Departamento empresa (Asegúrate de que el campo en Departamento sea correcto)
                sheet.cell(row=idx, column=start_col + 33, value=contratacion.departamento_empresa.nombre_departamento if contratacion.departamento_empresa else '')
                
                # Unidad de negocio
                sheet.cell(row=idx, column=start_col + 34, value=contratacion.unidad_de_negocio.nombre_unidad_de_negocio if contratacion.unidad_de_negocio else '')
                sheet.cell(row=idx, column=start_col + 35, value=contratacion.correo)
                
                # Tipo de contrato
                sheet.cell(row=idx, column=start_col + 36, value=contratacion.tipo_contrato.nombre_tipo_de_contrato if contratacion.tipo_contrato else '')
                
                # Información de la empresa
                sheet.cell(row=idx, column=start_col + 37, value=contratacion.telefono_empresa)
                sheet.cell(row=idx, column=start_col + 38, value=contratacion.direccionempresa)
                sheet.cell(row=idx, column=start_col + 39, value=contratacion.nombre_empresa.nombre_empresa if contratacion.nombre_empresa else '')
                
                # Hijos (Si o No)
                sheet.cell(row=idx, column=start_col + 40, value='Si' if contratacion.hijos else 'No')
                sheet.cell(row=idx, column=start_col + 41, value=contratacion.comision)
                sheet.cell(row=idx, column=start_col + 42, value=contratacion.bofa)
            except Exception as e:
                return HttpResponse(f"Error escribiendo datos en Excel en la fila {idx + 1}: {e}", status=500)

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="FormatoContratacion.xlsx"'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

#--------ANUNCIOS--------#  
def trabajaconnosotros_view(request):
    if request.method == 'GET':
        ciudades = Ciudades.objects.filter(estado='ACTIVO').order_by('nombre_ciudades')
        medios_reclutamiento = MedioReclutamiento.objects.filter(estado='ACTIVO').order_by('nombre_medio_de_reclutamiento')
        puestos = Puestos.objects.all()

        puestos_ids = set()
        for empleo in BolsaEmpleos.objects.all():
            puestos_ids.update(empleo.puestosaspira or [])
            puestos_ids.update(empleo.puestosaplica or [])

        puestos_en_bolsa = Puestos.objects.filter(id__in=puestos_ids)

        context = {
            'puestos': puestos,
            'puestos_en_bolsa': puestos_en_bolsa,
            'ciudades': ciudades,
            'medios_reclutamiento': medios_reclutamiento,
        }

        return render(request, 'anuncios/procesar_cv.html', context)
    if request.method == 'POST':
        try:
            # Obtener datos enviados en formato JSON
            data = json.loads(request.POST.get('data'))  # Datos JSON del formulario
            dni = data.get('dni')
            nombre_candidato = data.get('nombre_candidato')
            puestosaspira = data.get('puestoaspira', [])
            telefono = data.get('telefono')
            telefono2 = data.get('telefono2')
            estado = 'REGISTRADO'  # Estado automático
            ciudad_id = data.get('ciudad')
            medio_reclutamiento_id = data.get('mediosReclutamiento')
            edad = data.get('edad') or None  # Asignar None si está vacío
            fecha_nacimiento = data.get('fechanacimiento') or None  # Asignar None si está vacío
            estadocivil = data.get('estadocivil')
            nhijos = data.get('nhijos') or None  # Asignar None si está vacío
            direccion = data.get('direccion')
            experiencia = data.get('experiencia')

            # Añadir el prefijo "Aspiración Salarial: " a la observación
            observacion = f"Aspiración Salarial: {data.get('observacion', '')}"

            mediomovilizacion = data.get('mediomovilizacion')
            archivo = request.FILES.get('cv')  # Archivo PDF subido

            # Validación de campos obligatorios
            if not dni or not nombre_candidato or not telefono or not puestosaspira:
                return JsonResponse({'success': False, 'message': 'Los campos DNI, Nombre del Candidato, Teléfono y Puesto al que Aspira son obligatorios.'}, status=400)

            if BolsaEmpleos.objects.filter(dni=dni).exists():
                return JsonResponse({'success': False, 'message': 'El DNI que ingresaste ya está registrado en nuestro sistema.'}, status=400)

            # Validar que telefono y telefono2 no sean iguales entre sí
            if telefono == telefono2:
                return JsonResponse({'success': False, 'message': 'El teléfono principal y el secundario no pueden ser iguales.'}, status=400)

            # Validar que telefono y telefono2 no existan ya en la base de datos
            if BolsaEmpleos.objects.filter(telefono=telefono).exists() or BolsaEmpleos.objects.filter(telefono2=telefono).exists():
                return JsonResponse({'success': False, 'message': 'El número de teléfono principal ya está registrado.'}, status=400)

            if telefono2 and (BolsaEmpleos.objects.filter(telefono=telefono2).exists() or BolsaEmpleos.objects.filter(telefono2=telefono2).exists()):
                return JsonResponse({'success': False, 'message': 'El número de teléfono secundario ya está registrado.'}, status=400)

            # Validar formato de fecha
            if fecha_nacimiento:
                try:
                    fecha_nacimiento = timezone.datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'El formato de la fecha de nacimiento es inválido. Debe ser YYYY-MM-DD.'}, status=400)

            # Guardar archivo en ruta especificada con nombre único
            ruta = None
            nombredocumento = None
            if archivo:
                if archivo.content_type != 'application/pdf':
                    return JsonResponse({'success': False, 'message': 'Solo se permiten documentos en formato PDF.'}, status=400)
                
                # Generar nombre único para el archivo
                random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                nuevo_nombre_archivo = f"{random_str}_{archivo.name}"
                
                # Guardar archivo en reclutamiento/static/document
                ruta_relativa = os.path.join('reclutamiento', 'static', 'document', nuevo_nombre_archivo)
                ruta_archivo = os.path.join(settings.BASE_DIR, ruta_relativa)
                
                # Crear directorio si no existe
                os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)
                
                # Guardar el archivo en la ruta especificada
                with open(ruta_archivo, 'wb+') as destination:
                    for chunk in archivo.chunks():
                        destination.write(chunk)
                
                ruta = ruta_relativa
                nombredocumento = nuevo_nombre_archivo

            # Crear nueva entrada en BolsaEmpleos
            bolsa_empleo = BolsaEmpleos.objects.create(
                dni=dni,
                nombre_candidato=nombre_candidato,
                puestosaspira=puestosaspira,
                telefono=telefono,
                telefono2=telefono2,
                estado=estado,
                ciudad_id=ciudad_id,
                medio_reclutamiento_id=medio_reclutamiento_id,
                edad=edad,
                fecha_nacimiento=fecha_nacimiento,
                estadocivil=estadocivil,
                nhijos=nhijos,
                direccion=direccion,
                experiencia=experiencia,
                observacion=observacion,
                mediomovilizacion=mediomovilizacion,
                ruta=ruta,
                nombredocumento=nombredocumento
            )
            return JsonResponse({'success': True, 'message': 'Candidato registrado exitosamente'}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

#--------REQUISAS DE PUESTOS VIEW--------#
def requisaperfil_view(request, id):
    if request.method == 'GET':
        perfil_puesto = get_object_or_404(PerfilesPuestos, id=id)
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(estado='ACTIVO').order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(estado='ACTIVO').order_by('nombre_tipo_de_contrato').distinct()

        # Extraer solo las funciones del JSONField
        funciones_list = [funcion['funcion'] for funcion in perfil_puesto.funciones_cargo]
        funciones_text = ', '.join(funciones_list)  # Unir las funciones con comas

        context = {
            'perfil_puesto': perfil_puesto,
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'funciones_text': funciones_text,
            'all_tipos_contrato':all_tipos_contrato
        }

        return render(request, 'requisas/requisaperfil.html', context)
    elif request.method == 'POST':
        try:
            # Recibir los datos enviados desde el formulario con FormData
            salario_base = request.POST.get('salario_base') or None
            
            # Decodificar el JSON enviado con los valores de plan_de_compensacion y habilidades
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]'))
            montos = json.loads(request.POST.get('montos', '{}'))

            puesto = request.POST.get('puesto') or None
            sucursal_id = request.POST.get('sucursal') or None
            departamento_id = request.POST.get('departamento') or None
            funciones_cargo = request.POST.get('funciones_cargo') or None

            hora_inicio = request.POST.get('hora_inicio') or None
            hora_fin = request.POST.get('hora_fin') or None

            centrocostos = request.POST.get('centrocostos') or None
            puestonuevo = bool(int(request.POST.get('puestonuevo', 0)))  
            incapacidad = bool(int(request.POST.get('incapacidad', 0)))  
            reemplazo = bool(int(request.POST.get('reemplazo', 0)))

            tiempoprimercontrato = request.POST.get('tiempoprimercontrato') or None
            tipo_contrato_id = request.POST.get('tipo_contrato') or None
            motivo = request.POST.get('motivo') or None
            nombrereemplazar = request.POST.get('nombrereemplazar') or None
            formacionacademica = request.POST.get('formacionacademica') or None

            # Decodificar los JSON de habilidades
            habilidadesferreteras = json.loads(request.POST.get('habilidadesferreteras', '[]'))
            habilidadesinformaticas = json.loads(request.POST.get('habilidadesinformaticas', '[]'))
            habilidadespersonales = json.loads(request.POST.get('habilidadespersonales', '[]'))
            habilidadesanaliticas = json.loads(request.POST.get('habilidadesanaliticas', '[]'))
            materialesequipo = json.loads(request.POST.get('materialesequipo', '[]'))

            # Obtener instancias relacionadas
            sucursal = Sucursal.objects.get(id=sucursal_id) if sucursal_id else None
            departamento = Departamento.objects.get(id=departamento_id) if departamento_id else None
            tipo_contrato = TipoContrato.objects.get(id=tipo_contrato_id) if tipo_contrato_id else None

            # Crear la nueva Requisa sin fechaRecepcion
            nueva_requisa = Requisa.objects.create(
                salario_base=salario_base,
                plan_de_compensacion=plan_de_compensacion,
                montos=montos,
                puesto=puesto,
                sucursal=sucursal,
                departamento=departamento,
                hora_inicio=hora_inicio,
                hora_fin=hora_fin,
                centrocostos=centrocostos,
                puestonuevo=puestonuevo,
                incapacidad=incapacidad,
                reemplazo=reemplazo,
                tiempoprimercontrato=tiempoprimercontrato,
                tipo_contrato=tipo_contrato,
                motivo=motivo,
                nombrereemplazar=nombrereemplazar,
                formacionacademica=formacionacademica,
                habilidadesferreteras=habilidadesferreteras,
                habilidadesinformaticas=habilidadesinformaticas,
                habilidadespersonales=habilidadespersonales,
                habilidadesanaliticas=habilidadesanaliticas,
                materialesequipo=materialesequipo,
                funciones_cargo=funciones_cargo,
                fechaRecepcion=None,
                estado="EN PROCESO"
            )

            return JsonResponse({'success': True, 'message': 'Requisa registrada correctamente'})

        except Exception as e:
            # Si ocurre algún error, devolvemos una respuesta de error
            return JsonResponse({'success': False, 'message': str(e)})

def requisa_view(request):
    all_sucursales = Requisa.objects.values_list('sucursal__nombre_sucursal', flat=True).distinct()
    all_departamentos = Requisa.objects.values_list('departamento__nombre_departamento', flat=True).distinct()
    all_estados = Requisa.objects.values_list('estado', flat=True).distinct()
    all_fechas_creacion = Requisa.objects.annotate(fecha_creacion_dia=TruncDate('fechacreacion')).values_list('fecha_creacion_dia', flat=True).distinct()
    all_fechas_recepcion = Requisa.objects.annotate(fecha_recepcion_dia=TruncDate('fechaRecepcion')).values_list('fecha_recepcion_dia', flat=True).distinct()
    all_puestos = Requisa.objects.values_list('puesto', flat=True).distinct()

    search = request.GET.get('search', '')
    sucursales = request.GET.getlist('sucursal', [])
    departamentos = request.GET.getlist('departamento', [])
    estados = request.GET.getlist('estado', [])
    fechas_creacion = request.GET.getlist('fechacreacion', [])
    fechas_recepcion = request.GET.getlist('fecharecepcion', [])
    puestos = request.GET.getlist('puesto', [])

    requisas = Requisa.objects.all()

    if search:
        requisas = requisas.filter(
            Q(puesto__icontains=search) |
            Q(sucursal__nombre_sucursal__icontains=search) |
            Q(departamento__nombre_departamento__icontains=search) |
            Q(estado__icontains=search)
        )

    if sucursales:
        requisas = requisas.filter(sucursal__nombre_sucursal__in=sucursales)

    if departamentos:
        requisas = requisas.filter(departamento__nombre_departamento__in=departamentos)

    if estados:
        requisas = requisas.filter(estado__in=estados)

    if puestos:
        requisas = requisas.filter(puesto__in=puestos)

    if fechas_creacion:
        try:
            fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y').date() for f in fechas_creacion]
            requisas = requisas.filter(fechacreacion__date__in=fechas_seleccionadas)
        except ValueError:
            pass

    if fechas_recepcion:
        try:
            fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y').date() for f in fechas_recepcion]
            requisas = requisas.filter(fechaRecepcion__date__in=fechas_seleccionadas)
        except ValueError:
            pass

    paginator = Paginator(requisas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages
    current_page = page_obj.number
    start_page = max(1, current_page - 2)
    end_page = min(total_pages, current_page + 2)

    if current_page <= 3:
        end_page = min(5, total_pages)
    elif current_page >= total_pages - 2:
        start_page = max(1, total_pages - 4)

    page_range = range(start_page, end_page + 1)

    context = {
        'requisas': page_obj,
        'search': search,
        'page_range': page_range,
        'all_sucursales': all_sucursales,
        'all_departamentos': all_departamentos,
        'all_estados': all_estados,
        'all_fechas_creacion': all_fechas_creacion,
        'all_fechas_recepcion': all_fechas_recepcion,
        'all_puestos': all_puestos,
        'sucursales': sucursales,
        'departamentos': departamentos,
        'estados': estados,
        'fechas_creacion': fechas_creacion,
        'fechas_recepcion': fechas_recepcion,
        'puestos': puestos,
    }

    return render(request, 'requisas/requisa.html', context)

def requisaregistrar_view(request):
    if request.method == 'GET':
        # Cargar el formulario
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(estado='ACTIVO').order_by('nombre_tipo_de_contrato').distinct()

        context = {
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'all_tipos_contrato': all_tipos_contrato
        }

        return render(request, 'requisas/requisaregistrar.html', context)

    elif request.method == 'POST':
        try:
            # Recibir los datos enviados desde el formulario con FormData
            salario_base = request.POST.get('salario_base') or None
            
            # Decodificar el JSON enviado con los valores de plan_de_compensacion y habilidades
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]'))
            montos = json.loads(request.POST.get('montos', '{}'))

            puesto = request.POST.get('puesto') or None
            sucursal_id = request.POST.get('sucursal') or None
            departamento_id = request.POST.get('departamento') or None
            funciones_cargo = request.POST.get('funciones_cargo') or None

            hora_inicio = request.POST.get('hora_inicio') or None
            hora_fin = request.POST.get('hora_fin') or None

            centrocostos = request.POST.get('centrocostos') or None
            puestonuevo = bool(int(request.POST.get('puestonuevo', 0)))  
            incapacidad = bool(int(request.POST.get('incapacidad', 0)))  
            reemplazo = bool(int(request.POST.get('reemplazo', 0)))

            tiempoprimercontrato = request.POST.get('tiempoprimercontrato') or None
            tipo_contrato_id = request.POST.get('tipo_contrato') or None
            motivo = request.POST.get('motivo') or None
            nombrereemplazar = request.POST.get('nombrereemplazar') or None
            formacionacademica = request.POST.get('formacionacademica') or None

            # Decodificar los JSON de habilidades
            habilidadesferreteras = json.loads(request.POST.get('habilidadesferreteras', '[]'))
            habilidadesinformaticas = json.loads(request.POST.get('habilidadesinformaticas', '[]'))
            habilidadespersonales = json.loads(request.POST.get('habilidadespersonales', '[]'))
            habilidadesanaliticas = json.loads(request.POST.get('habilidadesanaliticas', '[]'))
            materialesequipo = json.loads(request.POST.get('materialesequipo', '[]'))

            # Obtener instancias relacionadas
            sucursal = Sucursal.objects.get(id=sucursal_id) if sucursal_id else None
            departamento = Departamento.objects.get(id=departamento_id) if departamento_id else None
            tipo_contrato = TipoContrato.objects.get(id=tipo_contrato_id) if tipo_contrato_id else None

            # Crear la nueva Requisa sin fechaRecepcion
            nueva_requisa = Requisa.objects.create(
                salario_base=salario_base,
                plan_de_compensacion=plan_de_compensacion,
                montos=montos,
                puesto=puesto,
                sucursal=sucursal,
                departamento=departamento,
                hora_inicio=hora_inicio,
                hora_fin=hora_fin,
                centrocostos=centrocostos,
                puestonuevo=puestonuevo,
                incapacidad=incapacidad,
                reemplazo=reemplazo,
                tiempoprimercontrato=tiempoprimercontrato,
                tipo_contrato=tipo_contrato,
                motivo=motivo,
                nombrereemplazar=nombrereemplazar,
                formacionacademica=formacionacademica,
                habilidadesferreteras=habilidadesferreteras,
                habilidadesinformaticas=habilidadesinformaticas,
                habilidadespersonales=habilidadespersonales,
                habilidadesanaliticas=habilidadesanaliticas,
                materialesequipo=materialesequipo,
                funciones_cargo=funciones_cargo,
                fechaRecepcion=None,
                estado="EN PROCESO"
            )

            return JsonResponse({'success': True, 'message': 'Requisa registrada correctamente'})

        except Exception as e:
            # Si ocurre algún error, devolvemos una respuesta de error
            return JsonResponse({'success': False, 'message': str(e)})

def aprobar_requisa_view(request, id):
    if request.method == 'POST':
        try:
            requisa = get_object_or_404(Requisa, id=id)
            # Cambiar el estado de la requisición a "APROBADO"
            requisa.estado = 'APROBADO'
            requisa.save()
            
            return JsonResponse({'success': True, 'message': 'Requisa aprobada correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

def cancelar_requisa_view(request, id):
    if request.method == 'POST':
        try:
            requisa = get_object_or_404(Requisa, id=id)
            # Cambiar el estado de la requisición a "CANCELADA"
            requisa.estado = 'CANCELADA'
            requisa.save()
            
            return JsonResponse({'success': True, 'message': 'Requisa cancelada correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

def duplicar_requisa_view(request, id):
    if request.method == 'POST':
        try:
            # Obtener la requisición original
            requisa_original = get_object_or_404(Requisa, id=id)
            
            # Crear una nueva requisición duplicando la original (omitimos el campo id)
            nueva_requisa = Requisa.objects.create(
                salario_base=requisa_original.salario_base,
                plan_de_compensacion=requisa_original.plan_de_compensacion,
                montos=requisa_original.montos,
                puesto=requisa_original.puesto,
                sucursal=requisa_original.sucursal,
                departamento=requisa_original.departamento,
                hora_inicio=requisa_original.hora_inicio,
                hora_fin=requisa_original.hora_fin,
                centrocostos=requisa_original.centrocostos,
                puestonuevo=requisa_original.puestonuevo,
                incapacidad=requisa_original.incapacidad,
                reemplazo=requisa_original.reemplazo,
                tiempoprimercontrato=requisa_original.tiempoprimercontrato,
                tipo_contrato=requisa_original.tipo_contrato,
                motivo=requisa_original.motivo,
                nombrereemplazar=requisa_original.nombrereemplazar,
                formacionacademica=requisa_original.formacionacademica,
                habilidadesferreteras=requisa_original.habilidadesferreteras,
                habilidadesinformaticas=requisa_original.habilidadesinformaticas,
                habilidadespersonales=requisa_original.habilidadespersonales,
                habilidadesanaliticas=requisa_original.habilidadesanaliticas,
                materialesequipo=requisa_original.materialesequipo,
                funciones_cargo=requisa_original.funciones_cargo,
                estado="EN PROCESO"  # Establecer el estado de la nueva requisición
            )
            
            return JsonResponse({'success': True, 'message': 'Requisa duplicada correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
        
def requisaupdate_view(request, id):
    if request.method == 'GET':
        requisa = get_object_or_404(Requisa, id=id)
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(estado='ACTIVO').order_by('nombre_tipo_de_contrato').distinct()

        context = {
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'all_tipos_contrato': all_tipos_contrato,
            'requisa': requisa,
        }
        return render(request, 'requisas/updaterequisa.html', context)

    if request.method == 'POST':
        try:
            requisa = get_object_or_404(Requisa, id=id)

            # Recibir los datos enviados desde el formulario con FormData
            salario_base = request.POST.get('salario_base')
            fecharecepcion = request.POST.get('fecharecepcion')

            # Decodificar el JSON enviado con los valores de plan_de_compensacion y montos
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]'))
            montos = json.loads(request.POST.get('montos', '{}'))

            puesto = request.POST.get('puesto')
            sucursal_id = request.POST.get('sucursal')
            departamento_id = request.POST.get('departamento')
            funciones_cargo = request.POST.get('funciones_cargo')

            hora_inicio = request.POST.get('hora_inicio')
            hora_fin = request.POST.get('hora_fin')

            centrocostos = request.POST.get('centrocostos')
            puestonuevo = bool(int(request.POST.get('puestonuevo', 0)))
            incapacidad = bool(int(request.POST.get('incapacidad', 0)))
            reemplazo = bool(int(request.POST.get('reemplazo', 0)))

            tiempoprimercontrato = request.POST.get('tiempoprimercontrato')
            tipo_contrato_id = request.POST.get('tipo_contrato')
            motivo = request.POST.get('motivo')
            nombrereemplazar = request.POST.get('nombrereemplazar')
            formacionacademica = request.POST.get('formacionacademica')

            # Decodificar los JSON de habilidades
            habilidadesferreteras = json.loads(request.POST.get('habilidadesferreteras', '[]'))
            habilidadesinformaticas = json.loads(request.POST.get('habilidadesinformaticas', '[]'))
            habilidadespersonales = json.loads(request.POST.get('habilidadespersonales', '[]'))
            habilidadesanaliticas = json.loads(request.POST.get('habilidadesanaliticas', '[]'))
            materialesequipo = json.loads(request.POST.get('materialesequipo', '[]'))

            # Actualizar los datos de la Requisa
            requisa.salario_base = salario_base
            requisa.fechaRecepcion = fecharecepcion if fecharecepcion else None
            requisa.plan_de_compensacion = plan_de_compensacion
            requisa.montos = montos
            requisa.puesto = puesto
            requisa.sucursal_id = sucursal_id if sucursal_id else None
            requisa.departamento_id = departamento_id if departamento_id else None
            requisa.hora_inicio = hora_inicio
            requisa.hora_fin = hora_fin
            requisa.centrocostos = centrocostos
            requisa.puestonuevo = puestonuevo
            requisa.incapacidad = incapacidad
            requisa.reemplazo = reemplazo
            requisa.tiempoprimercontrato = tiempoprimercontrato
            requisa.tipo_contrato_id = tipo_contrato_id if tipo_contrato_id else None
            requisa.motivo = motivo
            requisa.nombrereemplazar = nombrereemplazar
            requisa.formacionacademica = formacionacademica
            requisa.habilidadesferreteras = habilidadesferreteras
            requisa.habilidadesinformaticas = habilidadesinformaticas
            requisa.habilidadespersonales = habilidadespersonales
            requisa.habilidadesanaliticas = habilidadesanaliticas
            requisa.materialesequipo = materialesequipo
            requisa.funciones_cargo = funciones_cargo

            # Guardar la requisa actualizada
            requisa.save()

            return JsonResponse({'success': True, 'message': 'Requisa actualizada correctamente'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

def imprimirrequisa_view(request, id):
    if request.method == 'GET':
        requisa = get_object_or_404(Requisa, id=id)
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_tipos_contrato = TipoContrato.objects.filter(estado='ACTIVO').order_by('nombre_tipo_de_contrato').distinct()

        context = {
            'all_departamentos': all_departamentos,
            'all_sucursales': all_sucursales,
            'all_tipos_contrato': all_tipos_contrato,
            'requisa': requisa,
        }
        return render(request, 'requisas/imprimirrequisa.html', context)

#--------PERFILES DE PUESTOS VIEW--------#
def perfilpuesto_view(request):
    # Obtener todos los valores únicos para los filtros (tipo_perfil, empresa, departamento, etc.)
    all_tipos_perfil = PerfilesPuestos.objects.values_list('tipo_perfil', flat=True).distinct()
    all_empresas = PerfilesPuestos.objects.values_list('empresa', flat=True).distinct()
    all_departamentos = PerfilesPuestos.objects.values_list('departamento', flat=True).distinct()
    all_nombres_cargo = PerfilesPuestos.objects.values_list('nombre_cargo', flat=True).distinct()
    all_fechas_actualizacion = PerfilesPuestos.objects.annotate(fecha_actualizacion_dia=TruncDate('fechaactualizacion')).values_list('fecha_actualizacion_dia', flat=True).distinct()

    # Obtener parámetros de búsqueda y filtros desde el formulario
    search = request.GET.get('search', '')  # Búsqueda general
    tipos_perfil = request.GET.getlist('tipo_perfil', [])  # Lista de tipos de perfil seleccionados
    empresas = request.GET.getlist('empresa', [])  # Lista de empresas seleccionadas
    departamentos = request.GET.getlist('departamento', [])  # Lista de departamentos seleccionados
    nombres_cargo = request.GET.getlist('nombre_cargo', [])  # Lista de nombres de cargo seleccionados
    fechas_actualizacion = request.GET.getlist('fecha_actualizacion', [])  # Lista de fechas de actualización seleccionadas

    # Query inicial con todos los perfiles de puesto
    perfiles_puesto = PerfilesPuestos.objects.all()

    # Filtrar por búsqueda general
    if search:
        perfiles_puesto = perfiles_puesto.filter(
            Q(nombre_cargo__icontains=search) |
            Q(tipo_perfil__icontains=search) |
            Q(empresa__icontains=search) |
            Q(departamento__icontains=search)
        )

    # Filtrar por tipo de perfil si hay algún filtro aplicado
    if tipos_perfil:
        perfiles_puesto = perfiles_puesto.filter(tipo_perfil__in=tipos_perfil)

    # Filtrar por empresa si hay algún filtro aplicado
    if empresas:
        perfiles_puesto = perfiles_puesto.filter(empresa__in=empresas)

    # Filtrar por departamento si hay algún filtro aplicado
    if departamentos:
        perfiles_puesto = perfiles_puesto.filter(departamento__in=departamentos)

    # Filtrar por nombre de cargo si hay algún filtro aplicado
    if nombres_cargo:
        perfiles_puesto = perfiles_puesto.filter(nombre_cargo__in=nombres_cargo)

    # Filtrar por fecha de actualización si hay algún filtro aplicado
    if fechas_actualizacion:
        try:
            fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y %H:%M') for f in fechas_actualizacion]
            perfiles_puesto = perfiles_puesto.filter(fechaactualizacion__in=fechas_seleccionadas)
        except ValueError:
            pass  # Si alguna fecha es inválida, ignorar el filtro

    # Paginación: mostrar 10 registros por página
    paginator = Paginator(perfiles_puesto, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calcular el rango de páginas a mostrar en la paginación
    total_pages = paginator.num_pages
    current_page = page_obj.number
    start_page = max(1, current_page - 2)
    end_page = min(total_pages, current_page + 2)

    if current_page <= 3:
        end_page = min(5, total_pages)
    elif current_page >= total_pages - 2:
        start_page = max(1, total_pages - 4)

    page_range = range(start_page, end_page + 1)

    # Contexto para renderizar la plantilla
    context = {
        'perfiles_puesto': page_obj,
        'search': search,
        'page_range': page_range,
        'all_tipos_perfil': all_tipos_perfil,
        'all_empresas': all_empresas,
        'all_departamentos': all_departamentos,
        'all_nombres_cargo': all_nombres_cargo,
        'all_fechas_actualizacion': all_fechas_actualizacion,
        'tipos_perfil': tipos_perfil,
        'empresas': empresas,
        'departamentos': departamentos,
        'nombres_cargo': nombres_cargo,
        'fechas_actualizacion': fechas_actualizacion,
    }

    return render(request, 'perfilpuesto/perfilpuesto.html', context)

def obtener_pruebas_por_nivel(request):
    if request.method == 'GET':
        nivel = request.GET.get('nivel') 
        if nivel:
            pruebas = Spicosmart.objects.filter(nivel=nivel, estado='ACTIVO')
            pruebas_data = list(pruebas.values('nombre_prueba'))  # Obtener solo el nombre de las pruebas
            return JsonResponse({'success': True, 'pruebas': pruebas_data}, status=200)
        else:
            return JsonResponse({'success': False, 'message': 'Nivel no proporcionado'}, status=400)
        
def perfilpuestoregister_view(request):
    if request.method == 'GET':
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(estado='ACTIVO').order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()

        context = {
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos
        }
        return render(request, 'perfilpuesto/perfilpuestoregistro.html', context)

    elif request.method == 'POST':
        try:
            # Obtener datos del formulario, asignando null si el valor es vacío
            tipo_perfil = request.POST.get('tipo_perfil') or None
            empresa = request.POST.get('empresa') or None
            nombre_cargo = request.POST.get('nombre_cargo') or None
            departamento = request.POST.get('departamento') or None
            cargo_al_que_reporta = request.POST.get('cargo_al_que_reporta') or None
            cargo_que_le_reportan = request.POST.get('cargos_que_le_reportan') or None
            educacion_universitario = request.POST.get('educacion_universitario') or None
            postgrado_especializaciones = request.POST.get('postgrado_especializaciones') or None
            formacion_complementaria = request.POST.get('formacion_complementaria') or None
            idiomas = request.POST.get('idiomas') or None
            nivel_idioma = request.POST.get('nivel_idioma') or None
            anos_experiencia = request.POST.get('anos_experiencia') or None
            pensamiento_estrategico = request.POST.get('pensamiento_estrategico') or None
            enfoque_al_cliente = request.POST.get('enfoque_al_cliente') or None
            planificacion_y_organizacion = request.POST.get('planificacion_y_organizacion') or None
            comunicacion = request.POST.get('comunicacion') or None
            orientacion_al_logro = request.POST.get('orientacion_al_logro') or None
            mision_cargo = request.POST.get('mision_cargo') or None

            # Captura del campo de 'residir en el área', asignando null si está vacío
            residir_en_area = request.POST.get('residir_en_area') or None

            # Guardar imagen del organigrama si existe
            imagen = request.FILES.get('ruta_organigrama')
            nombre_imagen = None
            ruta_imagen = None
            if imagen:
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/organigrama/{nombre_imagen}"

                os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                # Guardar el archivo en la ruta correspondiente
                with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                    for chunk in imagen.chunks():
                        destino.write(chunk)

            # Capturar los campos dinámicos JSON (funciones, retos, materiales, etc.), con valores por defecto como lista vacía o null
            funciones_cargo = json.loads(request.POST.get('funciones_cargo', '[]') or '[]')
            retos = json.loads(request.POST.get('retos', '[]') or '[]')  
            materiales_equipos = json.loads(request.POST.get('materiales_equipos', '[]') or '[]')
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]') or '[]')
            beneficiospromaco = json.loads(request.POST.get('beneficiospromaco', '[]') or '[]')
            psicometricas = json.loads(request.POST.get('psicometricas', '[]') or '[]')
            principales_indicadores = json.loads(request.POST.get('principales_indicadores', '[]') or '[]')

            # Captura de otros campos, asignando null si están vacíos
            horario_turnos = request.POST.get('horario_turnos') or None
            otros = request.POST.get('otros') or None
            nivel_prueba = request.POST.get('nivel_prueba') or None
            montacargas = request.POST.get('montacargas') or None
            esquipo_pegado = request.POST.get('esquipo_pegado') or None

            # Crear nuevo perfil de puesto con todos los datos capturados, enviando null cuando corresponda
            perfil = PerfilesPuestos.objects.create(
                tipo_perfil=tipo_perfil,
                empresa=empresa,
                nombre_cargo=nombre_cargo,
                departamento=departamento,
                cargo_al_que_reporta=cargo_al_que_reporta,
                cargo_que_le_reportan=cargo_que_le_reportan,
                educacion_universitario=educacion_universitario,
                postgrado_especializaciones=postgrado_especializaciones,
                formacion_complementaria=formacion_complementaria,
                idiomas=idiomas,
                nivel_idioma=nivel_idioma,
                anos_experiencia=anos_experiencia,
                pensamiento_estrategico=pensamiento_estrategico,
                enfoque_al_cliente=enfoque_al_cliente,
                planificacion_y_organizacion=planificacion_y_organizacion,
                comunicacion=comunicacion,
                orientacion_al_logro=orientacion_al_logro,
                residir_en_area=residir_en_area,
                mision_cargo=mision_cargo,
                ruta_organigrama=ruta_imagen,
                nombre_organigrama=nombre_imagen,
                funciones_cargo=funciones_cargo,
                retos=retos,
                materiales_equipos=materiales_equipos,
                horario_turnos=horario_turnos,
                otros=otros,
                plan_de_compensacion=plan_de_compensacion,
                beneficiospromaco=beneficiospromaco,
                nivel_prueba=nivel_prueba,
                psicometricas=psicometricas,
                montacargas=montacargas,
                esquipo_pegado=esquipo_pegado,
                principales_indicadores=principales_indicadores
            )

            # Devolver una respuesta de éxito
            return JsonResponse({'success': True, 'message': 'Perfil de puesto registrado correctamente.'})

        except Exception as e:
            # Depuración de errores y POST data
            print(f"Datos POST: {request.POST}")
            print(f"Error: {str(e)}")  # Para depurar el error exacto
            return JsonResponse({'success': False, 'message': f'Error al procesar la solicitud: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

def imprimir_perfilpuesto_view(request, perfil_id):
    perfil_puesto = get_object_or_404(PerfilesPuestos, id=perfil_id)
    
    context = {
        'perfil_puesto': perfil_puesto,
    }

    # Renderiza la plantilla para la impresión
    return render(request, 'perfilpuesto/imprimirperfilpuesto.html', context)

def imprimir_perfilpuestocorto_view(request, perfil_id):
    perfil_puesto = get_object_or_404(PerfilesPuestos, id=perfil_id)
    
    context = {
        'perfil_puesto': perfil_puesto,
    }

    # Renderiza la plantilla para la impresión
    return render(request, 'perfilpuesto/imprimirperfilpuestocorto.html', context)

def update_perfilpuesto_view(request, id):
    if request.method == 'GET':
        perfil_puesto = get_object_or_404(PerfilesPuestos, id=id)
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(estado='ACTIVO').order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()

        # Valores predefinidos para materiales
        predefined_materials = ["COMPUTADOR", "CELULAR", "UNIFORME", "ESCRITORIO", "LICENCIA SAP", "LICENCIA IVEND", "SILLA", "CORREO"]
        otros_materiales = [material for material in perfil_puesto.materiales_equipos if material not in predefined_materials]
        otros_materiales = otros_materiales[0] if otros_materiales else ''

        # Valores predefinidos para plan de compensación y beneficios
        predefined_compensacion = ["SALARIO BASE", "COMISION", "BONO", "DEPRECIACION Y COMBUSTIBLE", "HOSPEDAJE"]
        predefined_beneficios = ["SEGURO MEDICO PRIVADO", "SEGURO SOCIAL", "PERMISO DE ESTUDIO", "BONO SUPERMERCADO", 
                                "BONO CAFE PROMACO", "DESCUENTO EN SUPERMERCADO", "DESCUENTO EN CLINICA Y LAVORATORIOS", 
                                "SEGURO DE VEHICULO"]

        # Limpiar espacios extra en plan de compensación y filtrar "Otros"
        plan_de_compensacion_limpio = [comp.strip() for comp in perfil_puesto.plan_de_compensacion]
        otros_compensacion = [comp for comp in plan_de_compensacion_limpio if comp not in predefined_compensacion]
        otros_compensacion = otros_compensacion[0] if otros_compensacion else ''

        context = {
            'perfil_puesto': perfil_puesto,
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos,
            'otros_materiales': otros_materiales,
            'otros_compensacion': otros_compensacion,  # Para el campo "Otros" en plan de compensación
            'predefined_materials': predefined_materials,  # Para comparar en el template
            'predefined_compensacion': predefined_compensacion,  # Para comparar en el template
            'predefined_beneficios': predefined_beneficios  # Para comparar en el template
        }

        return render(request, 'perfilpuesto/updateperfilpuesto.html', context)
    
    elif request.method == 'POST':
        try:
            # Obtener el perfil de puesto existente
            perfil = get_object_or_404(PerfilesPuestos, id=id)
            
            # Obtener datos del formulario, asignando None si el valor es vacío
            tipo_perfil = request.POST.get('tipo_perfil') or None
            empresa = request.POST.get('empresa') or None
            nombre_cargo = request.POST.get('nombre_cargo') or None
            departamento = request.POST.get('departamento') or None
            cargo_al_que_reporta = request.POST.get('cargo_al_que_reporta') or None
            cargo_que_le_reportan = request.POST.get('cargos_que_le_reportan') or None
            educacion_universitario = request.POST.get('educacion_universitario') or None
            postgrado_especializaciones = request.POST.get('postgrado_especializaciones') or None
            formacion_complementaria = request.POST.get('formacion_complementaria') or None
            idiomas = request.POST.get('idiomas') or None
            nivel_idioma = request.POST.get('nivel_idioma') or None
            anos_experiencia = request.POST.get('anos_experiencia') or None
            pensamiento_estrategico = request.POST.get('pensamiento_estrategico') or None
            enfoque_al_cliente = request.POST.get('enfoque_al_cliente') or None
            planificacion_y_organizacion = request.POST.get('planificacion_y_organizacion') or None
            comunicacion = request.POST.get('comunicacion') or None
            orientacion_al_logro = request.POST.get('orientacion_al_logro') or None
            mision_cargo = request.POST.get('mision_cargo') or None

            # Captura del campo de 'residir en el área', asignando None si está vacío
            residir_en_area = request.POST.get('residir_en_area') or None

            # Variables para imagen del organigrama
            imagen = request.FILES.get('ruta_organigrama')
            nombre_imagen = perfil.nombre_organigrama  # Mantener el nombre existente si no hay imagen nueva
            ruta_imagen = perfil.ruta_organigrama  # Mantener la ruta existente si no hay imagen nueva

            if imagen:
                # Si ya existe una imagen, eliminar la imagen anterior
                if perfil.ruta_organigrama:
                    ruta_anterior = os.path.join(settings.BASE_DIR, perfil.ruta_organigrama)
                    if os.path.exists(ruta_anterior):
                        try:
                            os.remove(ruta_anterior)  # Elimina la imagen anterior
                        except Exception as e:
                            return JsonResponse({'success': False, 'message': f'Error al eliminar la imagen anterior: {str(e)}'}, status=500)

                # Generar un nombre único para la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/organigrama/{nombre_imagen}"

                # Crear el directorio si no existe
                os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                # Guardar el archivo en la ruta correspondiente
                with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                    for chunk in imagen.chunks():
                        destino.write(chunk)

            # Capturar los campos dinámicos JSON (funciones, retos, materiales, etc.), con valores por defecto como lista vacía o None
            funciones_cargo = json.loads(request.POST.get('funciones_cargo', '[]') or '[]')
            retos = json.loads(request.POST.get('retos', '[]') or '[]')  
            materiales_equipos = json.loads(request.POST.get('materiales_equipos', '[]') or '[]')
            plan_de_compensacion = json.loads(request.POST.get('plan_de_compensacion', '[]') or '[]')
            beneficiospromaco = json.loads(request.POST.get('beneficiospromaco', '[]') or '[]')
            psicometricas = json.loads(request.POST.get('psicometricas', '[]') or '[]')
            principales_indicadores = json.loads(request.POST.get('principales_indicadores', '[]') or '[]')

            # Captura de otros campos, asignando None si están vacíos
            horario_turnos = request.POST.get('horario_turnos') or None
            otros = request.POST.get('otros') or None
            nivel_prueba = request.POST.get('nivel_prueba') or None
            montacargas = request.POST.get('montacargas') or None
            esquipo_pegado = request.POST.get('esquipo_pegado') or None

            # Asignar los valores al perfil existente
            perfil.tipo_perfil = tipo_perfil
            perfil.empresa = empresa
            perfil.nombre_cargo = nombre_cargo
            perfil.departamento = departamento
            perfil.cargo_al_que_reporta = cargo_al_que_reporta
            perfil.cargo_que_le_reportan = cargo_que_le_reportan
            perfil.educacion_universitario = educacion_universitario
            perfil.postgrado_especializaciones = postgrado_especializaciones
            perfil.formacion_complementaria = formacion_complementaria
            perfil.idiomas = idiomas
            perfil.nivel_idioma = nivel_idioma
            perfil.anos_experiencia = anos_experiencia
            perfil.pensamiento_estrategico = pensamiento_estrategico
            perfil.enfoque_al_cliente = enfoque_al_cliente
            perfil.planificacion_y_organizacion = planificacion_y_organizacion
            perfil.comunicacion = comunicacion
            perfil.orientacion_al_logro = orientacion_al_logro
            perfil.residir_en_area = residir_en_area
            perfil.mision_cargo = mision_cargo
            perfil.ruta_organigrama = ruta_imagen
            perfil.nombre_organigrama = nombre_imagen
            perfil.funciones_cargo = funciones_cargo
            perfil.retos = retos
            perfil.materiales_equipos = materiales_equipos
            perfil.horario_turnos = horario_turnos
            perfil.otros = otros
            perfil.plan_de_compensacion = plan_de_compensacion
            perfil.beneficiospromaco = beneficiospromaco
            perfil.nivel_prueba = nivel_prueba
            perfil.psicometricas = psicometricas
            perfil.montacargas = montacargas
            perfil.esquipo_pegado = esquipo_pegado
            perfil.principales_indicadores = principales_indicadores

            # Guardar el perfil actualizado
            perfil.save()

            # Devolver una respuesta de éxito
            return JsonResponse({'success': True, 'message': 'Perfil de puesto actualizado correctamente.'})

        except Exception as e:
            # Depuración de errores y POST data
            print(f"Datos POST: {request.POST}")
            print(f"Error: {str(e)}")  # Para depurar el error exacto
            return JsonResponse({'success': False, 'message': f'Error al procesar la solicitud: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

def updatecompleto_perfilpuesto_view(request, id):
    if request.method == 'GET':
        perfil_puesto = get_object_or_404(PerfilesPuestos, id=id)
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_puestos = Puestos.objects.filter(estado='ACTIVO').order_by('nombre_puestos').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()

        # Valores predefinidos para materiales
        predefined_materials = ["COMPUTADOR", "CELULAR", "UNIFORME", "ESCRITORIO", "LICENCIA SAP", "LICENCIA IVEND", "SILLA", "CORREO"]
        otros_materiales = [material for material in perfil_puesto.materiales_equipos if material not in predefined_materials]
        otros_materiales = otros_materiales[0] if otros_materiales else ''

        # Valores predefinidos para plan de compensación y beneficios
        predefined_compensacion = ["SALARIO BASE", "COMISION", "BONO", "DEPRECIACION Y COMBUSTIBLE", "HOSPEDAJE"]
        predefined_beneficios = ["SEGURO MEDICO PRIVADO", "SEGURO SOCIAL", "PERMISO DE ESTUDIO", "BONO SUPERMERCADO", 
                                "BONO CAFE PROMACO", "DESCUENTO EN SUPERMERCADO", "DESCUENTO EN CLINICA Y LAVORATORIOS", 
                                "SEGURO DE VEHICULO"]

        # Limpiar espacios extra en plan de compensación y filtrar "Otros"
        plan_de_compensacion_limpio = [comp.strip() for comp in perfil_puesto.plan_de_compensacion]
        otros_compensacion = [comp for comp in plan_de_compensacion_limpio if comp not in predefined_compensacion]
        otros_compensacion = otros_compensacion[0] if otros_compensacion else ''

        context = {
            'perfil_puesto': perfil_puesto,
            'all_empresas': all_empresas,
            'all_puestos': all_puestos,
            'all_departamentos': all_departamentos,
            'otros_materiales': otros_materiales,
            'otros_compensacion': otros_compensacion,  # Para el campo "Otros" en plan de compensación
            'predefined_materials': predefined_materials,  # Para comparar en el template
            'predefined_compensacion': predefined_compensacion,  # Para comparar en el template
            'predefined_beneficios': predefined_beneficios  # Para comparar en el template
        }

        return render(request, 'perfilpuesto/updatetotalperfilpuesto.html', context)

#--------CONTRATACIONES MULTI--------#  
def contratacionesform_multi_view(request):
    if request.method == 'GET':
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()


        all_municipios = MunicipioHonduras.objects.all().order_by('nombre_municipio').distinct()
        all_departamentos_hn = DepartamentoHonduras.objects.all().order_by('nombre_departamentohonduras').distinct()

        context = {
            'all_sucursales': all_sucursales,
            'all_municipios': all_municipios, 
            'all_departamentos_hn': all_departamentos_hn,
        }
        return render(request, 'contrataciones_multi/registrarcontrataciones_multi.html', context)
    if request.method == 'POST':
        try:
            # Obtener campos individuales y asignar None si están vacíos
            tipo_contratacion = request.POST.get('tipoIngreso') or None
            primer_nombre = request.POST.get('primerNombre') or None
            segundo_nombre = request.POST.get('segundoNombre') or None
            primer_apellido = request.POST.get('primerApellido') or None
            segundo_apellido = request.POST.get('segundoApellido') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None
            municipio = request.POST.get('municipio') or None
            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccion') or None
            dni = request.POST.get('cedula') or None
            estado_civil = request.POST.get('estadoCivil') or None
            profesion = request.POST.get('profesion') or None
            departamento = request.POST.get('departamento') or None
            telefono = request.POST.get('telefono') or None
            ultimo_grado_estudio = request.POST.get('ultimoGrado') or None

            # Datos de emergencia
            emergencia1 = request.POST.get('emergencia1') or None
            parentesco1 = request.POST.get('parentesco1') or None
            telefono_emergencia1 = request.POST.get('telefonoEmergencia1') or None
            enfermedad = request.POST.get('enfermedad') or None

            # Datos laborales
            puesto = request.POST.get('puesto') or None
            sucursal = request.POST.get('sucursal') or None
            salario = request.POST.get('salario') or None
            fecha_ingreso = request.POST.get('fechaIngreso') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombreBeneficiario') or None
            identidad_beneficiario = request.POST.get('identidadBeneficiario') or None
            parentesco_beneficiario = request.POST.get('parentescoBeneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentajeBeneficiario') or None

            # Imagen
            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Si hay imagen, generar un nombre único para el archivo
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/contrataciones/contrataciones_multi/{nombre_imagen}"

                # Crear directorio si no existe
                try:
                    os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)

                    # Guardar la imagen en la ruta especificada
                    with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                        for chunk in imagen.chunks():
                            destino.write(chunk)
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'Error al guardar la imagen: {str(e)}'}, status=500)

            # Crear nueva entrada en ContratacionEmpleadosmulti
            contratacion = ContratacionEmpleadosmulti.objects.create(
                tipo_contratacion=tipo_contratacion,
                nombre1=primer_nombre,
                nombre2=segundo_nombre,
                apellido1=primer_apellido,
                apellido2=segundo_apellido,
                fecha_nacimiento=fecha_nacimiento,
                municipio_id=municipio,
                genero=genero,
                direccionexacta=direccion,
                dni=dni,
                estado_civil=estado_civil,
                profecion_oficio=profesion,
                departamento_id=departamento,
                telefono=telefono,
                nombre_emergencia=emergencia1,
                parentesco=parentesco1,
                telefonoemergencia=telefono_emergencia1,
                enfermedad=enfermedad,
                puestos=puesto,
                sucursal_id=sucursal,
                salario=salario,
                fecha_ingreso=fecha_ingreso,
                ultimo_grado_estudio=ultimo_grado_estudio,
                nombre_beneficiario=nombre_beneficiario,
                dni_beneficiario=identidad_beneficiario,
                parentesco_beneficiario=parentesco_beneficiario,
                porcentaje=porcentaje_beneficiario,
                ruta=ruta_imagen,
                nombreimagen=nombre_imagen,
            )

            return JsonResponse({'success': True, 'message': 'Candidato registrado exitosamente'}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
                          
def contratacionesmulti_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')

        # Consulta los datos del nuevo modelo ContratacionEmpleadosmulti
        contrataciones_multi = ContratacionEmpleadosmulti.objects.select_related(
            'sucursal', 
            'municipio',  # Relaciona con el modelo Municipio para el filtro de ciudad
        )

        # Obtener valores únicos para los filtros
        municipios_options = set(contrataciones_multi.values_list('municipio__nombre_municipio', flat=True))
        puestos_options = set(contrataciones_multi.values_list('puestos', flat=True))
        sucursales_options = set(contrataciones_multi.values_list('sucursal__nombre_sucursal', flat=True))
        fechas_ingreso_options = set(contrataciones_multi.values_list('fecha_ingreso', flat=True))

        # Filtrar por los filtros seleccionados (si se aplican)
        municipio_filter = request.GET.getlist('municipio')
        puesto_filter = request.GET.getlist('puesto')
        sucursal_filter = request.GET.getlist('sucursal')
        fecha_ingreso_filter = request.GET.getlist('fecha_ingreso')

        # Aplicar filtros a la consulta solo si existen
        if municipio_filter:
            contrataciones_multi = contrataciones_multi.filter(municipio__nombre_municipio__in=municipio_filter)
        
        if puesto_filter:
            contrataciones_multi = contrataciones_multi.filter(puestos__in=puesto_filter)
        
        if sucursal_filter:
            contrataciones_multi = contrataciones_multi.filter(sucursal__nombre_sucursal__in=sucursal_filter)

        if fecha_ingreso_filter:
            contrataciones_multi = contrataciones_multi.filter(fecha_ingreso__in=fecha_ingreso_filter)

        # Filtro de búsqueda general
        if search:
            contrataciones_multi = contrataciones_multi.filter(
                models.Q(nombre1__icontains=search) |
                models.Q(nombre2__icontains=search) |
                models.Q(apellido1__icontains=search) |
                models.Q(apellido2__icontains=search) |
                models.Q(telefono__icontains=search) |
                models.Q(dni__icontains=search) |
                models.Q(correo__icontains=search)
            )

        # Paginación
        paginator = Paginator(contrataciones_multi, 10)  # 10 registros por página
        page_number = request.GET.get('page')  # Capturar el número de página actual
        page_obj = paginator.get_page(page_number)  # Obtener la página solicitada

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)  # Crear un rango de páginas a mostrar

        # Contexto para la plantilla
        context = {
            'contrataciones': page_obj,  # Pasar el objeto paginado
            'search': search,  # Pasar el valor de búsqueda si aplica
            'page_range': page_range,  # Pasar el rango de páginas
            'municipios_options': list(municipios_options),
            'puestos_options': list(puestos_options),
            'sucursales_options': list(sucursales_options),
            'fechas_ingreso_options': list(fechas_ingreso_options),
            'municipio_filter': municipio_filter,
            'puesto_filter': puesto_filter,
            'sucursal_filter': sucursal_filter,
            'fecha_ingreso_filter': fecha_ingreso_filter,
        }

        return render(request, 'contrataciones_multi/contrataciones_multi.html', context)
    if request.method == 'DELETE':
        if id:
            contratacion = get_object_or_404(ContratacionEmpleadosmulti, id=id)

            if contratacion.ruta:
                ruta_imagen = os.path.join(settings.BASE_DIR, contratacion.ruta)

                if os.path.exists(ruta_imagen):
                    os.remove(ruta_imagen)

            contratacion.delete()

            return JsonResponse({'success': True, 'message': 'Contratación eliminada con éxito.'})
        return JsonResponse({'success': False, 'message': 'ID de contratación no encontrado.'})

def imprimir_contratacion_multi_view(request, contratacion_id):
    contratacion_multi = get_object_or_404(ContratacionEmpleadosmulti, id=contratacion_id)
    
    context = {
        'contratacion_multi': contratacion_multi,
    }
    return render(request, 'contrataciones_multi/imprimir_contratacion_multi.html', context)

def updatecontrataciones_multi_view(request, contratacion_id):
    contratacion_multi = get_object_or_404(ContratacionEmpleadosmulti, id=contratacion_id)
    if request.method == 'GET':
        context = {
            'contratacion_multi': contratacion_multi,
            'all_municipios': MunicipioHonduras.objects.all(),
            'all_departamentos': DepartamentoHonduras.objects.all(),
            'all_sucursales': Sucursal.objects.all(),
        }
        return render(request, 'contrataciones_multi/updatecontrataciones_multi.html', context)

    if request.method == 'POST':
        try:
            primer_nombre = request.POST.get('nombre1') or None
            segundo_nombre = request.POST.get('nombre2') or None
            primer_apellido = request.POST.get('apellido1') or None
            segundo_apellido = request.POST.get('apellido2') or None
            fecha_nacimiento = request.POST.get('fechaNacimiento') or None

            # Claves foráneas deben ser números o None
            municipio = request.POST.get('municipio')
            municipio_id = int(municipio) if municipio and municipio.isdigit() else None

            genero = request.POST.get('genero') or None
            direccion = request.POST.get('direccionexacta') or None
            dni = request.POST.get('dni') or None
            estado_civil = request.POST.get('estadoCivil') or None
            profesion = request.POST.get('profecion_oficio') or None
            telefono = request.POST.get('telefono') or None

            # Datos laborales
            puesto = request.POST.get('puestos')
            sucursal = request.POST.get('sucursal')
            sucursal_id = int(sucursal) if sucursal and sucursal.isdigit() else None
            salario = request.POST.get('salario') or None
            fecha_ingreso = request.POST.get('fechaIngreso') or None

            # Datos de emergencia
            nombre_emergencia = request.POST.get('emergencia1') or None
            parentesco_emergencia = request.POST.get('parentesco1') or None
            telefono_emergencia = request.POST.get('telefonoEmergencia1') or None

            # Beneficiario
            nombre_beneficiario = request.POST.get('nombre_beneficiario') or None
            identidad_beneficiario = request.POST.get('dni_beneficiario') or None
            parentesco_beneficiario = request.POST.get('parentesco_beneficiario') or None
            porcentaje_beneficiario = request.POST.get('porcentaje') or None

            # Imagen
            imagen = request.FILES.get('imagen')
            nombre_imagen = None
            ruta_imagen = None

            if imagen:
                # Verificar si existe una imagen previa en el campo `ruta`
                if contratacion_multi.ruta:
                    ruta_anterior = os.path.join(settings.BASE_DIR, contratacion_multi.ruta)
                    
                    # Comprobar si la ruta anterior realmente existe en el sistema de archivos
                    if os.path.exists(ruta_anterior):
                        try:
                            os.remove(ruta_anterior)  # Eliminar la imagen anterior
                            print(f"Imagen anterior eliminada: {ruta_anterior}")  # Log para depuración
                        except Exception as e:
                            print(f"Error al eliminar la imagen anterior: {str(e)}")
                            return JsonResponse({'success': False, 'message': f'Error al eliminar la imagen anterior: {str(e)}'}, status=500)
                    else:
                        print(f"La ruta de la imagen anterior no existe: {ruta_anterior}")  # Log para depuración

                # Guardar la nueva imagen
                nombre_imagen = f"{uuid.uuid4().hex[:4]}_{imagen.name}"
                ruta_imagen = f"reclutamiento/static/img/contrataciones/contrataciones_multi/{nombre_imagen}"
                os.makedirs(os.path.dirname(os.path.join(settings.BASE_DIR, ruta_imagen)), exist_ok=True)
                
                # Guardar el archivo en la nueva ruta
                with open(os.path.join(settings.BASE_DIR, ruta_imagen), 'wb+') as destino:
                    for chunk in imagen.chunks():
                        destino.write(chunk)

                # Actualizar el modelo con la ruta y el nombre de la nueva imagen
                contratacion_multi.ruta = ruta_imagen
                contratacion_multi.nombreimagen = nombre_imagen

            # Actualizar el modelo con otros campos
            contratacion_multi.nombre1 = primer_nombre
            contratacion_multi.nombre2 = segundo_nombre
            contratacion_multi.apellido1 = primer_apellido
            contratacion_multi.apellido2 = segundo_apellido
            contratacion_multi.fecha_nacimiento = fecha_nacimiento
            contratacion_multi.municipio_id = municipio_id
            contratacion_multi.genero = genero
            contratacion_multi.direccionexacta = direccion
            contratacion_multi.dni = dni
            contratacion_multi.estado_civil = estado_civil
            contratacion_multi.profecion_oficio = profesion
            contratacion_multi.telefono = telefono
            contratacion_multi.puestos = puesto
            contratacion_multi.sucursal_id = sucursal_id
            contratacion_multi.fecha_ingreso = fecha_ingreso
            contratacion_multi.salario = salario

            # Datos de emergencia
            contratacion_multi.nombre_emergencia = nombre_emergencia
            contratacion_multi.parentesco = parentesco_emergencia
            contratacion_multi.telefonoemergencia = telefono_emergencia

            # Beneficiario
            contratacion_multi.nombre_beneficiario = nombre_beneficiario
            contratacion_multi.dni_beneficiario = identidad_beneficiario
            contratacion_multi.parentesco_beneficiario = parentesco_beneficiario
            contratacion_multi.porcentaje = porcentaje_beneficiario

            # Guardar los cambios
            contratacion_multi.save()

            return JsonResponse({'success': True, 'message': 'Candidato actualizado exitosamente'}, status=200)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

class ExportExcelContratacionMulti(View):
    def get(self, request, *args, **kwargs):
        template_path = os.path.join('reclutamiento/static/templates/FormatoContratacion_multi.xlsx')
        
        # Cargar la plantilla de Excel
        try:
            wb = load_workbook(template_path)
            sheet = wb.active  # Ajusta si la hoja tiene un nombre específico
        except Exception as e:
            return HttpResponse(f"Error al cargar la plantilla de Excel: {e}", status=500)

        # Obtener los datos de ContratacionEmpleadosmulti
        contrataciones = ContratacionEmpleadosmulti.objects.all().order_by('-id')
        
        # Definir la celda de inicio (C5)
        start_row = 5
        start_col = 3  # Columna C es la columna 3

        # Escribir los datos en la hoja de Excel
        for idx, contratacion in enumerate(contrataciones, start=start_row):
            try:
                sheet.cell(row=idx, column=start_col, value=contratacion.id)
                sheet.cell(row=idx, column=start_col + 1, value=contratacion.tipo_contratacion)
                sheet.cell(row=idx, column=start_col + 2, value=contratacion.nombre1)
                sheet.cell(row=idx, column=start_col + 3, value=contratacion.nombre2)
                sheet.cell(row=idx, column=start_col + 4, value=contratacion.apellido1)
                sheet.cell(row=idx, column=start_col + 5, value=contratacion.apellido2)
                sheet.cell(row=idx, column=start_col + 6, value=contratacion.fecha_nacimiento)
                
                # Departamento Honduras
                sheet.cell(row=idx, column=start_col + 7, value=contratacion.departamento.nombre_departamentohonduras if contratacion.departamento else '')
                
                # Municipio Honduras
                sheet.cell(row=idx, column=start_col + 8, value=contratacion.municipio.nombre_municipio if contratacion.municipio else '')
                
                sheet.cell(row=idx, column=start_col + 9, value=contratacion.genero)
                sheet.cell(row=idx, column=start_col + 10, value=contratacion.dni)
                sheet.cell(row=idx, column=start_col + 11, value=contratacion.telefono)
                sheet.cell(row=idx, column=start_col + 12, value=contratacion.profecion_oficio)
                sheet.cell(row=idx, column=start_col + 13, value=contratacion.estado_civil)

                
                # Datos de contacto de emergencia
                sheet.cell(row=idx, column=start_col + 14, value=contratacion.nombre_emergencia)
                sheet.cell(row=idx, column=start_col + 15, value=contratacion.parentesco)
                sheet.cell(row=idx, column=start_col + 16, value=contratacion.telefonoemergencia)
                
                # Enfermedad
                sheet.cell(row=idx, column=start_col + 17, value=contratacion.enfermedad)
                
                # Dirección exacta y último grado de estudio
                sheet.cell(row=idx, column=start_col + 18, value=contratacion.direccionexacta)
                sheet.cell(row=idx, column=start_col + 19, value=contratacion.ultimo_grado_estudio)
                
                # Puestos y salario
                sheet.cell(row=idx, column=start_col + 20, value=contratacion.puestos)
                sheet.cell(row=idx, column=start_col + 21, value=contratacion.salario)
                
                # Sucursal y fecha de ingreso
                sheet.cell(row=idx, column=start_col + 22, value=contratacion.sucursal.nombre_sucursal if contratacion.sucursal else '')
                sheet.cell(row=idx, column=start_col + 23, value=contratacion.fecha_ingreso)
                
                # Beneficiario
                sheet.cell(row=idx, column=start_col + 24, value=contratacion.nombre_beneficiario)
                sheet.cell(row=idx, column=start_col + 25, value=contratacion.dni_beneficiario)
                sheet.cell(row=idx, column=start_col + 26, value=contratacion.parentesco_beneficiario)
                sheet.cell(row=idx, column=start_col + 27, value=contratacion.porcentaje)
            except Exception as e:
                return HttpResponse(f"Error escribiendo datos en Excel en la fila {idx + 1}: {e}", status=500)

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="FormatoContratacion_multi.xlsx"'

        # Guardar el libro de trabajo en la respuesta
        wb.save(response)

        return response

#--------CESANTIAs VIEW--------#
@csrf_exempt
def cesantias_view(request, id=None):
    if request.method == 'GET':
        # Obtener todos los registros de estados, años, autorizadores y fechas de extensión
        all_estados_pago = Cesantias.objects.values_list('estado_pago', flat=True).distinct()
        all_estados_empleado = Cesantias.objects.values_list('estado_empleado', flat=True).distinct()
        all_años = Cesantias.objects.values_list('año', flat=True).distinct()
        all_autorizadores = Cesantias.objects.values_list('nombre_autoriza', flat=True).distinct()
        all_fechas_extencion = Cesantias.objects.values_list('fecha_extencion', flat=True).distinct()
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()



        # Obtener parámetros de búsqueda y filtros desde el formulario
        search = request.GET.get('search', '')
        estados_pago = request.GET.getlist('estado_pago', [])  # Lista de estados de pago seleccionados
        estados_empleado = request.GET.getlist('estado_empleado', [])  # Lista de estados de empleado seleccionados
        años = request.GET.getlist('año', [])  # Lista de años seleccionados
        autorizadores = request.GET.getlist('nombre_autoriza', [])  # Lista de nombres de autorizadores seleccionados
        fechas_extencion = request.GET.getlist('fecha_extencion', [])  # Lista de fechas de extensión seleccionadas

        # Query inicial con todas las cesantías
        cesantias = Cesantias.objects.all()

        # Filtrar por búsqueda general (nombre autoriza, nombre empleado, correlativo, año)
        if search:
            try:
                search_date = datetime.strptime(search, '%d/%m/%Y').date()
                cesantias = cesantias.filter(Q(fecha_extencion=search_date))
            except ValueError:
                cesantias = cesantias.filter(
                    Q(nombre_autoriza__icontains=search) |
                    Q(nombre_empleado__icontains=search) |
                    Q(correlativo__icontains=search) |
                    Q(año__icontains=search)
                )

        # Filtrar por estado de pago si hay algún filtro aplicado
        if estados_pago:
            cesantias = cesantias.filter(estado_pago__in=estados_pago)

        # Filtrar por estado de empleado si hay algún filtro aplicado
        if estados_empleado:
            cesantias = cesantias.filter(estado_empleado__in=estados_empleado)

        # Filtrar por año si hay algún filtro aplicado
        if años:
            cesantias = cesantias.filter(año__in=años)

        # Filtrar por nombre del autorizador si hay algún filtro aplicado
        if autorizadores:
            cesantias = cesantias.filter(nombre_autoriza__in=autorizadores)

        # Filtrar por fechas de extensión si hay algún filtro aplicado
        if fechas_extencion:
            try:
                fechas_seleccionadas = [datetime.strptime(f, '%d/%m/%Y').date() for f in fechas_extencion]
                cesantias = cesantias.filter(fecha_extencion__in=fechas_seleccionadas)
            except ValueError:
                pass  # Si alguna fecha es inválida, ignorar el filtro

        # Paginación: mostrar 10 registros por página
        paginator = Paginator(cesantias, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar en la paginación
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Contexto para renderizar la plantilla
        context = {
            'cesantias': page_obj,
            'all_sucursales':all_sucursales,
            'all_empresas':all_empresas,
            'all_departamentos':all_departamentos,
            'search': search,
            'page_range': page_range,
            'all_estados_pago': all_estados_pago,
            'all_estados_empleado': all_estados_empleado,
            'all_años': all_años,
            'all_autorizadores': all_autorizadores,
            'all_fechas_extencion': all_fechas_extencion,  # Fechas de extensión disponibles
            'estados_pago': estados_pago,
            'estados_empleado': estados_empleado,
            'años': años,  # Años seleccionados
            'autorizadores': autorizadores,
            'fechas_extencion': fechas_extencion  # Fechas de extensión seleccionadas
        }

        return render(request, 'cesantias/cesantias.html', context)
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Crear nueva instancia de Cesantias con los datos del formulario
            cesantia = Cesantias(
                correlativo=data['correlativo'],
                nombre_autoriza=data['nombre_autoriza'],
                dni_autoriza=data['dni_autoriza'],
                cargo_autoriza=data['cargo_autoriza'],
                empresa=Empresas.objects.get(id=data['empresa']),
                sucursal=Sucursal.objects.get(id=data['sucursal']),
                departamento=Departamento.objects.get(id=data['departamento']),
                nombre_empleado=data['nombre_empleado'],
                dni_empleado=data['dni_empleado'],
                fecha_inicial=data['fecha_inicial'],
                fecha_final=data['fecha_final'],
                fecha_extencion=data['fecha_extencion'],
                sueldo_actual=data['sueldo_actual'],
                cesantia_actual=data['cesantia_actual'],
                cesantia_final=data['cesantia_final'],
                porcentaje=data['porcentaje'],
                año=data['año'],
                estado_empleado=data['estado_empleado'],
                estado_pago=data['estado_pago'],
                n_cheke=data.get('n_cheke', None)  # Puede no estar presente en el payload
            )

            # Guardar la nueva cesantía
            cesantia.save()

            # Responder con éxito
            return JsonResponse({'success': True, 'message': 'Cesantía registrada correctamente.'})

        except Exception as e:
            # En caso de error, devolver una respuesta con el error
            return JsonResponse({'success': False, 'message': str(e)})
        
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            # Buscar el registro a actualizar usando el id
            cesantia = get_object_or_404(Cesantias, id=id)

            # Actualizar los campos
            cesantia.nombre_autoriza = data['nombre_autoriza']
            cesantia.dni_autoriza = data['dni_autoriza']
            cesantia.cargo_autoriza = data['cargo_autoriza']
            cesantia.sucursal_id = data['sucursal']
            cesantia.empresa_id = data['empresa']
            cesantia.departamento_id = data['departamento']
            cesantia.nombre_empleado = data['nombre_empleado']
            cesantia.dni_empleado = data['dni_empleado']
            cesantia.fecha_inicial = data['fecha_inicial']
            cesantia.fecha_final = data['fecha_final']
            cesantia.fecha_extencion = data['fecha_extencion']
            cesantia.sueldo_actual = data['sueldo_actual']
            cesantia.cesantia_actual = data['cesantia_actual']
            cesantia.cesantia_final = data['cesantia_final']
            cesantia.porcentaje = data['porcentaje']
            cesantia.año = data['año']
            cesantia.estado_empleado = data['estado_empleado']
            cesantia.estado_pago = data['estado_pago']
            cesantia.n_cheke = data['n_cheke']

            # Guardar los cambios
            cesantia.save()

            return JsonResponse({'success': True, 'message': 'Cesantía actualizada correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Método no permitido.'})

def numero_a_letras_con_decimales(numero):
    parte_entera = int(numero)
    parte_decimal = int(round((numero - parte_entera) * 100))

    parte_entera_letras = num2words(parte_entera, lang='es').upper()

    if parte_decimal > 0:
        parte_decimal_letras = num2words(parte_decimal, lang='es').upper()
        return f"{parte_entera_letras} PUNTO {parte_decimal_letras}"
    else:
        return parte_entera_letras

def imprimircesantias_view(request, id):
    cesantia = get_object_or_404(Cesantias, id=id)

    # Convertir a palabras usando la nueva función para manejar decimales correctamente
    cesantia_letras = numero_a_letras_con_decimales(cesantia.cesantia_actual)
    cesantia_letras_final = numero_a_letras_con_decimales(cesantia.cesantia_final)
    
    # Formatear números con separador de miles y dos decimales
    cesantia_actual_formateado = "{:,.2f}".format(cesantia.cesantia_actual)
    cesantia_final_formateado = "{:,.2f}".format(cesantia.cesantia_final)
    sueldo_actual_formateado = "{:,.2f}".format(cesantia.sueldo_actual)

    context = {
        'cesantia': cesantia,
        'cesantia_letras': cesantia_letras,
        'cesantia_letras_final': cesantia_letras_final,
        'cesantia_actual_formateado': cesantia_actual_formateado,
        'cesantia_final_formateado': cesantia_final_formateado,
        'sueldo_actual_formateado': sueldo_actual_formateado
    }
    
    return render(request, 'cesantias/imprimircesantias.html', context)

def exportar_cesantias(request):
    datos = (
        Cesantias.objects
        .select_related('empresa', 'sucursal', 'departamento') 
        .values(
            'fecha_extencion', 
            'empresa__nombre_empresa', 
            'sucursal__nombre_sucursal',  
            'nombre_empleado', 
            'dni_empleado',
            'departamento__nombre_departamento',
            'año', 
            'cesantia_final'
        )
    )

    # Convertir los datos a un DataFrame de pandas
    df = pd.DataFrame(list(datos))
    
    # Crear un archivo Excel
    excel_file = "cesantias.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={excel_file}'

    # Escribir el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Cesantias')

    return response
#--------TELEFONIA--------#
def telefonosinventario_view(request, id=None):
    if request.method == 'GET':
        all_nombres_telefono = Inventariotelefonos.objects.values_list('nombretelefono', flat=True).distinct()

        search = request.GET.get('search', '')
        nombres_telefono = request.GET.getlist('nombretelefono', [])

        # Obtener todos los inventarios
        inventarios = Inventariotelefonos.objects.all()

        # Filtrar por búsqueda
        if search:
            inventarios = inventarios.filter(
                Q(nombretelefono__icontains=search) |
                Q(correlativo__icontains=search) 
            )

        # Filtrar por nombres de teléfono seleccionados
        if nombres_telefono:
            inventarios = inventarios.filter(nombretelefono__in=nombres_telefono)

        # Formatear el valor total con comas para miles y punto para decimales
        for inventario in inventarios:
            inventario.valortotal_formateado = "{:,.2f}".format(inventario.valortotal)

        # Paginación
        paginator = Paginator(inventarios, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'inventarios': page_obj,
            'search': search,
            'page_range': page_range,
            'all_nombres_telefono': all_nombres_telefono,
            'nombres_telefono': nombres_telefono,
        }

        # Crear una lista de inventarios para el script JSON
        context['inventarios_json'] = [
            {
                'id': inventario.id,
                'correlativo': inventario.correlativo,
                'nombretelefono': inventario.nombretelefono,
                'valortotal': inventario.valortotal,
                'estadotelefono': inventario.estadotelefono,
                'estado': inventario.estado,
            }
            for inventario in inventarios
        ]

        return render(request, 'inventario/inventariotelefono.html', context)


    elif request.method == 'POST':
        try:
            # Parsear los datos JSON recibidos desde el fetch
            data = json.loads(request.body)
            nombre_telefono = data.get('nombretelefono')
            valor_telefono = float(data.get('valortelefono', 0))
            cantidad = int(data.get('cantidad', 1))  # Cantidad de registros a crear
            estado = data.get('estado', 'REGISTRADO')
            estadotelefono = data.get('estadotelefono')

            # Validación básica
            if not nombre_telefono or valor_telefono <= 0 or cantidad <= 0:
                return JsonResponse({'success': False, 'message': 'Datos no válidos.'}, status=400)
            
            # Crear los registros según la cantidad especificada
            for _ in range(cantidad):
                Inventariotelefonos.objects.create(
                    nombretelefono=nombre_telefono,
                    estado=estado,
                    valortotal=valor_telefono,
                    estadotelefono = estadotelefono
                )
            
            return JsonResponse({'success': True, 'message': f'{cantidad} teléfonos registrados exitosamente.'}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error al procesar la solicitud.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        
    elif request.method == 'DELETE':
        if id:
            telefono = get_object_or_404(Inventariotelefonos, id=id)
            telefono.delete()
            return JsonResponse({'success': True, 'message': 'Teléfono eliminado con éxito.'})

        return JsonResponse({'success': False, 'message': 'ID de teléfono no encontrado.'})
        
    elif request.method == 'PUT':
        telefono_id = id  # Asegúrate de que el ID se pase correctamente a la vista
        telefono = get_object_or_404(Inventariotelefonos, id=telefono_id)

        # Obtener los datos del cuerpo de la solicitud
        try:
            data = json.loads(request.body)
            valortelefono = data.get('valortelefono')
            estadotelefono = data.get('estadotelefono')
            estado = data.get('estado')

            # Actualizar los campos del objeto
            if valortelefono is not None:
                telefono.valortotal = float(valortelefono)  # Asegúrate de convertir a float

            if estadotelefono is not None:
                telefono.estadotelefono = estadotelefono

            if estado is not None:
                # Si el nuevo estado es "REGISTRADO" y el anterior era "ASIGNADO" o "REASIGNADO", actualizar correlativo a null
                if estado == "REGISTRADO" and telefono.estado in ["ASIGNADO", "REASIGNADO"]:
                    telefono.correlativo = None  # Establecer correlativo a null
                telefono.estado = estado

            # Guardar los cambios en la base de datos
            telefono.save()

            return JsonResponse({'success': True, 'message': 'Teléfono actualizado con éxito.'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error al decodificar los datos JSON.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Método no permitido.'})

def obtener_id_telefono(request):
    nombre_telefono = request.GET.get('nombretelefono', None)
    estadotelefono = request.GET.get('estadotelefono', None)  # Nuevo parámetro de estado de teléfono

    if nombre_telefono and estadotelefono:
        # Filtrar por nombre, estado 'REGISTRADO' y estado del teléfono ('NUEVO' o 'USADO')
        telefono = (Inventariotelefonos.objects
                    .filter(nombretelefono=nombre_telefono, estado='REGISTRADO', estadotelefono=estadotelefono)
                    .order_by('id')
                    .first())

        if telefono:
            return JsonResponse({'success': True, 'id': telefono.id})
        else:
            return JsonResponse({'success': False, 'message': 'No se encontró el teléfono en el inventario.'}, status=404)

    return JsonResponse({'success': False, 'message': 'Parámetros nombretelefono o estadotelefono no proporcionados.'}, status=400)

def telefonos_view(request):
    if request.method == 'GET':
        all_unidades_negocio = Unidad_Negocio.objects.filter(estado='ACTIVO').order_by('nombre_unidad_de_negocio').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_telefonosinventario = (Inventariotelefonos.objects
                                .filter(estado='REGISTRADO')
                                .values('nombretelefono', 'estadotelefono')
                                .distinct()
                                .order_by('nombretelefono', 'estadotelefono'))

        # Search and filter functionality
        search = request.GET.get('search', '')
        telefonia = Telefonia.objects.all()

        if search:
            try:
                # Intentar parsear la fecha si se busca en el formato día/mes/año
                search_date = datetime.strptime(search, '%d/%m/%Y').date()
                telefonia = telefonia.filter(fecha=search_date)
            except ValueError:
                # Si no es una fecha válida, buscar en los demás campos
                telefonia = telefonia.filter(
                    Q(correlativo__icontains=search) |
                    Q(nombre__icontains=search) |
                    Q(dni__icontains=search) |
                    Q(unidad_de_negocio__nombre_unidad_de_negocio__icontains=search) |
                    Q(departamento__nombre_departamento__icontains=search) |
                    Q(nombretelefono__icontains=search)
                )

        # Paginación
        paginator = Paginator(telefonia, 10)  # 10 registros por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Rango de paginación
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'all_unidades_negocio': all_unidades_negocio,
            'all_departamentos': all_departamentos,
            'all_telefonosinventario': all_telefonosinventario,
            'telefonia': page_obj,  # Pasamos los datos paginados
            'search': search,
            'page_range': page_range
        }
        return render(request, 'inventario/telefonos.html', context)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            correlativo = data.get('correlativo')  # Captura el correlativo
            fecha = data.get('fecha')
            nombre = data.get('nombre')
            dni = data.get('dni')
            MEI = data.get('MEI')
            unidad_negocio_id = data.get('unidadnegocio')
            departamento_id = data.get('departamento')
            telefono_id = data.get('telefono') 
            lineatelefonica = data.get('lineatelefono')
            caracter = data.get('caracter')
            tiempopago = data.get('tiempopago')
            quincenas = data.get('quincenas')
            observacion = data.get('observacion')

            # Obtén el objeto `Inventariotelefonos` para el `telefono_id`
            inventario_telefono1 = Inventariotelefonos.objects.filter(id=telefono_id, estado='REGISTRADO').first()
            
            # Verifica que el teléfono existe en el inventario y tiene el estado `REGISTRADO`
            if not inventario_telefono1:
                return JsonResponse({'success': False, 'message': 'No se encontró el teléfono en el inventario.'}, status=404)
            
            valor_total = inventario_telefono1.valortotal
            estado_telefono = inventario_telefono1.estadotelefono  # Obtiene el `estadotelefono` desde `Inventariotelefonos`

            # Validación de campos obligatorios
            if not nombre or not dni or not unidad_negocio_id or not departamento_id or not telefono_id or not caracter:
                return JsonResponse({'success': False, 'message': 'Faltan campos obligatorios.'}, status=400)
            
            # Crear el registro en `Telefonia` con `estadotelefono`
            telefonia = Telefonia.objects.create(
                correlativo=correlativo, 
                fecha=fecha,
                nombre=nombre,
                dni=dni,
                unidad_de_negocio_id=unidad_negocio_id,
                departamento_id=departamento_id,
                nombretelefono=telefono_id,
                MEI=MEI,
                lineatelefonica=lineatelefonica,
                caracter=caracter,
                valor=valor_total,
                tiempopago=tiempopago,
                quinsena=quincenas,
                estado='ASIGNADO',
                observacion=observacion,
                estadotelefono=estado_telefono  # Asigna el `estadotelefono` al campo correspondiente en `Telefonia`
            )

            # Actualizar el estado en `Inventariotelefonos`
            inventario_telefono1.estado = 'ASIGNADO'
            inventario_telefono1.correlativo = correlativo
            inventario_telefono1.save()

            return JsonResponse({'success': True, 'message': 'Registro de telefonía y actualización de inventario exitosos.'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error en los datos enviados.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

def imprimir_telefono_view(request, id):
    telefonia = get_object_or_404(Telefonia, id=id)
    inventario_telefono = get_object_or_404(Inventariotelefonos, id=telefonia.nombretelefono)

    telefonia.valor_formateado = "{:,.2f}".format(telefonia.valor) if telefonia.valor else None
    if telefonia.quinsena:
        # Verificar si el número es entero o tiene decimales
        if telefonia.quinsena.is_integer():
            telefonia.quincena_formateada = "{:,.0f}".format(telefonia.quinsena)  # Sin decimales
        else:
            telefonia.quincena_formateada = "{:,.2f}".format(telefonia.quinsena)  # Con dos decimales
    else:
        telefonia.quincena_formateada = None

    context = {
        'telefonia': telefonia,
        'nombre_telefono': inventario_telefono.nombretelefono,
    }
    return render(request, 'inventario/impimirtelefonos.html', context)

@csrf_exempt
def telefonosreasignar_view(request, correlativo):
    if request.method == 'PUT':
        print(f"Buscando correlativo: {correlativo}")  
        try:
            data = json.loads(request.body.decode('utf-8'))

            telefonia = get_object_or_404(Telefonia, correlativo=correlativo)

            # Actualizar los campos en Telefonia
            telefonia.asignacionnueva = data.get('asignacionnueva')
            telefonia.estado = 'REASIGNADO'
            telefonia.save()

            # Buscar el registro en Inventariotelefonos por correlativo
            inventario_telefono = Inventariotelefonos.objects.filter(correlativo=correlativo, estado='ASIGNADO').first()

            if inventario_telefono:
                inventario_telefono.estado = 'REASIGNADO'
                inventario_telefono.save()
            else:
                return JsonResponse({'success': False, 'message': 'No se encontró el teléfono en el inventario.'}, status=404)

            return JsonResponse({'success': True, 'message': 'Registro reasignado correctamente'})

        except Telefonia.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Correlativo no encontrado'}, status=404)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

def deduccion_telefono(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        correlativo = data.get('correlativo')
        fecha_extravio = data.get('fecha_extravio')

        if not correlativo or not fecha_extravio:
            return JsonResponse({'success': False, 'message': 'Datos incompletos.'})

        telefono_inventario = get_object_or_404(Inventariotelefonos, correlativo=correlativo)
        telefono_asignacion = get_object_or_404(Telefonia, correlativo=correlativo)

        valor_total = telefono_inventario.valortotal
        fecha_asignacion = telefono_asignacion.fecha
        fecha_extravio_dt = datetime.strptime(fecha_extravio, '%Y-%m-%d').date()

        # Calcular los meses de uso hasta la fecha de extravío
        meses_uso = (fecha_extravio_dt.year - fecha_asignacion.year) * 12 + (fecha_extravio_dt.month - fecha_asignacion.month) + 1

        # Definir tasa de depreciación y calcular el valor de depreciación mensual
        tasa_depreciacion_mensual = 5.2632 / 100
        depreciacion_mensual = round(valor_total * tasa_depreciacion_mensual)

        # Calcular el tiempo de pago restante y el valor mensual a pagar
        vida_util_meses = 18
        tiempo_pago_restante = vida_util_meses - meses_uso
        valor_mensual_a_pagar = depreciacion_mensual

        # Actualizar el estado en inventario y telefonía, y valores de deducción
        telefono_inventario.estado = 'DEDUCIDO'
        telefono_inventario.save()

        telefono_asignacion.estado = 'DEDUCIDO'
        telefono_asignacion.fechaextravio = fecha_extravio_dt
        telefono_asignacion.tiempopago = tiempo_pago_restante
        telefono_asignacion.valorquincena = valor_mensual_a_pagar
        telefono_asignacion.save()

        return JsonResponse({
            'success': True,
            'message': 'Deducción calculada y guardada correctamente',
            'valor_mensual_a_pagar': valor_mensual_a_pagar,
            'tiempo_pago_restante': tiempo_pago_restante
        })
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


def exportar_inventario_telefonos(request):
    # Obtener todos los registros de Inventariotelefonos
    datos = (
        Inventariotelefonos.objects
        .values(
            'id', 
            'nombretelefono', 
            'correlativo', 
            'estado', 
            'estadotelefono', 
            'valortotal', 
            'fechacreacion'  # Incluye la fecha de creación
        )
    )

    # Convertir los datos a un DataFrame de pandas
    df = pd.DataFrame(list(datos))

    # Formatear la fecha de creación para que solo muestre la fecha
    df['fechacreacion'] = pd.to_datetime(df['fechacreacion']).dt.date  # Convertir a solo fecha

    # Crear un archivo Excel
    excel_file = "inventario_telefonos.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={excel_file}'

    # Escribir el DataFrame en el archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='InventarioTelefonos')

    return response
#--------ASISTENCIA--------#

def registrar_asistencia_view(request):
    if request.method == "GET":
        jefes = Jefes.objects.filter(estado='ACTIVO')
        meses = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        
        # Generar el rango de años
        años = list(range(2024, 2035))  # Ajusta el rango según tus necesidades
        
        return render(request, 'asistencia/asistenciaregistro.html', {'jefes': jefes, 'meses': meses, 'años': años})

def cargar_colaboradores(request, jefe_id):
    colaboradores = Colaboradores.objects.filter(jefe_id=jefe_id, estado='ACTIVO')
    data = {
        "colaboradores": list(colaboradores.values('id', 'nombrecolaborador')),
    }
    return JsonResponse(data)

def cargar_semanas(request, jefe_id, año, mes):
    colaboradores = Colaboradores.objects.filter(jefe_id=jefe_id, estado='ACTIVO')
    horarios = HorarioJefes.objects.filter(jefe_id=jefe_id, estado='ACTIVO')
    
    semanas = obtener_semanas_del_mes(año, mes)  # Llama a la función que obtiene las semanas del mes
    rangos_semanales = []
    
    for semana in semanas:
        inicio = semana[0]
        fin = semana[1]
        
        # Determina si incluye domingo y sábado
        incluye_domingo = inicio.weekday() <= 6 and fin.weekday() >= 6
        incluye_sabado = inicio.weekday() <= 5 and fin.weekday() >= 5

        rangos_semanales.append({
            "inicio": inicio.strftime("%Y-%m-%d"),
            "fin": fin.strftime("%Y-%m-%d"),
            "incluye_domingo": incluye_domingo,
            "incluye_sabado": incluye_sabado  # Agrega la validación del sábado
        })
    
    # Filtrar colaboradores que no tienen asistencia registrada para el mes y año seleccionados
    colaboradores_sin_asistencia = []
    for colaborador in colaboradores:
        asistencia_existe = Asistencia.objects.filter(
            colaborador=colaborador, mes=mes, año=año
        ).exists()
        
        if not asistencia_existe:
            colaboradores_sin_asistencia.append({
                'id': colaborador.id,
                'nombrecolaborador': colaborador.nombrecolaborador
            })

    data = {
        "colaboradores": colaboradores_sin_asistencia,
        "horarios": list(horarios.values('hora_inicio', 'hora_fin')),
        "semanas": rangos_semanales
    }
    return JsonResponse(data)

def obtener_semanas_del_mes(año, mes):
    semanas = []
    calendario = calendar.monthcalendar(año, mes)
    for semana in calendario:
        dia_inicio = next(dia for dia in semana if dia != 0)
        dia_fin = next((dia for dia in reversed(semana) if dia != 0), dia_inicio)
        fecha_inicio = datetime(año, mes, dia_inicio)
        fecha_fin = datetime(año, mes, dia_fin)
        semanas.append((fecha_inicio, fecha_fin))
    return semanas

def buscar_jefe_por_identidad(request, identidad):
    try:
        jefe = Jefes.objects.get(identidadjefe=identidad)
        data = {
            'success': True,
            'jefe': {
                'id': jefe.id,
                'codigo': jefe.codigo,
                'nombrejefe': jefe.nombrejefe,
            }
        }
    except Jefes.DoesNotExist:
        data = {'success': False, 'message': 'Jefe no encontrado.'}
    return JsonResponse(data)

@csrf_exempt
def registro_asistencia(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            jefe_id = data.get("jefe_id")
            mes = data.get("mes")
            año = data.get("año")
            colaboradores_data = data.get("colaboradores", [])

            jefe = get_object_or_404(Jefes, id=jefe_id)

            for colaborador_data in colaboradores_data:
                colaborador_id = colaborador_data.get("colaborador_id")
                semanas_data = colaborador_data.get("semanas", [])

                if not all(
                    semana.get("horario_semana") and
                    semana.get("almuerzo_inicio") and
                    semana.get("almuerzo_fin")
                    for semana in semanas_data
                ):
                    print(f"Colaborador {colaborador_id} tiene semanas incompletas, omitiendo.")
                    continue

                colaborador = get_object_or_404(Colaboradores, id=colaborador_id)
                asistencia, created = Asistencia.objects.get_or_create(
                    colaborador=colaborador,
                    jefe=jefe,
                    mes=mes,
                    año=año
                )

                for semana_data in semanas_data:
                    semana_inicio = semana_data.get("semana_inicio")
                    semana_fin = semana_data.get("semana_fin")
                    horario_semana = semana_data.get("horario_semana")
                    horario_sabado = semana_data.get("horario_sabado")
                    almuerzo_inicio = semana_data.get("almuerzo_inicio")
                    almuerzo_fin = semana_data.get("almuerzo_fin")
                    dia_libre_domingo = semana_data.get("dia_libre_domingo", False)
                    fecha_domingo_libre = semana_data.get("fecha_domingo_libre")

                    if fecha_domingo_libre == "PAGADO":
                        fecha_domingo_libre = "PAGADO"  # Guardamos "PAGADO" si está marcado
                    elif dia_libre_domingo:
                        fecha_domingo_libre = fecha_domingo_libre  # Guardamos la fecha seleccionada

                    try:
                        semana_inicio = datetime.strptime(semana_inicio, "%Y-%m-%d").date() if semana_inicio else None
                        semana_fin = datetime.strptime(semana_fin, "%Y-%m-%d").date() if semana_fin else None
                    except ValueError as ve:
                        print(f"Error en formato de fecha para colaborador {colaborador_id}: {ve}")
                        return JsonResponse({'success': False, 'message': 'Formato de fecha inválido en semana_inicio o semana_fin.'}, status=400)

                    if not semana_inicio or not semana_fin:
                        print(f"Saltando semana para colaborador {colaborador_id} debido a fechas inválidas.")
                        continue

                    RegistroAsistencia.objects.update_or_create(
                        asistencia=asistencia,
                        semana_inicio=semana_inicio,
                        semana_fin=semana_fin,
                        defaults={
                            'horario_semana': horario_semana,
                            'horario_sabado': horario_sabado,
                            'almuerzo_inicio': almuerzo_inicio,
                            'almuerzo_fin': almuerzo_fin,
                            'dia_libre_domingo': dia_libre_domingo,
                            'fecha_domingo_libre': fecha_domingo_libre
                        }
                    )

            return JsonResponse({'success': True, 'message': 'Asistencia registrada correctamente.'})

        except Exception as e:
            print("Error en registro_asistencia:", e)
            return JsonResponse({'success': False, 'message': 'Ocurrió un error al registrar la asistencia.'}, status=500)

    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)


def exportar_asistencia(request):
    asistencias = Asistencia.objects.prefetch_related('registros', 'colaborador').all()
    
    data = []

    for asistencia in asistencias:
        colaborador = asistencia.colaborador
        colaborador_nombre = colaborador.nombrecolaborador
        codigocolaborador = colaborador.codigocolaborador
        mes = asistencia.mes
        mes_nombre = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                      'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes - 1]
        año = asistencia.año
        
        row = {
            'año': año,
            'mes': mes_nombre,
            'codigocolaborador': codigocolaborador,
            'colaborador': colaborador_nombre,
        }

        last_day = (datetime(año, mes + 1, 1) - timedelta(days=1)).day
        for dia in range(1, last_day + 1):
            fecha_dia = datetime(año, mes, dia)
            row[fecha_dia.strftime('%d/%m/%Y')] = 'AUSENTE'

        for registro in asistencia.registros.all():
            semana_inicio = registro.semana_inicio
            
            for dia in range(7):
                fecha_dia = semana_inicio + timedelta(days=dia)
                if fecha_dia.month == mes and fecha_dia.year == año:
                    if dia < 5:
                        if registro.horario_semana:
                            row[fecha_dia.strftime('%d/%m/%Y')] = registro.horario_semana.split(' - ')[0]
                    elif dia == 5:
                        if registro.horario_sabado:
                            row[fecha_dia.strftime('%d/%m/%Y')] = registro.horario_sabado.split(' - ')[0]
                    elif dia == 6:
                        if registro.dia_libre_domingo:
                            if registro.fecha_domingo_libre:
                                row[fecha_dia.strftime('%d/%m/%Y')] = 'PAGADO'
                            else:
                                row[fecha_dia.strftime('%d/%m/%Y')] = 'LIBRE'

        data.append(row)

    df = pd.DataFrame(data)

    df['mes_num'] = df['mes'].apply(lambda x: ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 
                                                   'Junio', 'Julio', 'Agosto', 'Septiembre', 
                                                   'Octubre', 'Noviembre', 'Diciembre'].index(x) + 1)

    df.sort_values(by=['año', 'mes_num'], inplace=True)
    df.drop(columns=['mes_num'], inplace=True)

    excel_file = "asistencia.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={excel_file}'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Asistencia')

    return response
#---------------------------------------------------------------------------------------------------------------------------------#
def rolesregistrados_view(request):
    if request.method == 'GET':
        return render(request, 'asistencia/rolesregistrados.html')
    
def consultarroles_view(request):
    if request.method == 'GET':
        jefes = Jefes.objects.filter(asistencias__isnull=False).distinct()
        return render(request, 'asistencia/consultaroles.html', {'jefes': jefes})


def buscar_jefe_por_id(request, jefe_id):
    try:
        jefe = Jefes.objects.get(id=jefe_id)
        data = {
            'success': True,
            'jefe': {
                'id': jefe.id,
                'codigo': jefe.codigo,
                'nombrejefe': jefe.nombrejefe,
            }
        }
    except Jefes.DoesNotExist:
        data = {'success': False, 'message': 'Jefe no encontrado.'}
    
    return JsonResponse(data)


def buscar_jefe_por_identidad_registrados(request, identidad):
    try:
        jefe = Jefes.objects.get(identidadjefe=identidad)
        
        # Obtener meses y años de asistencia para este jefe
        meses = Asistencia.objects.filter(jefe=jefe).values('mes').distinct()
        años = Asistencia.objects.filter(jefe=jefe).values('año').distinct()
        
        meses_nombres = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }

        meses_disponibles = {mes['mes']: meses_nombres[mes['mes']] for mes in meses if mes['mes'] in meses_nombres}
        años_disponibles = [año['año'] for año in años]

        data = {
            'success': True,
            'jefe': {
                'id': jefe.id,
                'codigo': jefe.codigo,
                'nombrejefe': jefe.nombrejefe,
            },
            'meses': meses_disponibles,
            'años': años_disponibles
        }
    except Jefes.DoesNotExist:
        data = {'success': False, 'message': 'Jefe no encontrado.'}
    return JsonResponse(data)

def cargar_roles_registrados(request, jefe_id, mes, año):
    registros = RegistroAsistencia.objects.filter(
        asistencia__jefe_id=jefe_id,
        asistencia__mes=mes,
        asistencia__año=año
    ).select_related('asistencia', 'asistencia__colaborador')

    data = []
    for registro in registros:
        data.append({
            'nombre_colaborador': registro.asistencia.colaborador.nombrecolaborador,
            'semana_inicio': registro.semana_inicio,
            'semana_fin': registro.semana_fin,
            'horario_semana': registro.horario_semana,
            'horario_sabado': registro.horario_sabado,
            'almuerzo_inicio': registro.almuerzo_inicio,
            'almuerzo_fin': registro.almuerzo_fin,
            'dia_libre_domingo': registro.dia_libre_domingo,
            'fecha_domingo_libre': registro.fecha_domingo_libre,
        })

    return JsonResponse({'success': True, 'data': data})
#---------------------------------------------------------------------------------------------------------------------------------#

def jefes_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')  
        jefes = Jefes.objects.all()

        if search:
            jefes = jefes.filter(
                models.Q(codigo__icontains=search) |
                models.Q(nombrejefe__icontains=search)
            )

        paginator = Paginator(jefes, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        context = {
            'jefes': page_obj,
            'search': search,
            'page_range': page_range,
        }

        return render(request, 'modulos/jefes.html', context)
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            codigo = data.get('codigo')
            nombrejefe = data.get('nombrejefe')
            estado = data.get('estado')
            identidadjefe = data.get('identidadjefe0')

            if not nombrejefe:
                return JsonResponse({'success': False, 'message': 'El nombre del jefe es obligatorio.'}, status=400)

            if Jefes.objects.filter(codigo=codigo).exists():
                return JsonResponse({'success': False, 'message': 'El Jefe ya está registrado.'}, status=409)

            jefes = Jefes.objects.create(
                codigo=codigo,
                nombrejefe=nombrejefe,
                estado=estado,
                identidadjefe=identidadjefe 
            )

            return JsonResponse({'success': True, 'message': 'Jefe registrado correctamente.'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Formato de JSON no válido.'}, status=400)
        except Exception as e:
            # Asegúrate de obtener más detalles del error para poder depurarlo
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

        
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            codigo = data.get('codigo')
            nombrejefe = data.get('nombrejefe')
            estado = data.get('estado')
            identidadjefe = data.get('identidadjefe')

            jefe = Jefes.objects.get(id=id)

            jefe.codigo = codigo
            jefe.nombrejefe = nombrejefe
            jefe.estado = estado
            jefe.identidadjefe = identidadjefe
            jefe.save()

            return JsonResponse({'success': True, 'message': 'Jefe actualizado correctamente.'}, status=200)
        except Jefes.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Jefe no encontrado.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def colaboradores_view(request, id=None):
    if request.method == 'GET':
        search = request.GET.get('search', '')  # Parámetro de búsqueda
        estado = request.GET.get('estado', '')  # Parámetro de filtro por estado

        # Obtener todos los colaboradores
        colaboradores = Colaboradores.objects.all()

        # Filtrar por búsqueda en múltiples campos
        if search:
            colaboradores = colaboradores.filter(
                models.Q(nombrecolaborador__icontains=search) |
                models.Q(sucursal__nombre_sucursal__icontains=search) |
                models.Q(empresa__nombre_empresa__icontains=search) |
                models.Q(unidad_de_negocio__nombre_unidad_de_negocio__icontains=search) |
                models.Q(departamento__nombre_departamento__icontains=search) |
                models.Q(jefe__nombrejefe__icontains=search) |
                models.Q(codigocolaborador__icontains=search)
            )

        # Filtrar por estado
        if estado:
            colaboradores = colaboradores.filter(estado=estado)

        # Paginación
        paginator = Paginator(colaboradores, 10)  # 10 colaboradores por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calcular el rango de páginas a mostrar
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Obtener listas para los selects del formulario
        all_sucursales = Sucursal.objects.filter(estado='ACTIVO').order_by('nombre_sucursal').distinct()
        all_empresas = Empresas.objects.filter(estado='ACTIVO').order_by('nombre_empresa').distinct()
        all_unidades_negocio = Unidad_Negocio.objects.filter(estado='ACTIVO').order_by('nombre_unidad_de_negocio').distinct()
        all_departamentos = Departamento.objects.filter(estado='ACTIVO').order_by('nombre_departamento').distinct()
        all_jefes = Jefes.objects.filter(estado="ACTIVO").order_by('nombrejefe').distinct()

        context = {
            'colaboradores': page_obj,
            'search': search,
            'estado': estado,
            'page_range': page_range,
            'all_sucursales': all_sucursales,
            'all_empresas': all_empresas,
            'all_unidades_negocio': all_unidades_negocio,
            'all_departamentos': all_departamentos,
            'all_jefes': all_jefes
        }

        return render(request, 'modulos/colaboradores.html', context)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            nombrecolaborador = data.get('nombrecolaborador')
            sucursal_id = data.get('sucursal')
            empresa_id = data.get('empresa')
            unidadnegocio_id = data.get('unidadnegocio')
            departamento_id = data.get('departamento')
            jefe_id = data.get('jefes')
            estado = data.get('estado')
            codigocolaborador = data.get('codigocolaborador')

            # Manejar los valores en blanco o None
            sucursal = Sucursal.objects.get(id=sucursal_id) if sucursal_id else None
            empresa = Empresas.objects.get(id=empresa_id) if empresa_id else None
            unidad_negocio = Unidad_Negocio.objects.get(id=unidadnegocio_id) if unidadnegocio_id else None
            departamento = Departamento.objects.get(id=departamento_id) if departamento_id else None
            jefe = Jefes.objects.get(id=jefe_id) if jefe_id else None

            # Crear el nuevo colaborador
            colaborador = Colaboradores.objects.create(
                nombrecolaborador=nombrecolaborador,
                sucursal=sucursal,
                empresa=empresa,
                unidad_de_negocio=unidad_negocio,
                departamento=departamento,
                jefe=jefe,
                estado=estado,
                codigocolaborador=codigocolaborador
            )

            return JsonResponse({'success': True, 'message': 'Colaborador registrado correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

        
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            colaborador = Colaboradores.objects.get(id=id)

            colaborador.nombrecolaborador = data.get('nombrecolaborador', colaborador.nombrecolaborador) or None
            colaborador.codigocolaborador = data.get('codigocolaborador', colaborador.codigocolaborador) or None

            sucursal_id = data.get('sucursal_id')
            colaborador.sucursal = Sucursal.objects.get(id=sucursal_id) if sucursal_id else None

            empresa_id = data.get('empresa_id')
            colaborador.empresa = Empresas.objects.get(id=empresa_id) if empresa_id else None

            unidad_negocio_id = data.get('unidad_de_negocio_id')
            colaborador.unidad_de_negocio = Unidad_Negocio.objects.get(id=unidad_negocio_id) if unidad_negocio_id else None

            departamento_id = data.get('departamento_id')
            colaborador.departamento = Departamento.objects.get(id=departamento_id) if departamento_id else None

            jefe_id = data.get('jefe_id')
            colaborador.jefe = Jefes.objects.get(id=jefe_id) if jefe_id else None

            colaborador.estado = data.get('estado', colaborador.estado) or None

            colaborador.save()

            return JsonResponse({'success': True, 'message': 'Colaborador actualizado correctamente'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def horariosjefes_view(request, id=None):
    if request.method == "GET":
        search = request.GET.get('search', '')
        estado = request.GET.get('estado', '')

        # Filtrado inicial de los horarios
        horarios = HorarioJefes.objects.all()

        # Realizar el filtrado en los campos del modelo relacionado Jefes
        if search:
            horarios = horarios.filter(
                models.Q(jefe__nombrejefe__icontains=search) |
                models.Q(jefe__codigo__icontains=search)
            )
        
        # Filtrar por estado si se proporciona
        if estado:
            horarios = horarios.filter(estado=estado)

        # Paginación
        paginator = Paginator(horarios, 10)  # 10 items por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Cálculo del rango de páginas
        total_pages = paginator.num_pages
        current_page = page_obj.number
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)

        # Ajuste del rango para que siempre muestre hasta 5 páginas
        if current_page <= 3:
            end_page = min(5, total_pages)
        elif current_page >= total_pages - 2:
            start_page = max(1, total_pages - 4)

        page_range = range(start_page, end_page + 1)

        # Datos adicionales
        all_jefes = Jefes.objects.filter(estado="ACTIVO").order_by('nombrejefe').distinct()

        context = {
            'horarios': page_obj,
            'search': search,
            'estado': estado,
            'page_range': page_range,
            'all_jefes': all_jefes
        }
        return render(request, 'Modulos/horariosjefes.html', context)


    elif request.method == "POST":
        try:
            data = json.loads(request.body.decode('utf-8'))

            horarios = HorarioJefes(
                jefe=Jefes.objects.get(id=data['jefe']),
                hora_inicio=data['hora_inicio'],
                hora_fin=data['hora_fin'],
                estado=data['estado']
            )
            horarios.save()

            return JsonResponse({'success': True, 'message': 'Horario registrado correctamente'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    elif request.method == "PUT":
        if not id:
            return JsonResponse({'success': False, 'message': 'ID del horario no proporcionado'}, status=400)

        try:
            data = json.loads(request.body.decode('utf-8'))
            horario = get_object_or_404(HorarioJefes, id=id)

            horario.jefe = Jefes.objects.get(id=data['jefe'])
            horario.hora_inicio = data['hora_inicio']
            horario.hora_fin = data['hora_fin']
            horario.estado = data['estado']
            horario.save()

            return JsonResponse({'success': True, 'message': 'Horario actualizado correctamente'})

        except Jefes.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'El jefe seleccionado no existe'}, status=404)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
